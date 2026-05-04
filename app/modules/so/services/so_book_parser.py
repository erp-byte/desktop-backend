"""
Parser for Sales Order Book Excel files (e.g. "Sales Order Book Jan-March 21 CFPL.xlsx").

Uses a state machine to group line items under their parent SO header.
GST is at the SO header level and gets apportioned to lines by value.

Column mapping is detected dynamically from the header rows so the parser
works even when the Excel has fewer or reordered columns.
"""

import io
import logging
import re
from datetime import date, datetime

import openpyxl

from app.core.helpers import safe_float_zero as safe_float, safe_str

logger = logging.getLogger(__name__)

# Locations to strip from customer names
_LOCATIONS = (
    "maharashtra", "mah", "pune", "nashik", "mumbai", "navi mumbai", "thane", "nagpur",
    "telangana", "hyderabad", "karnataka", "bengaluru", "bangalore", "chennai", "tamil nadu",
    "ahmedabad", "gujarat", "rajasthan", "delhi", "ncr", "punjab", "haryana", "jharkhand",
    "kerala", "trichy", "andhra pradesh", "andra pradesh", "khurdha", "turbhe", "bhiwandi",
)
_LOCATION_PATTERN = re.compile(
    r"[\s,\-]*\b(?:" + "|".join(re.escape(loc) for loc in _LOCATIONS) + r")\b[\s,\-]*$",
    re.IGNORECASE,
)
_PAREN_PATTERN = re.compile(r"\s*\(.*?\)")

# ---------------------------------------------------------------------------
# Known header aliases → canonical field name
# Each key is a lowercase substring that can appear in a column header.
# Order matters: first match wins, so put more specific patterns first.
# ---------------------------------------------------------------------------
_HEADER_ALIASES: list[tuple[str, str]] = [
    # Date
    ("date", "date"),
    # Customer / party
    ("party", "customer_name"),
    ("customer", "customer_name"),
    ("buyer", "customer_name"),
    # Voucher type
    ("voucher type", "voucher_type"),
    ("vch type", "voucher_type"),
    ("type", "voucher_type"),
    # SO number
    ("vch no", "so_number"),
    ("voucher no", "so_number"),
    ("so no", "so_number"),
    ("order no", "so_number"),
    # Order reference
    ("order ref", "order_reference"),
    ("ref no", "order_reference"),
    # Narration
    ("narration", "narration"),
    # Payment terms
    ("payment", "payment_terms"),
    # Other references
    ("other ref", "other_references"),
    # Terms of delivery
    ("terms of delivery", "terms_of_delivery"),
    ("delivery terms", "terms_of_delivery"),
    # Quantity
    ("quantity", "quantity"),
    ("qty", "quantity"),
    # Alt units
    ("alt", "alt_units"),
    # Rate
    ("rate", "rate_inr"),
    # Particulars / SKU name (item name in line rows)
    ("particulars", "particulars"),
    ("item name", "particulars"),
    ("sku", "particulars"),
    # Amount / value
    ("amount", "amount_inr"),
    ("value", "amount_inr"),
    # Gross total
    ("gross total", "gross_total"),
    # Sales GST local
    ("sales gst local", "sales_gst_local"),
    ("sales.*local", "sales_gst_local"),
    # CGST
    ("cgst", "cgst_amount"),
    # SGST
    ("sgst", "sgst_amount"),
    # IGST
    ("igst", "igst_amount"),
    # Packing charges
    ("packing", "packing_charges"),
    # Round off
    ("round", "round_off"),
    # Freight
    ("freight", "freight_charges"),
    # Sales export
    ("export", "sales_export"),
]

# Fallback: fixed positional mapping (original Tally SO Book layout)
_DEFAULT_COL_MAP = {
    0: "date", 1: "particulars", 2: "voucher_type", 3: "so_number",
    4: "order_reference", 5: "narration", 6: "payment_terms",
    7: "other_references", 8: "terms_of_delivery",
    9: "quantity", 10: "alt_units", 11: "rate_inr", 12: "amount_inr",
    13: "gross_total", 14: "sales_gst_local",
    15: "cgst_amount", 16: "sgst_amount", 17: "igst_amount",
    18: "packing_charges", 19: "round_off", 20: "freight_charges",
    21: "sales_export",
}


def _parse_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    return s if s else None


def _clean_customer_name(name: str) -> str:
    """Derive common_customer_name from full customer name."""
    cleaned = _PAREN_PATTERN.sub("", name)
    cleaned = _LOCATION_PATTERN.sub("", cleaned)
    cleaned = re.sub(r"[\s,\-]+$", "", cleaned).strip()
    return cleaned[:50]


# ---------------------------------------------------------------------------
# Dynamic header detection
# ---------------------------------------------------------------------------

def _detect_columns(ws) -> tuple[dict[str, int], int]:
    """Scan the first 15 rows to find a header row and build field→col index map.

    Returns (col_map, data_start_row).
    col_map: { canonical_field_name: column_index }
    data_start_row: 1-indexed row where data begins (row after the header).
    """
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        cells = [safe_str(c).lower() if c is not None else "" for c in row]
        text = " ".join(cells)

        # A header row should contain at least a few key indicators
        indicators = sum(1 for kw in ("date", "party", "particulars", "quantity", "amount",
                                       "voucher", "customer", "vch no", "so no", "rate")
                         if kw in text)
        if indicators < 3:
            continue

        # Build col_map from detected header
        col_map: dict[str, int] = {}
        used_fields: set[str] = set()
        for col_idx, cell_text in enumerate(cells):
            if not cell_text:
                continue
            for alias, field in _HEADER_ALIASES:
                if field in used_fields:
                    continue
                if re.search(alias, cell_text):
                    col_map[field] = col_idx
                    used_fields.add(field)
                    break

        if len(col_map) >= 3:
            logger.info("Detected SO Book headers at row %d: %s", row_idx, col_map)
            return col_map, row_idx + 1

    # Fallback: use default positional mapping, data starts at row 13
    logger.warning("Could not detect SO Book headers, using default column layout")
    col_map = {field: idx for idx, field in _DEFAULT_COL_MAP.items()}
    return col_map, 13


def _get(row: tuple, col_map: dict[str, int], field: str, default=None):
    """Safe field access from a row using the dynamic column map."""
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return default
    return row[idx]


# ---------------------------------------------------------------------------
# Row classification
# ---------------------------------------------------------------------------

def _is_header_row(row: tuple, col_map: dict[str, int]) -> bool:
    """Header rows have a date value in the date column."""
    date_val = _get(row, col_map, "date")
    return date_val is not None and str(date_val).strip() != ""


def _is_line_row(row: tuple, col_map: dict[str, int]) -> bool:
    """Line rows have no date but have a quantity."""
    date_val = _get(row, col_map, "date")
    if date_val is not None and str(date_val).strip() != "":
        return False
    qty = _get(row, col_map, "quantity")
    return qty is not None and str(qty).strip() != ""


def _is_grand_total_row(row: tuple, col_map: dict[str, int]) -> bool:
    particulars = safe_str(_get(row, col_map, "particulars"))
    if not particulars:
        # Also check col index 1 as fallback
        val = row[1] if len(row) > 1 else None
        particulars = safe_str(val) if val else None
    return particulars is not None and particulars.lower().startswith("grand total")


# ---------------------------------------------------------------------------
# Main parser
# ---------------------------------------------------------------------------

def parse_so_book(file_bytes: bytes) -> list[dict]:
    """
    Parse a Sales Order Book Excel file.
    Returns a list of SO dicts, each with a 'lines' array.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    col_map, data_start = _detect_columns(ws)

    current_so = None
    all_orders = []
    warnings = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start, values_only=True), start=data_start):
        row_vals = tuple(row)

        # Skip grand total row
        if _is_grand_total_row(row_vals, col_map):
            continue

        if _is_header_row(row_vals, col_map):
            # Flush previous SO
            if current_so is not None:
                all_orders.append(current_so)

            customer_name = safe_str(_get(row_vals, col_map, "particulars")) or \
                            safe_str(_get(row_vals, col_map, "customer_name")) or ""
            common_customer_name = _clean_customer_name(customer_name)
            igst = safe_float(_get(row_vals, col_map, "igst_amount", 0))

            current_so = {
                "so_number": safe_str(_get(row_vals, col_map, "so_number")),
                "so_date": _parse_date(_get(row_vals, col_map, "date")),
                "customer_name": customer_name,
                "common_customer_name": common_customer_name,
                "company": "CFPL",
                "voucher_type": safe_str(_get(row_vals, col_map, "voucher_type")),
                "payment_terms": safe_str(_get(row_vals, col_map, "payment_terms")),
                "order_reference": safe_str(_get(row_vals, col_map, "order_reference")),
                "narration": safe_str(_get(row_vals, col_map, "narration")),
                "other_references": safe_str(_get(row_vals, col_map, "other_references")),
                "terms_of_delivery": safe_str(_get(row_vals, col_map, "terms_of_delivery")),
                "is_interstate": igst > 0,
                "gross_total": safe_float(_get(row_vals, col_map, "gross_total", 0)),
                "sales_gst_local": safe_float(_get(row_vals, col_map, "sales_gst_local", 0)),
                "cgst_amount": safe_float(_get(row_vals, col_map, "cgst_amount", 0)),
                "sgst_amount": safe_float(_get(row_vals, col_map, "sgst_amount", 0)),
                "igst_amount": igst,
                "packing_charges": safe_float(_get(row_vals, col_map, "packing_charges", 0)),
                "round_off": safe_float(_get(row_vals, col_map, "round_off", 0)),
                "freight_charges": safe_float(_get(row_vals, col_map, "freight_charges", 0)),
                "sales_export": safe_float(_get(row_vals, col_map, "sales_export", 0)),
                "lines": [],
            }

        elif _is_line_row(row_vals, col_map):
            if current_so is None:
                warnings.append(f"Orphan line at row {row_idx}: {safe_str(_get(row_vals, col_map, 'particulars'))}")
                continue

            alt_raw = _get(row_vals, col_map, "alt_units")
            alt_units = safe_float(alt_raw) if alt_raw is not None and str(alt_raw).strip() != "" else None

            current_so["lines"].append({
                "sku_name": safe_str(_get(row_vals, col_map, "particulars")),
                "quantity": safe_float(_get(row_vals, col_map, "quantity", 0)),
                "alt_units": alt_units,
                "uom": "CTN" if alt_units is not None else "PCS",
                "rate_inr": safe_float(_get(row_vals, col_map, "rate_inr", 0)),
                "amount_inr": safe_float(_get(row_vals, col_map, "amount_inr", 0)),
            })

    # Flush last SO
    if current_so is not None:
        all_orders.append(current_so)

    wb.close()

    # Apportion header-level charges to lines by value ratio and add line numbers.
    # Bug 10: packing_charges and freight_charges exist at SO header level but were
    # zeroed at line level.  Apportion them the same way as GST so the per-line
    # total_amount_inr and GST reconciliation total check are correct.
    for so in all_orders:
        lines = so["lines"]
        sum_values = sum(l["amount_inr"] for l in lines)

        for i, line in enumerate(lines, start=1):
            line["line_number"] = i
            if sum_values > 0:
                ratio = line["amount_inr"] / sum_values
                line["igst_amount"] = round(so["igst_amount"] * ratio, 3)
                line["sgst_amount"] = round(so["sgst_amount"] * ratio, 3)
                line["cgst_amount"] = round(so["cgst_amount"] * ratio, 3)
                line["packing_amount"] = round(so["packing_charges"] * ratio, 3)
                line["freight_amount"] = round(so["freight_charges"] * ratio, 3)
            else:
                line["igst_amount"] = 0.0
                line["sgst_amount"] = 0.0
                line["cgst_amount"] = 0.0
                line["packing_amount"] = 0.0
                line["freight_amount"] = 0.0
            line["total_amount_inr"] = round(
                line["amount_inr"]
                + line["igst_amount"] + line["sgst_amount"] + line["cgst_amount"]
                + line["packing_amount"] + line["freight_amount"],
                3,
            )

        so["total_line_items"] = len(lines)

    # Validation
    for so in all_orders:
        sn = so["so_number"]
        if len(so["lines"]) == 0:
            warnings.append(f"SO {sn} has no line items")
        if not so["gross_total"]:
            warnings.append(f"SO {sn} has zero gross total")
        line_sum = sum(l["amount_inr"] for l in so["lines"])
        if so["sales_gst_local"] and abs(line_sum - so["sales_gst_local"]) > 1:
            warnings.append(f"SO {sn}: line value sum ({line_sum:.2f}) != sales_gst_local ({so['sales_gst_local']:.2f})")
        if so["is_interstate"] and so["sgst_amount"] > 0:
            warnings.append(f"SO {sn}: GST type conflict — interstate but SGST > 0")

    for w in warnings:
        logger.warning("SO Book: %s", w)

    logger.info(
        "Parsed SO Book: %d SOs, %d total lines, columns detected: %s",
        len(all_orders), sum(len(so["lines"]) for so in all_orders),
        list(col_map.keys()),
    )

    return all_orders
