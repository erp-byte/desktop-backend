"""Sales Register Excel parser — converts Excel bytes into grouped row dicts."""

import io
import logging
from collections import defaultdict

import openpyxl

from app.core.helpers import safe_float, safe_str

logger = logging.getLogger(__name__)

# Column header → internal key mapping
HEADER_MAP = {
    "Date": "date",
    "Customer Name": "customer_name",
    "Common Customer Name": "common_customer_name",
    "Article": "article",
    "Main GRP": "main_grp",
    "Sub-Group": "sub_group",
    "Voucher Type": "voucher_type",
    "Company": "company",
    "Sales Order No.": "so_number",
    "Qty.": "qty",
    "Rate": "rate",
    "Without GST Amt.": "without_gst_amt",
    "IGST": "igst",
    "SGST": "sgst",
    "CGST": "cgst",
    "APMC": "apmc",
    "Packing": "packing",
    "Freight": "freight",
    "Processing": "processing",
    "With GST Amt.": "with_gst_amt",
}

# We need two UOM columns and GRP — handle by position after header detection
POSITIONAL_FIELDS = {"UOM_H": "uom_h", "GRP_I": "grp_i"}


def parse_sales_register(file_bytes: bytes) -> dict[str, list[dict]]:
    """
    Parse Sales Register Excel from bytes.
    Groups rows by Sales Order No.
    Returns { so_number: [row_dicts] }.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    # Find header row — scan first 5 rows for "Sales Order No."
    header_row_idx = None
    headers = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=False), start=1):
        for cell in row:
            val = safe_str(cell.value)
            if val == "Sales Order No.":
                header_row_idx = row_idx
                break
        if header_row_idx:
            # Map all headers in this row
            for cell in row:
                val = safe_str(cell.value)
                if val and val in HEADER_MAP:
                    headers[cell.column - 1] = HEADER_MAP[val]
            break

    if header_row_idx is None:
        wb.close()
        raise ValueError("Could not find 'Sales Order No.' header in Excel file")

    # Detect UOM (col H = index 7) and GRP (col I = index 8) by position
    # These have duplicate header names so we use position
    headers[7] = "uom"  # Column H (0-indexed = 7)
    headers[8] = "grp_code"  # Column I (0-indexed = 8)

    # Parse data rows
    groups: dict[str, list[dict]] = defaultdict(list)
    row_count = 0

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=False):
        values = {headers[i]: cell.value for i, cell in enumerate(row) if i in headers}

        so_number = safe_str(values.get("so_number"))
        if not so_number:
            continue

        row_data = {
            "so_number": so_number,
            "date": values.get("date"),
            "customer_name": safe_str(values.get("customer_name")),
            "common_customer_name": safe_str(values.get("common_customer_name")),
            "article": safe_str(values.get("article")),
            "item_category": safe_str(values.get("main_grp")),
            "sub_category": safe_str(values.get("sub_group")),
            "uom": safe_float(values.get("uom")),
            "grp_code": safe_str(values.get("grp_code")),
            "voucher_type": safe_str(values.get("voucher_type")),
            "company": safe_str(values.get("company")),
            "quantity": safe_float(values.get("qty")),
            "rate_inr": safe_float(values.get("rate")),
            "amount_inr": safe_float(values.get("without_gst_amt")),
            "igst_amount": safe_float(values.get("igst")) or 0,
            "sgst_amount": safe_float(values.get("sgst")) or 0,
            "cgst_amount": safe_float(values.get("cgst")) or 0,
            "apmc_amount": safe_float(values.get("apmc")) or 0,
            "packing_amount": safe_float(values.get("packing")) or 0,
            "freight_amount": safe_float(values.get("freight")) or 0,
            "processing_amount": safe_float(values.get("processing")) or 0,
            "total_amount_inr": safe_float(values.get("with_gst_amt")),
        }
        groups[so_number].append(row_data)
        row_count += 1

    wb.close()
    logger.info("Parsed %d rows into %d SOs from Excel", row_count, len(groups))
    return dict(groups)
