"""Master data ingest — BOM headers, BOM lines, process routes, machines.

Reads three Excel files at startup and populates:
  - bom_header       (from FG_Master_Completion.xlsx)
  - bom_process_route (derived from Process Category column)
  - bom_line          (from BOM_Enrichment.xlsx)
  - machine           (from Floorwise utility dada.xlsx)

Idempotent: skips if bom_header already has data.
"""

import csv
import logging
import re
from pathlib import Path

import openpyxl

from app.core.helpers import (
    PROCESS_CLASS_LOCK,
    ROUTING_LOCK,
    SFG_CODE_LOCK,
    insert_with_pk_retry,
    new_short_time_id,
    normalise_key,
)
from app.modules.so.services.item_matcher import MasterItem, match_sku

logger = logging.getLogger(__name__)

# Plug CSVs live under the repo's docs/ tree, NOT under app/data — resolve them
# by ABSOLUTE path off the repo root so a Workspace-subfolder CWD can't break
# the lookup. master_ingest.py is at app/modules/production/services/, so
# parents[4] is the server_replica root.
_REPO_ROOT = Path(__file__).resolve().parents[4]
_SFG_PLUG_DIR = _REPO_ROOT / "docs" / "flows" / "Candor_SFG_JobCard_Integration_2026-06-14"
SFG_CATALOG_PLUG = _SFG_PLUG_DIR / "SFG_Catalog_Plug.csv"
JC_ROUTING_PLUG = _SFG_PLUG_DIR / "JC_Routing_Plug.csv"
JC_BOM_PLUG = _SFG_PLUG_DIR / "JC_BOM_Plug.csv"
# 056 enrichment / reconciliation plug files (companion set, same 2026-06-17 gen).
JC_STAGE1_MASTER = _SFG_PLUG_DIR / "JC_Stage1_Create_WIP_Master.csv"
PROCESS_CATEGORY_MAP = _SFG_PLUG_DIR / "ProcessCategory_to_Operation.csv"
JC_STAGE2_PACKING = _SFG_PLUG_DIR / "JC_Stage2_Final_FG_Packing.csv"
SFG_RESOLUTION_MAP = _SFG_PLUG_DIR / "SFG_Resolution_Map.csv"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _safe_float(val) -> float | None:
    if val is None:
        return None
    try:
        return round(float(val), 3)
    except (ValueError, TypeError):
        return None


def _safe_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _split_comma(val) -> list[str]:
    """Split comma-separated string, strip each part, remove empties."""
    if not val:
        return []
    return [p.strip() for p in str(val).split(",") if p.strip()]


def _split_process_category(val) -> list[str]:
    """Split process category by '+', handling unicode thin spaces and parens."""
    if not val:
        return []
    s = str(val)
    # Normalize unicode thin/non-breaking spaces around +
    s = re.sub(r'[\u2009\u00a0\s]*\+[\u2009\u00a0\s]*', '+', s)
    parts = [p.strip().strip('()') for p in s.split('+')]
    return [p for p in parts if p]


def _derive_entity(factory: str | None) -> str | None:
    if not factory:
        return None
    f = str(factory).strip().upper()
    if 'W202' in f:
        return 'cfpl'
    if 'A185' in f:
        return 'cdpl'
    return None


# ---------------------------------------------------------------------------
# 1. FG Master → bom_header + bom_process_route
# ---------------------------------------------------------------------------

async def ingest_fg_master(conn, file_path: Path, master_items: list[MasterItem]) -> dict:
    """Read FG_Master_Fill sheet → bom_header + bom_process_route. Returns header_map."""

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb['FG_Master_Fill']

    header_map: dict[tuple[str, str | None], int] = {}  # (fg_name, variant) → bom_id
    headers_created = 0
    routes_created = 0
    skipped = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # Row 0 = header, Row 1 = template row ("Particluars")
        if i < 2:
            continue

        vals = list(row)
        while len(vals) < 16:
            vals.append(None)

        fg_name = _safe_str(vals[1])
        if not fg_name or fg_name == 'Particluars':
            skipped += 1
            continue

        group = _safe_str(vals[2])
        sub_group = _safe_str(vals[3])
        process_category = _safe_str(vals[4])
        business_unit = _safe_str(vals[5])
        factory = _safe_str(vals[6])
        floors = _split_comma(vals[7])
        machines = _split_comma(vals[8])
        pack_size_kg = _safe_float(vals[9])
        shelf_life_days = _safe_int(vals[10])
        gst_rate = _safe_float(vals[11])
        hsn_sac = _safe_str(vals[12])
        inventory_group = _safe_str(vals[13])
        customer_code = _safe_str(vals[14])

        entity = _derive_entity(factory)

        # Use Excel Group as item_group; fallback to fuzzy match
        item_group = group
        if not item_group and master_items:
            matched, _score = match_sku(fg_name, master_items)
            if matched:
                item_group = matched.group

        bom_id = await conn.fetchval(
            """
            INSERT INTO bom_header (
                fg_sku_name, customer_name, pack_size_kg, item_group,
                entity, sub_group, process_category, business_unit,
                factory, floors, machines, shelf_life_days,
                gst_rate, hsn_sac, inventory_group, customer_code
            ) VALUES ($1, NULL, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13, $14, $15)
            ON CONFLICT DO NOTHING
            RETURNING bom_id
            """,
            fg_name, pack_size_kg, item_group,
            entity, sub_group, process_category, business_unit,
            factory, floors or None, machines or None, shelf_life_days,
            gst_rate, hsn_sac, inventory_group, customer_code,
        )

        if bom_id is None:
            # Already exists — fetch existing id
            bom_id = await conn.fetchval(
                "SELECT bom_id FROM bom_header WHERE fg_sku_name = $1 AND customer_name IS NULL LIMIT 1",
                fg_name,
            )
            if bom_id is None:
                skipped += 1
                continue
        else:
            headers_created += 1

        header_map[(fg_name, None)] = bom_id

        # Process route from Process Category (split by +)
        steps = _split_process_category(process_category)
        for step_num, step_name in enumerate(steps, 1):
            stage = step_name.lower().replace(' ', '_')
            await conn.execute(
                """
                INSERT INTO bom_process_route (bom_id, step_number, process_name, stage)
                VALUES ($1, $2, $3, $4)
                ON CONFLICT (bom_id, step_number) DO NOTHING
                """,
                bom_id, step_num, step_name, stage,
            )
            routes_created += 1

    wb.close()
    logger.info("FG Master: %d headers, %d routes, %d skipped", headers_created, routes_created, skipped)
    return header_map


# ---------------------------------------------------------------------------
# 1b. SFG catalogue → all_sku (item_type='sfg')
# ---------------------------------------------------------------------------

async def _all_sku_has_sfg_code(conn) -> bool:
    """True if the all_sku.sfg_code column (added by 051) exists."""
    return bool(await conn.fetchval(
        """
        SELECT EXISTS (
            SELECT 1 FROM information_schema.columns
             WHERE table_name = 'all_sku' AND column_name = 'sfg_code'
        )
        """
    ))


async def ingest_sfg_master(conn, file_path: Path | None = None) -> dict:
    """Load the canonical SFG items (SFG_Catalog_Plug.csv) into all_sku as
    item_type='sfg'.

    db/051_sfg_seed_catalog.sql prepares the schema (widens all_sku.sku_id to
    BIGINT, adds sfg_code + its unique index); this loader does the DATA:
      * re-types an existing bulk article matched by (normalised) name in place,
        claiming its SFG#### and PRESERVING its sku_id / real item_group / uom / gst;
      * inserts the derived net-new SFGs with an application-supplied 8-digit
        BIGINT primary key via new_short_time_id() + insert_with_pk_retry()
        (sale_group='wip', uom left NULL — the plug's 'kg' is conceptual and
        all_sku.uom is NUMERIC).

    MUST run inside an outer transaction (insert_with_pk_retry uses savepoints).
    Idempotent (keyed on sfg_code). Safe in a half-migrated DB: skips with a
    warning if the all_sku.sfg_code column (051) is absent.
    """
    path = Path(file_path) if file_path else SFG_CATALOG_PLUG
    if not path.exists():
        logger.warning("SFG catalogue ingest skipped — plug not found: %s", path)
        return {"status": "file_not_found", "inserted": 0, "retyped": 0}

    if not await _all_sku_has_sfg_code(conn):
        logger.warning(
            "SFG catalogue ingest skipped — all_sku.sfg_code missing; "
            "run scripts/migrate.py (051_sfg_seed_catalog.sql) first"
        )
        return {"status": "schema_missing", "inserted": 0, "retyped": 0}

    # Serialise concurrent loaders (cold-start storm of multiple workers, or a
    # migrate path overlapping app startup) on the same advisory lock next_sfg_code
    # uses. Transaction-scoped — auto-released at COMMIT/ROLLBACK. Combined with
    # the per-row ON CONFLICT / NOT EXISTS guards this makes the whole load
    # race-safe (no UniqueViolation can abort the shared master-ingest txn).
    await conn.execute("SELECT pg_advisory_xact_lock($1)", SFG_CODE_LOCK)

    # Read + normalise; dedupe defensively by code (last wins). Track how many
    # rows the plug itself expects to be re-types (already_in_all_sku='Y') so we
    # can warn on flag drift (e.g. SFG0334 is mis-flagged 'N' but pre-exists).
    with path.open(encoding="utf-8-sig", newline="") as fh:
        seed: dict[str, tuple[str, str | None]] = {}
        retype_expected = 0
        for r in csv.DictReader(fh):
            code = normalise_key(r.get("sfg_code"))
            name = normalise_key(r.get("particulars"))
            if not code or not name:
                continue
            group = normalise_key(r.get("item_group")) or None
            seed[code] = (name, group)
            if normalise_key(r.get("already_in_all_sku")).upper() == "Y":
                retype_expected += 1

    # Idempotency guard keyed on the CANONICAL codes (not a raw item_type='sfg'
    # count): a foreign/partial 'sfg' population must not mask a missing plug.
    codes = list(seed.keys())
    present = await conn.fetchval(
        "SELECT count(*) FROM all_sku WHERE sfg_code = ANY($1::text[])", codes
    )
    if present and present >= len(codes):
        logger.info(
            "SFG catalogue already ingested (%d/%d canonical codes present) — skipping",
            present, len(codes),
        )
        return {"status": "already_ingested", "inserted": 0, "retyped": 0,
                "sfg_rows": present}

    # Build a NORMALISED-name index of claimable rows (sfg_code IS NULL) so an
    # NBSP/mojibake-bearing all_sku name still matches the clean plug name — the
    # join is normalised on BOTH sides (normalise_key is Python-only, so the
    # match is done here rather than in SQL). Lowest sku_id wins on a name clash.
    claimable: dict[str, list[tuple]] = {}
    for row in await conn.fetch(
        "SELECT sku_id, particulars, item_type FROM all_sku WHERE sfg_code IS NULL"
    ):
        k = normalise_key(row["particulars"])
        if k:
            claimable.setdefault(k, []).append((row["sku_id"], row["item_type"]))
    for k in claimable:
        claimable[k].sort()  # ascending sku_id

    retyped = 0
    inserted = 0
    retype_log: list[tuple[str, str, str | None]] = []
    for code, (name, group) in seed.items():
        # (a) Re-type an existing article matched by NORMALISED name, claiming
        #     the code + tagging sale_group='wip' (consistent with net-new rows),
        #     while PRESERVING sku_id / item_group / uom / gst. The NOT EXISTS +
        #     row lock make it safe to run concurrently (the loser updates 0 rows
        #     and falls through to the idempotent insert).
        match = claimable.get(name)
        if match:
            sku_id, prev_type = match[0]
            status = await conn.execute(
                """
                UPDATE all_sku
                   SET item_type = 'sfg', sfg_code = $1, sale_group = 'wip'
                 WHERE sku_id = $2 AND sfg_code IS NULL
                   AND NOT EXISTS (SELECT 1 FROM all_sku WHERE sfg_code = $1)
                """,
                code, sku_id,
            )
            if status.rsplit(" ", 1)[-1] != "0":  # robust vs 'UPDATE N'
                retyped += 1
                retype_log.append((code, name, prev_type))
                claimable[name].pop(0)             # don't let another code re-claim it
                continue

        # (b) Insert the derived net-new SFG with an app-supplied 8-digit BIGINT
        #     PK. new_short_time_id() is re-evaluated on every retry so a same-ms
        #     sku_id collision resolves with a fresh id; ON CONFLICT (sfg_code)
        #     DO NOTHING makes the sfg_code key concurrency-safe (a racing worker
        #     that already inserted this code -> 0 rows, no aborting violation).
        async def _insert_sfg(_code=code, _name=name, _group=group):
            return await conn.fetchval(
                """
                INSERT INTO all_sku (sku_id, sfg_code, particulars, item_type, item_group, sale_group)
                VALUES ($1, $2, $3, 'sfg', $4, 'wip')
                ON CONFLICT (sfg_code) WHERE sfg_code IS NOT NULL DO NOTHING
                RETURNING sku_id
                """,
                new_short_time_id(), _code, _name, _group,
            )

        new_id = await insert_with_pk_retry(conn, _insert_sfg, max_retries=5)
        if new_id is not None:
            inserted += 1

    # Audit + drift check: which existing rows were reclassified, and whether the
    # re-typed count matches the plug's own already_in_all_sku='Y' expectation.
    if retype_log:
        logger.info(
            "SFG re-typed %d existing rows in place (sku_id preserved): %s",
            len(retype_log),
            ", ".join(f"{c}<-[{p}]{n[:32]}" for c, n, p in retype_log[:60]),
        )
    if retyped != retype_expected:
        logger.warning(
            "SFG re-type count %d != plug already_in_all_sku='Y' count %d — plug "
            "flag drift (e.g. SFG0334 pre-exists as a sellable FG but is flagged "
            "'N'); the matching article was reclassified to 'sfg'. Verify the catalogue.",
            retyped, retype_expected,
        )
    logger.info(
        "SFG catalogue ingest: %d inserted, %d re-typed (target %d)",
        inserted, retyped, len(seed),
    )
    return {"status": "ok", "inserted": inserted, "retyped": retyped,
            "target": len(seed)}


# ---------------------------------------------------------------------------
# 1c. Process-category classification (Slice 2)
# ---------------------------------------------------------------------------
# Canonical per-token classification, derived from ProcessCategory_to_Operation.csv.
# The Process Category column is split on '+' into individual tokens (so the CSV's
# combined rows like "Blending + Bar Forming" are decomposed here per token), and
# real data carries "X (Bulk Packaging" artifacts (unbalanced paren from the
# split) — both handled by matching on the LEADING token (text before '(').
#
# G2 LOCKED: Sorting = inline (never a buffered SFG stage); Packaging family =
# terminal (produces the FG).

STAGE_CREATE_WIP = "Create WIP"
STAGE_FINAL_FG = "Final FG"
STAGE_INLINE = "inline"

# transform token -> practical operation (each produces / feeds an SFG)
_TRANSFORM_OPS = {
    "de-seeding": "De-Seeding",
    "deseeding": "De-Seeding",
    "blanching": "Blanch & Slice",
    "slicing/dicing/slivering": "Blanch & Slice",
    "slicing": "Blanch & Slice",
    "dicing": "Blanch & Slice",
    "slivering": "Blanch & Slice",
    "blending": "Blend & Form",
    "bar forming": "Blend & Form",
    "roasting": "Roasting",            # -> "Roast & Flavour/Salt" if seasoning co-present
    "flavouring": "Roast & Flavour/Salt",
    "salting": "Roast & Flavour/Salt",
    "stuffing": "Stuffing",
    "enrobing": "Enrobe / Choco-Coat",
    "chocolate": "Enrobe / Choco-Coat",
    # Bar-line token (G4): a bar-line "Mixing" step is a Create-WIP transform —
    # same op family as Blending (it blends/forms the bar mass), so it produces
    # an SFG, not packaging.
    "mixing": "Blend & Form",
}
_SEASONING_TOKENS = {"flavouring", "salting"}
# terminal tokens (produce the FG). "flow" is the bar-line G4 standalone token =
# flow-wrap abbreviated (terminal / Final-FG), distinct from the full "flow wrap".
_TERMINAL_TOKENS = {
    "packaging", "packing", "bulk packaging", "master carton", "mono carton", "monocarton",
    "flow wrap", "flow", "krugger", "x-ray", "xray", "weighing",
}
# inline tokens (no buffered stage) — G2: sorting is inline
_INLINE_TOKENS = {"sorting", "receiving"}


def _canon_process_token(name: str) -> str:
    """Leading canonical token: lowercased, trimmed, dropping a trailing
    '(...' artifact left by the '+' split (e.g. 'Roasting (Bulk Packaging')."""
    t = (name or "").strip().lower()
    return t.split("(")[0].strip()


def classify_route_steps(step_names: list[str]) -> list[dict]:
    """Classify an FG's ORDERED process-route tokens into per-step
    (practical_operation, stage_bucket, produces_sfg).

    Combine rule (ground reality): if the step set has Roasting AND a seasoning
    token, the Roasting step is labelled the combined 'Roast & Flavour/Salt'.
    Unknown tokens are left unclassified (NULL) rather than guessed.
    """
    canon = [_canon_process_token(s) for s in step_names]
    has_season = any(t in _SEASONING_TOKENS for t in canon)
    out: list[dict] = []
    for name, t in zip(step_names, canon):
        if t in _TRANSFORM_OPS:
            op = _TRANSFORM_OPS[t]
            if t == "roasting" and has_season:
                op = "Roast & Flavour/Salt"
            out.append({"process_name": name, "practical_operation": op,
                        "stage_bucket": STAGE_CREATE_WIP, "produces_sfg": True})
        elif t in _TERMINAL_TOKENS:
            out.append({"process_name": name, "practical_operation": "Packaging",
                        "stage_bucket": STAGE_FINAL_FG, "produces_sfg": False})
        elif t in _INLINE_TOKENS:
            out.append({"process_name": name, "practical_operation": None,
                        "stage_bucket": STAGE_INLINE, "produces_sfg": False})
        else:
            out.append({"process_name": name, "practical_operation": None,
                        "stage_bucket": None, "produces_sfg": False})
    return out


async def ingest_process_category_rules(conn, file_path: Path | None = None,
                                        *, full: bool = False) -> dict:
    """Backfill bom_process_route.practical_operation + stage_bucket (Slice 2).

    Re-entrant + idempotent: by default only re-classifies routes that have at
    least one UNCLASSIFIED step (stage_bucket IS NULL), re-running each such
    route's FULL ordered step set so the combine rule is correct when a step is
    appended later (e.g. by sync_fg_master). Pass ``full=True`` to re-label every
    route (used by the cold ingest). A no-op when nothing needs classifying.

    Concurrency-safe: takes a transaction-scoped advisory lock so two startups
    can't run it at once. MUST run inside an outer transaction.
    """
    await conn.execute("SELECT pg_advisory_xact_lock($1)", PROCESS_CLASS_LOCK)

    if full:
        rows = await conn.fetch(
            "SELECT bom_id, step_number, process_name FROM bom_process_route "
            "ORDER BY bom_id, step_number"
        )
    else:
        # Only routes with an unclassified step — but fetch ALL their steps so the
        # combine rule sees the full ordered set.
        rows = await conn.fetch(
            """
            SELECT bom_id, step_number, process_name
              FROM bom_process_route
             WHERE bom_id IN (
                   SELECT DISTINCT bom_id FROM bom_process_route WHERE stage_bucket IS NULL)
             ORDER BY bom_id, step_number
            """
        )

    by_bom: dict[int, list] = {}
    for r in rows:
        by_bom.setdefault(r["bom_id"], []).append(r)

    params: list[tuple] = []
    create_wip = 0
    for bom_id, steps in by_bom.items():
        cls = classify_route_steps([s["process_name"] for s in steps])
        for s, c in zip(steps, cls):
            params.append((c["practical_operation"], c["stage_bucket"], bom_id, s["step_number"]))
            if c["stage_bucket"] == STAGE_CREATE_WIP:
                create_wip += 1

    if params:
        await conn.executemany(
            """
            UPDATE bom_process_route
               SET practical_operation = $1, stage_bucket = $2
             WHERE bom_id = $3 AND step_number = $4
            """,
            params,
        )

    logger.info(
        "Process-category classification: %d steps labelled (%d Create-WIP, %d routes)",
        len(params), create_wip, len(by_bom),
    )
    return {"status": "ok", "steps_classified": len(params),
            "create_wip_steps": create_wip, "routes": len(by_bom)}


# ---------------------------------------------------------------------------
# 1d. JC Routing → bom_process_route (the 2-stage chain + SFG seam, Slice 3)
# ---------------------------------------------------------------------------

def _route_stage_slug(s: str | None) -> str:
    """Slugify an operation/label into a clean lowercase ``stage`` token —
    alphanumeric runs joined by '_' ('Blend & Form' -> 'blend_form',
    'Roast & Flavour/Salt' -> 'roast_flavour_salt'). Empty -> '' so the caller
    can supply a fallback. Never yields a 'pack'-bearing token for a WIP op, so
    a Create-WIP step is never mis-sorted as packing by order_steps_packing_last."""
    return re.sub(r"[^a-z0-9]+", "_", (s or "").strip().lower()).strip("_")


async def ingest_jc_routing(conn, file_path: Path | None = None,
                            *, full: bool = False) -> dict:
    """Load the authoritative 2-stage routing chain (JC_Routing_Plug.csv) into
    bom_process_route, REPLACING each covered article's route with its 1-or-2
    step chain and stamping the SFG seam columns (Slice 3).

    Per plug article (resolved to a bom_id via bom_header on the NORMALISED name
    + entity; falls back to name alone when unambiguous):
      * 2-stage: step 1 = Create WIP (output_kind=SFG, output_code=SFG####),
        step 2 = Final FG (input_kind=SFG, input_code=SFG####, output_kind=FG);
      * pack-only: a single Final FG step (input RM / output FG).
    The route is REPLACED, not merged: each step is upserted by
    (bom_id, step_number) and any surplus steps left by the earlier
    Process-Category split (step_number > the plug's step count) are deleted.
    process_name <- stage_label; stage <- 'packing' for the terminal Final FG
    step (so order_steps_packing_last keeps it last + EGA detection works) and a
    slug of the WIP operation for a Create-WIP step. input_code/output_code are
    nulled for non-SFG kinds (the plug writes 'RM'/'FG' into the code columns).

    This supersedes ingest_fg_master's naive per-FG derivation for the 1085
    plug-covered articles; articles NOT in the plug keep that derivation (then
    get classified by ingest_process_category_rules) — so the non-routed path is
    untouched.

    Re-entrant + idempotent. full=True (cold ingest) processes every plug
    article; full=False (warm path) skips articles whose route is already
    seam-stamped (any step carries output_kind), so warm re-runs are cheap.
    Concurrency-safe via a transaction-scoped advisory lock. MUST run inside an
    outer transaction. Skips with a warning if the plug is absent.
    """
    path = Path(file_path) if file_path else JC_ROUTING_PLUG
    if not path.exists():
        logger.warning("JC routing ingest skipped — plug not found: %s", path)
        return {"status": "file_not_found", "articles": 0, "steps": 0}

    # Serialise concurrent loaders (cold-start storm / migrate overlap) — the
    # DELETE-then-upsert replacement is not safe to interleave across txns.
    await conn.execute("SELECT pg_advisory_xact_lock($1)", ROUTING_LOCK)

    # ---- read + group the plug by (normalised article, entity) ----
    with path.open(newline="", encoding="utf-8-sig") as fh:
        plug_rows = list(csv.DictReader(fh))

    by_art: dict[tuple[str, str | None], list[dict]] = {}
    for r in plug_rows:
        art = normalise_key(r.get("article"))
        if not art:
            continue
        entity = (_safe_str(r.get("entity")) or "").lower() or None
        by_art.setdefault((art, entity), []).append(r)

    # ---- resolve article -> bom_id via the generic (customer_name IS NULL)
    #      bom_header rows (matching ingest_fg_master's keying) ----
    headers = await conn.fetch(
        "SELECT bom_id, fg_sku_name, entity FROM bom_header WHERE customer_name IS NULL"
    )
    by_name_entity: dict[tuple[str, str | None], int] = {}
    by_name: dict[str, int] = {}
    name_dupe: set[str] = set()
    entity_dupe: set[tuple[str, str | None]] = set()
    for h in headers:
        nk = normalise_key(h["fg_sku_name"])
        if not nk:
            continue
        ent = (_safe_str(h["entity"]) or "").lower() or None
        key = (nk, ent)
        # Detect a same-(name,entity) collision on the PRIMARY lookup too — not
        # just the by_name fallback — so a duplicate generic bom_header never
        # silently routes the plug to an arbitrary bom_id (there is no unique
        # constraint on bom_header(fg_sku_name, entity)).
        if key in by_name_entity and by_name_entity[key] != h["bom_id"]:
            entity_dupe.add(key)
        by_name_entity[key] = h["bom_id"]
        if nk in by_name and by_name[nk] != h["bom_id"]:
            name_dupe.add(nk)
        by_name[nk] = h["bom_id"]

    # ---- warm-path scope: bom_ids whose route is already seam-stamped ----
    stamped: set[int] = set()
    if not full:
        rows = await conn.fetch(
            "SELECT DISTINCT bom_id FROM bom_process_route WHERE output_kind IS NOT NULL"
        )
        stamped = {r["bom_id"] for r in rows}

    articles = 0
    steps_written = 0
    two_stage = 0
    unresolved = 0
    ambiguous = 0
    skipped_stamped = 0

    for (art, entity), rows in by_art.items():
        if (art, entity) in entity_dupe:
            # Ambiguous: >1 generic bom_header with this exact (name, entity).
            # Refuse to guess — skip + warn rather than route to an arbitrary one.
            ambiguous += 1
            logger.warning("JC routing: ambiguous (name=%r, entity=%r) -> multiple "
                           "bom_header rows; route skipped", art, entity)
            continue
        bom_id = by_name_entity.get((art, entity))
        if bom_id is None and art not in name_dupe:
            # entity mismatch (e.g. bom_header.entity NULL) — resolve by name
            # alone, but only when that name is unambiguous across entities.
            bom_id = by_name.get(art)
        if bom_id is None:
            unresolved += 1
            continue
        if not full and bom_id in stamped:
            skipped_stamped += 1
            continue

        steps = sorted(rows, key=lambda r: _safe_int(r.get("stage_seq")) or 0)
        step_nums = [n for n in (_safe_int(r.get("stage_seq")) for r in steps) if n and n >= 1]
        if not step_nums:
            unresolved += 1
            continue
        max_step = max(step_nums)

        # Replace: drop every route step NOT in the plug's authoritative step
        # set (not just the tail). `> max_step` would orphan a stale middle step
        # if the plug ever has a gap (e.g. [1,3]); matching the exact set is gap-
        # safe. The plug is contiguous 1..N today, so this is defensive.
        await conn.execute(
            "DELETE FROM bom_process_route WHERE bom_id=$1 AND step_number <> ALL($2::int[])",
            bom_id, step_nums,
        )
        for r in steps:
            step_no = _safe_int(r.get("stage_seq")) or 0
            if step_no < 1:
                continue
            # Canonicalise the bucket so an NBSP/case variant in the plug can't
            # slip the terminal-step detection (mirrors the normalise_key care on
            # the article-name join). 'final fg' -> 'Final FG', 'create wip' ->
            # 'Create WIP'; anything else is stored as-cleaned.
            bucket_raw = _safe_str(r.get("stage_bucket"))
            bucket_norm = (normalise_key(bucket_raw) or "").casefold()
            is_final = bucket_norm == "final fg"
            stage_bucket = ("Final FG" if is_final
                            else "Create WIP" if bucket_norm == "create wip"
                            else bucket_raw)
            label = _safe_str(r.get("stage_label")) or stage_bucket or f"Step {step_no}"
            wip_op = _safe_str(r.get("create_wip_operation"))
            input_kind = _safe_str(r.get("input_kind"))
            output_kind = _safe_str(r.get("output_kind"))
            input_code = _safe_str(r.get("input_code"))
            output_code = _safe_str(r.get("output_code"))
            # The code columns are only meaningful for SFG kinds; the plug writes
            # the literal 'RM'/'FG' into them for non-SFG rows — null those out.
            if (input_kind or "").upper() != "SFG":
                input_code = None
            if (output_kind or "").upper() != "SFG":
                output_code = None
            stage = ("packing" if is_final
                     else (_route_stage_slug(wip_op) or "create_wip"))
            await conn.execute(
                """
                INSERT INTO bom_process_route (
                    bom_id, step_number, process_name, stage,
                    practical_operation, stage_bucket,
                    input_kind, output_kind, input_code, output_code
                ) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10)
                ON CONFLICT (bom_id, step_number) DO UPDATE SET
                    process_name        = EXCLUDED.process_name,
                    stage               = EXCLUDED.stage,
                    practical_operation = EXCLUDED.practical_operation,
                    stage_bucket        = EXCLUDED.stage_bucket,
                    input_kind          = EXCLUDED.input_kind,
                    output_kind         = EXCLUDED.output_kind,
                    input_code          = EXCLUDED.input_code,
                    output_code         = EXCLUDED.output_code
                """,
                bom_id, step_no, label, stage,
                wip_op, stage_bucket,
                input_kind, output_kind, input_code, output_code,
            )
            steps_written += 1
        articles += 1
        if max_step >= 2:
            two_stage += 1

    logger.info(
        "JC routing: %d articles routed (%d two-stage), %d steps written "
        "(%d unresolved, %d ambiguous, %d already-stamped)",
        articles, two_stage, steps_written, unresolved, ambiguous, skipped_stamped,
    )
    return {"status": "ok", "articles": articles, "two_stage": two_stage,
            "ambiguous": ambiguous,
            "steps": steps_written, "unresolved": unresolved,
            "skipped_stamped": skipped_stamped}


# ---------------------------------------------------------------------------
# 1e. JC BOM → bom_line SFG inputs (Slice 4, gate G1 = Option B)
# ---------------------------------------------------------------------------

async def ingest_jc_bom(conn, file_path: Path | None = None) -> dict:
    """Load the SFG consumption lines (JC_BOM_Plug.csv, item_type='sfg') into
    bom_line so a Stage-2 / pack-of-existing-SFG card consumes the SFG as a real
    BOM input (Slice 4, gate G1 Option B).

    The plug has 5017 lines (rm2784/pm1759/fg373/sfg101). The RM/PM/FG lines
    already exist in bom_line via ingest_bom_lines (BOM_Enrichment.xlsx) — we do
    NOT re-load them (that would duplicate). Only the **101 net-new sfg lines**
    are inserted here, each:
      * resolved to the parent FG's bom_id via bom_header on the NORMALISED
        parent_article name (customer_name IS NULL; skipped when ambiguous),
      * material_sku_name = the SFG#### code (so consumption / WIP inventory /
        the chain seam all key on the same code),
      * item_type='sfg', quantity_per_unit = bom_qty, consumed_at_stage from the
        plug ('Final FG (opening RM)'), appended at MAX(line_number)+1.

    Idempotent: skips any (bom_id, SFG####) sfg line already present (one upfront
    query), so cold and warm re-runs converge. MUST run inside an outer
    transaction. Skips with a warning if the plug is absent.  (dep: 050
    consumed_at_stage column, Slice-1 catalogue, Slice-3 routing for bom_header)
    """
    path = Path(file_path) if file_path else JC_BOM_PLUG
    if not path.exists():
        logger.warning("JC BOM ingest skipped — plug not found: %s", path)
        return {"status": "file_not_found", "sfg_lines": 0}

    with path.open(newline="", encoding="utf-8-sig") as fh:
        plug_rows = list(csv.DictReader(fh))

    # Parsed-count audit (acceptance: rm2784/pm1759/fg373/sfg101).
    counts: dict[str, int] = {}
    sfg_rows: list[dict] = []
    for r in plug_rows:
        it = (_safe_str(r.get("item_type")) or "").lower()
        counts[it] = counts.get(it, 0) + 1
        if it == "sfg":
            sfg_rows.append(r)

    # Resolve parent_article -> bom_id via generic bom_header (name only — the
    # plug carries no entity column). Ambiguous normalised names are skipped.
    headers = await conn.fetch(
        "SELECT bom_id, fg_sku_name FROM bom_header WHERE customer_name IS NULL"
    )
    by_name: dict[str, int] = {}
    name_dupe: set[str] = set()
    for h in headers:
        nk = normalise_key(h["fg_sku_name"])
        if not nk:
            continue
        if nk in by_name and by_name[nk] != h["bom_id"]:
            name_dupe.add(nk)
        by_name[nk] = h["bom_id"]

    # Per-bom next line_number (append after the existing RM/PM lines), and the
    # set of (bom_id, SFG####) sfg lines already loaded (idempotency).
    existing = await conn.fetch(
        "SELECT bom_id, line_number, item_type, material_sku_name FROM bom_line"
    )
    max_line: dict[int, int] = {}
    have_sfg: set[tuple[int, str]] = set()
    for e in existing:
        max_line[e["bom_id"]] = max(max_line.get(e["bom_id"], 0), e["line_number"] or 0)
        if (e["item_type"] or "").lower() == "sfg":
            have_sfg.add((e["bom_id"], e["material_sku_name"]))

    inserted = 0
    unresolved = 0
    ambiguous = 0
    skipped_existing = 0
    seen_in_run: set[tuple[int, str]] = set()

    for r in sfg_rows:
        art = normalise_key(r.get("parent_article"))
        code = _safe_str(r.get("sfg_code"))
        if not art or not code:
            unresolved += 1
            continue
        if art in name_dupe:
            ambiguous += 1
            logger.warning("JC BOM: ambiguous parent_article %r -> multiple "
                           "bom_header rows; sfg line skipped", art)
            continue
        bom_id = by_name.get(art)
        if bom_id is None:
            unresolved += 1
            continue
        key = (bom_id, code)
        if key in have_sfg or key in seen_in_run:
            skipped_existing += 1
            continue
        seen_in_run.add(key)
        line_no = max_line.get(bom_id, 0) + 1
        max_line[bom_id] = line_no
        await conn.execute(
            """
            INSERT INTO bom_line (
                bom_id, line_number, material_sku_name, item_type,
                quantity_per_unit, uom, loss_pct, godown, consumed_at_stage
            ) VALUES ($1, $2, $3, 'sfg', $4, $5, 0, $6, $7)
            ON CONFLICT (bom_id, line_number) DO NOTHING
            """,
            bom_id, line_no, code,
            _safe_float(r.get("bom_qty")) or 0,
            "kg", _safe_str(r.get("godown")),
            _safe_str(r.get("consumed_at_stage")) or "Final FG (opening RM)",
        )
        inserted += 1

    logger.info(
        "JC BOM: parsed %s; inserted %d sfg lines (%d unresolved, %d ambiguous, "
        "%d already-present)",
        counts, inserted, unresolved, ambiguous, skipped_existing,
    )
    return {"status": "ok", "parsed_counts": counts, "sfg_lines": inserted,
            "unresolved": unresolved, "ambiguous": ambiguous,
            "skipped_existing": skipped_existing}


# ---------------------------------------------------------------------------
# 2. BOM Enrichment → bom_line
# ---------------------------------------------------------------------------

async def ingest_bom_lines(
    conn, file_path: Path, header_map: dict, master_items: list[MasterItem],
) -> dict:
    """Read BOM_Enrichment sheet → bom_line."""

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb['BOM_Enrichment']

    lines_created = 0
    matched = 0
    unmatched = 0
    skipped = 0
    line_counters: dict[int, int] = {}  # bom_id → next line_number

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 1:  # header row
            continue

        vals = list(row)
        while len(vals) < 13:
            vals.append(None)

        fg_name = _safe_str(vals[1])
        if not fg_name:
            skipped += 1
            continue

        variant = _safe_str(vals[2])  # BOM Variant — NULL = default
        component = _safe_str(vals[3])
        mat_type = _safe_str(vals[4])
        godown = _safe_str(vals[5])
        uom = _safe_str(vals[6])
        qty_per_unit = _safe_float(vals[7])
        loss_pct = _safe_float(vals[8])
        unit_rate = _safe_float(vals[9])
        process_stage_raw = _safe_str(vals[10])

        if not component or qty_per_unit is None:
            skipped += 1
            continue

        # Clean process_stage
        process_stage = process_stage_raw
        if process_stage and 'will be done' in process_stage.lower():
            process_stage = None

        # Lookup bom_id
        key = (fg_name, variant)
        bom_id = header_map.get(key)
        if bom_id is None and variant is not None:
            # Try default BOM
            bom_id = header_map.get((fg_name, None))
        if bom_id is None:
            skipped += 1
            continue

        # Line number
        line_num = line_counters.get(bom_id, 0) + 1
        line_counters[bom_id] = line_num

        # Normalize item_type
        item_type = mat_type.lower() if mat_type else 'rm'

        # Fuzzy match component against all_sku
        if master_items:
            mi, score = match_sku(component, master_items)
            if mi:
                matched += 1
            else:
                unmatched += 1

        await conn.execute(
            """
            INSERT INTO bom_line (
                bom_id, line_number, material_sku_name, item_type,
                quantity_per_unit, uom, loss_pct, godown,
                unit_rate_inr, process_stage
            ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10)
            ON CONFLICT (bom_id, line_number) DO NOTHING
            """,
            bom_id, line_num, component, item_type,
            qty_per_unit, uom, loss_pct or 0, godown,
            unit_rate, process_stage,
        )
        lines_created += 1

    wb.close()
    logger.info("BOM Lines: %d created, %d matched, %d unmatched, %d skipped",
                lines_created, matched, unmatched, skipped)
    return {"lines_created": lines_created, "matched": matched, "unmatched": unmatched}


# ---------------------------------------------------------------------------
# 3. Floorwise machine list → machine
# ---------------------------------------------------------------------------

async def ingest_machines(conn, file_path: Path) -> dict:
    """Read floorwise machine list → machine table. Inserts ALL machines."""

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb['floorwise machine list']

    machines_created = 0
    current_area = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # header row
            continue

        vals = list(row)
        while len(vals) < 4:
            vals.append(None)

        area = _safe_str(vals[0])
        machine_name = _safe_str(vals[1])

        if area and area != 'Area':
            current_area = area

        if machine_name and not area:
            # This is a machine under current_area
            await conn.execute(
                """
                INSERT INTO machine (machine_name, floor, factory, entity, allocation)
                VALUES ($1, $2, 'W202', 'cfpl', 'idle')
                """,
                machine_name, current_area,
            )
            machines_created += 1

    wb.close()
    logger.info("Machines: %d created", machines_created)
    return {"machines_created": machines_created}


# ---------------------------------------------------------------------------
# 4. Derive machine_capacity from FG Master (machine × stage × group)
# ---------------------------------------------------------------------------

# Default kg/hr rates by stage — reasonable food processing industry defaults.
# These are starting estimates; editable later via API.
_DEFAULT_CAPACITY_BY_STAGE = {
    'sorting':                          150,
    'sorting (bulk packaging':          200,
    'sorting (bulk packaging)':         200,
    'roasting':                         100,
    'roasting (bulk packaging':         120,
    'roasting (bulk packaging)':        120,
    'packaging':                        120,
    'flavouring':                       80,
    'flavouring (bulk packaging':       100,
    'flavouring (bulk packaging)':      100,
    'blanching':                        100,
    'slicing/dicing/slivering':         60,
    'slicing/dicing/slivering (bulk packaging': 80,
    'slicing/dicing/slivering (bulk packaging)': 80,
    'de-seeding':                       50,
    'stuffing':                         40,
    'blending':                         80,
    'bar forming':                      60,
    'chocolate':                        50,
    'chocolate (bulk packaging':        60,
    'chocolate (bulk packaging)':       60,
}
_DEFAULT_FALLBACK = 80  # kg/hr when stage not in table

# Multipliers per product group — heavier/harder items are slower
_GROUP_MULTIPLIER = {
    'DATES':            1.2,   # heavy, sticky
    'CASHEW':           1.0,
    'ALMOND':           0.95,
    'PISTA':            0.9,
    'PEANUTS':          1.1,   # small, fast
    'RAISIN':           0.9,
    'SEEDS':            1.0,
    'TRAIL MIX':        0.85,  # mixed, slower
    'BARS & CEREALS':   0.7,   # complex processing
    'FESTIVE HAMPERS':  0.5,   # manual assembly
    'CRANBERRY':        0.9,
    'ANJEER':           0.8,
    'APRICOT':          0.85,
    'BLUEBERRY':        0.85,
    'MAKHANA':          1.0,
    'PREMIUM NUTS':     0.9,
    'PRUNES':           0.85,
    'WALNUT':           0.9,
    'DEHYDRATED FRUITS': 0.9,
    'BLACKCURRANT':     0.85,
    'BLACKBERRY':       0.85,
    'TAJIR':            1.0,
    'SPICES':           0.6,
}


async def derive_machine_capacity(conn, fg_master_path: Path) -> dict:
    """Derive machine_capacity from FG Master machine×stage×group mappings.

    Uses default kg/hr rates by stage, adjusted by product group multiplier.
    """
    wb = openpyxl.load_workbook(fg_master_path, read_only=True, data_only=True)
    ws = wb['FG_Master_Fill']

    # Collect unique (machine_name, stage, group) combos
    combos: set[tuple[str, str, str]] = set()

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        vals = list(row)
        while len(vals) < 16:
            vals.append(None)

        fg_name = _safe_str(vals[1])
        group = _safe_str(vals[2])
        process_cat = _safe_str(vals[4])
        machines_raw = _safe_str(vals[8])

        if not fg_name or fg_name == 'Particluars':
            continue
        if not machines_raw or not process_cat or not group:
            continue

        machines = _split_comma(machines_raw)
        stages = _split_process_category(process_cat)

        for m in machines:
            for s in stages:
                combos.add((m, s, group))

    wb.close()

    # Build machine name → machine_id lookup from DB
    rows = await conn.fetch("SELECT machine_id, machine_name FROM machine")
    # Normalize lookup: lowercase stripped
    machine_lookup: dict[str, int] = {}
    for r in rows:
        key = r['machine_name'].strip().lower()
        machine_lookup[key] = r['machine_id']

    capacity_created = 0
    skipped = 0

    for machine_name, stage_name, group in combos:
        machine_key = machine_name.strip().lower()
        machine_id = machine_lookup.get(machine_key)
        if machine_id is None:
            skipped += 1
            continue

        stage_lower = stage_name.lower().strip()
        base_rate = _DEFAULT_CAPACITY_BY_STAGE.get(stage_lower, _DEFAULT_FALLBACK)
        multiplier = _GROUP_MULTIPLIER.get(group, 0.9)
        capacity = round(base_rate * multiplier, 3)

        await conn.execute(
            """
            INSERT INTO machine_capacity (machine_id, stage, item_group, capacity_kg_per_hr)
            VALUES ($1, $2, $3, $4)
            ON CONFLICT (machine_id, stage, item_group) DO NOTHING
            """,
            machine_id, stage_lower, group, capacity,
        )
        capacity_created += 1

    logger.info("Machine capacity: %d entries created, %d skipped (machine not in DB)", capacity_created, skipped)
    return {"capacity_created": capacity_created, "skipped": skipped}


# ---------------------------------------------------------------------------
# Stock Ingest (Physical Stock + A-185 Stock Take → floor_inventory)
# ---------------------------------------------------------------------------

# Map Excel location columns to floor_location values
# Physical Stock CFPL sheet: pairs of (Qty, KG) columns starting at col 6
_CFPL_LOCATIONS = [
    (6, 7, "rm_store"),           # W-202 Store
    (8, 9, "production_floor"),   # W-202 Lower Basement
    (10, 11, "production_floor"), # W-202 Upper Basement
    (12, 13, "production_floor"), # W-202 1st Floor
    (14, 15, "production_floor"), # W-202 2nd Floor
    (16, 17, "production_floor"), # W-202 Barline
    (18, 19, "production_floor"), # W-202 Terrace
    (20, 21, "offgrade"),         # W-202 Off-Grade
]

# A-185 Stock Take Consolidated sheet: pairs of (Qty, KG) starting at col 6
_A185_LOCATIONS = [
    (6, 7, "rm_store"),           # Rack & Dock Area
    (8, 9, "production_floor"),   # Production
    (10, 11, "production_floor"), # Packing Area
    (12, 13, "fg_store"),         # Cold Area
    (14, 15, "offgrade"),         # Off-Grade
]


async def ingest_stock(conn, data_dir: Path) -> dict:
    """Ingest stock from Physical Stock.xlsx (CFPL) + A-185 Stock Take.xlsx (A-185).

    Populates floor_inventory table. Idempotent — skips if floor_inventory has data.
    """
    count = await conn.fetchval("SELECT COUNT(*) FROM floor_inventory")
    if count and count > 0:
        return {"status": "already_ingested", "count": count}

    inserted = 0

    # --- Physical Stock (CFPL) ---
    cfpl_file = data_dir / "Physical Stock.xlsx"
    if cfpl_file.exists():
        wb = openpyxl.load_workbook(cfpl_file, read_only=True, data_only=True)
        ws = wb['CFPL']
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 3:  # skip header rows (row 1 = totals, row 2 = locations, row 3 = column headers)
                continue
            vals = list(row)
            while len(vals) < 40:
                vals.append(None)

            item_name = _safe_str(vals[1])
            fg_rm = _safe_str(vals[2])
            group = _safe_str(vals[3])
            pack_size = vals[5]

            if not item_name:
                continue

            item_type = 'rm' if fg_rm and fg_rm.upper() == 'RM' else 'fg'

            for qty_col, kg_col, floor_loc in _CFPL_LOCATIONS:
                kg_val = vals[kg_col] if kg_col < len(vals) else None
                if kg_val and isinstance(kg_val, (int, float)) and kg_val > 0:
                    await conn.execute(
                        """
                        INSERT INTO floor_inventory (sku_name, item_type, floor_location, quantity_kg, entity)
                        VALUES ($1, $2, $3, $4, 'cfpl')
                        ON CONFLICT (sku_name, floor_location, lot_number, entity)
                        DO UPDATE SET quantity_kg = floor_inventory.quantity_kg + $4, last_updated = NOW()
                        """,
                        item_name, item_type, floor_loc, float(kg_val),
                    )
                    inserted += 1

        wb.close()

    # --- A-185 Stock Take ---
    a185_file = data_dir / "A-185 Stock Take.xlsx"
    if a185_file.exists():
        wb = openpyxl.load_workbook(a185_file, read_only=True, data_only=True)
        ws = wb['Consolidated']
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 3:
                continue
            vals = list(row)
            while len(vals) < 16:
                vals.append(None)

            item_name = _safe_str(vals[1])
            fg_rm = _safe_str(vals[2])

            if not item_name:
                continue

            item_type = 'rm' if fg_rm and fg_rm.upper() == 'RM' else 'fg'

            for qty_col, kg_col, floor_loc in _A185_LOCATIONS:
                kg_val = vals[kg_col] if kg_col < len(vals) else None
                if kg_val and isinstance(kg_val, (int, float)) and kg_val > 0:
                    await conn.execute(
                        """
                        INSERT INTO floor_inventory (sku_name, item_type, floor_location, quantity_kg, entity)
                        VALUES ($1, $2, $3, $4, 'cdpl')
                        ON CONFLICT (sku_name, floor_location, lot_number, entity)
                        DO UPDATE SET quantity_kg = floor_inventory.quantity_kg + $4, last_updated = NOW()
                        """,
                        item_name, item_type, floor_loc, float(kg_val),
                    )
                    inserted += 1

        wb.close()

    logger.info("Stock ingest: %d inventory entries created", inserted)
    return {"inserted": inserted}


# ---------------------------------------------------------------------------
# 056 enrichment / reconciliation loaders (the 4 companion plug files).
# Each self-guards on a missing table (to_regclass) so a half-migrated DB skips
# with a warning instead of erroring, mirroring ingest_sfg_master's column guard.
# All are idempotent (ON CONFLICT DO UPDATE) and re-entrant.
# ---------------------------------------------------------------------------

async def _table_exists(conn, name: str) -> bool:
    return await conn.fetchval("SELECT to_regclass($1)", name) is not None


async def ingest_sfg_attributes(conn, file_path: Path | None = None) -> dict:
    """Load the rich SFG attributes (JC_Stage1_Create_WIP_Master.csv) into
    sfg_attributes. base_recipe / create_wip_operation / sfg_origin / item_group /
    consumed_by_fgs_count only — suggested_shelf_life_days is intentionally NOT
    ingested (WIP carries no shelf life). Keyed on sfg_code (1:1 with all_sku).
    Idempotent; safe to re-run."""
    path = Path(file_path) if file_path else JC_STAGE1_MASTER
    if not path.exists():
        logger.warning("SFG attributes ingest skipped — file not found: %s", path)
        return {"status": "file_not_found", "rows": 0}
    if not await _table_exists(conn, "sfg_attributes"):
        logger.warning("SFG attributes ingest skipped — sfg_attributes table missing (run 056)")
        return {"status": "schema_missing", "rows": 0}

    rows = 0
    with path.open(encoding="utf-8-sig", newline="") as fh:
        for r in csv.DictReader(fh):
            code = normalise_key(r.get("sfg_code"))
            if not code:
                continue
            # va_article / primary_bu (GAP B): present only in the absent
            # SFG_WIP_Item_List_FINAL source — .get() yields None for the current
            # plug, so the columns stay NULL until that file is dropped in.
            await conn.execute(
                """
                INSERT INTO sfg_attributes
                    (sfg_code, base_recipe, create_wip_operation, sfg_origin,
                     item_group, consumed_by_fgs_count, va_article, primary_bu)
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
                ON CONFLICT (sfg_code) DO UPDATE SET
                    base_recipe           = EXCLUDED.base_recipe,
                    create_wip_operation  = EXCLUDED.create_wip_operation,
                    sfg_origin            = EXCLUDED.sfg_origin,
                    item_group            = EXCLUDED.item_group,
                    consumed_by_fgs_count = EXCLUDED.consumed_by_fgs_count,
                    va_article            = COALESCE(EXCLUDED.va_article, sfg_attributes.va_article),
                    primary_bu            = COALESCE(EXCLUDED.primary_bu, sfg_attributes.primary_bu)
                """,
                code,
                _safe_str(r.get("base_recipe")),
                _safe_str(r.get("create_wip_operation")),
                _safe_str(r.get("sfg_origin")),
                _safe_str(r.get("item_group")),
                _safe_int(r.get("consumed_by_fgs_count")),
                _safe_str(r.get("va_article")),
                _safe_str(r.get("primary_bu")),
            )
            rows += 1
    logger.info("SFG attributes ingest: %d rows upserted", rows)
    return {"status": "ok", "rows": rows}


async def ingest_stage_catalog(conn, file_path: Path | None = None) -> dict:
    """Load the §6 token->operation map (ProcessCategory_to_Operation.csv) into
    stage_catalog. The token string is the key. produces_sfg 'Y'->TRUE.
    Reference config (the route classifier transcribes the same 9 rules)."""
    path = Path(file_path) if file_path else PROCESS_CATEGORY_MAP
    if not path.exists():
        logger.warning("Stage catalog ingest skipped — file not found: %s", path)
        return {"status": "file_not_found", "rows": 0}
    if not await _table_exists(conn, "stage_catalog"):
        logger.warning("Stage catalog ingest skipped — stage_catalog table missing (run 056)")
        return {"status": "schema_missing", "rows": 0}

    rows = 0
    with path.open(encoding="utf-8-sig", newline="") as fh:
        for r in csv.DictReader(fh):
            token = _safe_str(r.get("process_category_token(s)"))
            if not token:
                continue
            produces = (normalise_key(r.get("produces_sfg")).upper() == "Y")
            await conn.execute(
                """
                INSERT INTO stage_catalog
                    (process_category_tokens, practical_operation, stage_bucket,
                     produces_sfg, note)
                VALUES ($1, $2, $3, $4, $5)
                ON CONFLICT (process_category_tokens) DO UPDATE SET
                    practical_operation = EXCLUDED.practical_operation,
                    stage_bucket        = EXCLUDED.stage_bucket,
                    produces_sfg        = EXCLUDED.produces_sfg,
                    note                = EXCLUDED.note
                """,
                token,
                _safe_str(r.get("practical_operation")),
                _safe_str(r.get("stage_bucket")),
                produces,
                _safe_str(r.get("note")),
            )
            rows += 1
    logger.info("Stage catalog ingest: %d rules upserted", rows)
    return {"status": "ok", "rows": rows}


async def ingest_fg_sfg_input_map(conn, file_path: Path | None = None) -> dict:
    """Load the FG->stage-2-input map (JC_Stage2_Final_FG_Packing.csv) into
    fg_sfg_input_map (all 1085 FGs, incl. pack-only input_code='RM'). Idempotent
    on (fg_name, input_code)."""
    path = Path(file_path) if file_path else JC_STAGE2_PACKING
    if not path.exists():
        logger.warning("FG->SFG input map ingest skipped — file not found: %s", path)
        return {"status": "file_not_found", "rows": 0}
    if not await _table_exists(conn, "fg_sfg_input_map"):
        logger.warning("FG->SFG input map ingest skipped — fg_sfg_input_map table missing (run 056)")
        return {"status": "schema_missing", "rows": 0}

    rows = 0
    with path.open(encoding="utf-8-sig", newline="") as fh:
        for r in csv.DictReader(fh):
            fg = _safe_str(r.get("fg_name"))
            code = _safe_str(r.get("input_code"))
            if not fg or not code:
                continue
            await conn.execute(
                """
                INSERT INTO fg_sfg_input_map
                    (fg_name, cycle_archetype, final_stage_label, input_code,
                     input_sfg_article, input_kind)
                VALUES ($1, $2, $3, $4, $5, $6)
                ON CONFLICT (fg_name, input_code) DO UPDATE SET
                    cycle_archetype   = EXCLUDED.cycle_archetype,
                    final_stage_label = EXCLUDED.final_stage_label,
                    input_sfg_article = EXCLUDED.input_sfg_article,
                    input_kind        = EXCLUDED.input_kind
                """,
                fg,
                _safe_str(r.get("cycle_archetype")),
                _safe_str(r.get("final_stage_label")),
                code,
                _safe_str(r.get("input_sfg_article")),
                _safe_str(r.get("input_kind")),
            )
            rows += 1
    logger.info("FG->SFG input map ingest: %d rows upserted", rows)
    return {"status": "ok", "rows": rows}


async def ingest_sfg_resolution_map(conn, file_path: Path | None = None) -> dict:
    """Load the dedup provenance (SFG_Resolution_Map.csv) into sfg_resolution_map.
    Idempotent on (sfg_code, fg_sample)."""
    path = Path(file_path) if file_path else SFG_RESOLUTION_MAP
    if not path.exists():
        logger.warning("SFG resolution map ingest skipped — file not found: %s", path)
        return {"status": "file_not_found", "rows": 0}
    if not await _table_exists(conn, "sfg_resolution_map"):
        logger.warning("SFG resolution map ingest skipped — sfg_resolution_map table missing (run 056)")
        return {"status": "schema_missing", "rows": 0}

    rows = 0
    with path.open(encoding="utf-8-sig", newline="") as fh:
        for r in csv.DictReader(fh):
            code = normalise_key(r.get("sfg_code"))
            fg = _safe_str(r.get("fg_sample"))
            if not code or not fg:
                continue
            await conn.execute(
                """
                INSERT INTO sfg_resolution_map
                    (sfg_code, fg_sample, base_recipe, create_wip_operation,
                     sfg_origin, sfg_article, num_fgs)
                VALUES ($1, $2, $3, $4, $5, $6, $7)
                ON CONFLICT (sfg_code, fg_sample) DO UPDATE SET
                    base_recipe          = EXCLUDED.base_recipe,
                    create_wip_operation = EXCLUDED.create_wip_operation,
                    sfg_origin           = EXCLUDED.sfg_origin,
                    sfg_article          = EXCLUDED.sfg_article,
                    num_fgs              = EXCLUDED.num_fgs
                """,
                code,
                fg,
                _safe_str(r.get("base_recipe")),
                _safe_str(r.get("create_wip_operation")),
                _safe_str(r.get("sfg_origin")),
                _safe_str(r.get("sfg_article")),
                _safe_int(r.get("num_fgs")),
            )
            rows += 1
    logger.info("SFG resolution map ingest: %d rows upserted", rows)
    return {"status": "ok", "rows": rows}


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

async def run_master_ingest(pool, data_dir: Path, master_items: list) -> dict | None:
    """Run all master data ingests. Idempotent — skips if data already exists."""

    count = await pool.fetchval("SELECT COUNT(*) FROM bom_header")
    if count and count > 0:
        # Check if machine_capacity needs backfill
        cap_count = await pool.fetchval("SELECT COUNT(*) FROM machine_capacity")
        if cap_count == 0:
            fg_file = data_dir / "FG_Master_Completion.xlsx"
            if fg_file.exists():
                async with pool.acquire() as conn:
                    async with conn.transaction():
                        cap_result = await derive_machine_capacity(conn, fg_file)
                logger.info("Backfilled machine capacity: %d entries", cap_result["capacity_created"])

        # Check if stock needs ingest
        stock_count = await pool.fetchval("SELECT COUNT(*) FROM floor_inventory")
        if stock_count == 0:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    stock_result = await ingest_stock(conn, data_dir)
                logger.info("Backfilled stock: %s", stock_result)
        else:
            logger.info("Master data already ingested (%d headers, %d capacity, %d stock), skipping",
                         count, cap_count, stock_count)

        # Check if the SFG catalogue needs backfill (Slice 1). Self-guards on a
        # missing sfg_code column and on an already-loaded catalogue.
        sfg_count = await pool.fetchval(
            "SELECT COUNT(*) FROM all_sku WHERE item_type = 'sfg'"
        )
        if not sfg_count:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    sfg_result = await ingest_sfg_master(conn)
                logger.info("Backfilled SFG catalogue: %s", sfg_result)

        # Slice 3: load the 2-stage routing + SFG seam for any not-yet-stamped
        # plug article (re-entrant — skips articles whose route already carries an
        # output_kind). Runs BEFORE classification so the plug's authoritative
        # stage_bucket isn't re-derived from the soon-to-be-replaced process_name.
        async with pool.acquire() as conn:
            async with conn.transaction():
                routing_result = await ingest_jc_routing(conn)
        if routing_result.get("steps"):
            logger.info("Backfilled JC routing: %s", routing_result)

        # Slice 4 (gate G1 = Option B): backfill the 101 sfg consumption bom_lines.
        # Idempotent (skips (bom_id, SFG####) lines already present), so this is a
        # cheap no-op once loaded. Needs bom_header (present on the warm path).
        async with pool.acquire() as conn:
            async with conn.transaction():
                jc_bom_result = await ingest_jc_bom(conn)
        if jc_bom_result.get("sfg_lines"):
            logger.info("Backfilled JC BOM (sfg lines): %s", jc_bom_result)

        # Classify any UNCLASSIFIED routes (Slice 2). Re-entrant + idempotent —
        # cheap no-op when every route is already labelled, but catches routes
        # added/changed by a later sync_fg_master run (the function scopes itself
        # to routes with a NULL stage_bucket step). The plug routes above are now
        # stamped, so this only touches non-routed articles.
        async with pool.acquire() as conn:
            async with conn.transaction():
                cls_result = await ingest_process_category_rules(conn)
        if cls_result["steps_classified"]:
            logger.info("Backfilled process classification: %s", cls_result)

        # 056: the 4 companion plug files (rich SFG attrs, §6 stage catalog,
        # FG->input reconciliation, dedup provenance). Each self-guards on a
        # missing table + is idempotent, so this is a cheap no-op once loaded.
        async with pool.acquire() as conn:
            async with conn.transaction():
                attrs_result = await ingest_sfg_attributes(conn)
                stage_cat_result = await ingest_stage_catalog(conn)
                input_map_result = await ingest_fg_sfg_input_map(conn)
                resolution_result = await ingest_sfg_resolution_map(conn)
        logger.info("Backfilled SFG companion data: attrs=%s stage_catalog=%s input_map=%s resolution=%s",
                    attrs_result, stage_cat_result, input_map_result, resolution_result)

        # 068: FG → canonical SFG catalogue (data/sfg_canonical_map.csv). One-time
        # load — skip when already populated so it doesn't slow every boot.
        from app.modules.production.services.sfg_canonical import ingest_canonical_map
        canon_count = await pool.fetchval("SELECT COUNT(*) FROM sfg_canonical_map") \
            if await pool.fetchval("SELECT to_regclass('sfg_canonical_map')") else None
        if canon_count == 0:
            async with pool.acquire() as conn:
                async with conn.transaction():
                    canon_result = await ingest_canonical_map(conn)
                logger.info("Backfilled SFG canonical map: %s", canon_result)
        return None

    fg_file = data_dir / "FG_Master_Completion.xlsx"
    bom_file = data_dir / "BOM_Enrichment.xlsx"
    machine_file = data_dir / "Floorwise utility dada.xlsx"

    for f in [fg_file, bom_file, machine_file]:
        if not f.exists():
            logger.warning("Master ingest skipped — file not found: %s", f)
            return None

    async with pool.acquire() as conn:
        async with conn.transaction():
            # SFG catalogue first so the codes exist before any later routing
            # build references them (Slice 3+); independent of FG/BOM for Slice 1.
            sfg_result = await ingest_sfg_master(conn)
            header_map = await ingest_fg_master(conn, fg_file, master_items)
            bom_result = await ingest_bom_lines(conn, bom_file, header_map, master_items)
            machine_result = await ingest_machines(conn, machine_file)
            capacity_result = await derive_machine_capacity(conn, fg_file)
            stock_result = await ingest_stock(conn, data_dir)
            # Slice 2: classify the freshly-built routes (practical op + stage bucket).
            await ingest_process_category_rules(conn, full=True)
            # Slice 3: load the authoritative 2-stage routing + SFG seam, REPLACING
            # the naive Process-Category routes for the plug-covered articles. Runs
            # last so the plug has the final word over the classifier.
            routing_result = await ingest_jc_routing(conn, full=True)
            # Slice 4 (gate G1 = Option B): the 101 sfg consumption bom_lines so a
            # Stage-2 / pack-of-existing-SFG card consumes the SFG as a BOM input.
            jc_bom_result = await ingest_jc_bom(conn)
            # 056: rich SFG attributes + §6 stage catalog + FG->input reconciliation
            # + dedup provenance (the 4 companion plug files). Idempotent.
            await ingest_sfg_attributes(conn)
            await ingest_stage_catalog(conn)
            await ingest_fg_sfg_input_map(conn)
            await ingest_sfg_resolution_map(conn)

    logger.info(
        "Master data ingest complete: %d headers, %d BOM lines, %d machines, "
        "%d capacity, %d stock entries, SFG %s, routing %s, jc_bom %s",
        len(header_map), bom_result["lines_created"],
        machine_result["machines_created"], capacity_result["capacity_created"],
        stock_result["inserted"], sfg_result, routing_result, jc_bom_result,
    )
    return {
        "headers": len(header_map),
        "bom_lines": bom_result["lines_created"],
        "machines": machine_result["machines_created"],
        "capacity_entries": capacity_result["capacity_created"],
        "stock_entries": stock_result["inserted"],
        "sfg": sfg_result,
        "routing": routing_result,
        "jc_bom": jc_bom_result,
    }
