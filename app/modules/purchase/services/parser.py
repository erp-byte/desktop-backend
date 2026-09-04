"""Purchase Order Book Excel parser — state-machine parser like so_book_parser."""

import io
import logging
from datetime import date, datetime

import openpyxl

from app.core.helpers import safe_float as _safe_float_or_none, safe_str as _safe_str

logger = logging.getLogger(__name__)


def _parse_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    return s if s else None


# ── Dynamic column detection ─────────────────────────────────────────────
#
# Column POSITIONS are not stable between exports. Tally emits a column only
# for ledgers actually used in the period, and the optional voucher columns
# ("Terms of Payment", "Other References", "Terms of Delivery") come and go:
#
#   Jan-Mar 26 : [9]=Quantity [10]=Alt. Units [11]=Rate [12]=Value
#   Aug 26     : [6]=Quantity  [7]=Alt. Units  [8]=Rate  [9]=Value
#
# Reading fixed indices against the Aug layout silently loaded the Value column
# as pack_count and the SGST column as total_amount. Positions are therefore
# resolved from the header row by NAME, exactly like so_book_parser does.
#
# Matching is plain lowercase substring, first match wins per field, scanning
# columns left to right. Order matters — put the specific pattern first:
#   • "order reference" before anything that would also match "Other References"
#   • "packing charges" NOT "packing", or "Purchase of Packing Material" wins
#   • "freight" NOT "local",  or "Purchase of Raw Material - Local" wins
_HEADER_ALIASES: tuple[tuple[str, str], ...] = (
    ("order reference", "order_reference_no"),
    ("voucher type", "voucher_type"),
    ("vch type", "voucher_type"),
    ("voucher no", "po_number"),
    ("vch no", "po_number"),
    ("narration", "narration"),
    ("particulars", "particulars"),
    ("date", "date"),
    ("alt. units", "alt_units"),
    ("alt units", "alt_units"),
    ("quantity", "quantity"),
    ("qty", "quantity"),
    ("rate", "rate"),
    ("gross total", "gross_total"),
    ("value", "value"),
    ("sgst", "sgst_amount"),
    ("cgst", "cgst_amount"),
    ("igst", "igst_amount"),
    ("round off", "round_off"),
    ("apmc", "apmc_tax"),
    ("packing charges", "packing_charges"),
    ("loading", "loading_unloading_charges"),
    ("transportation", "freight_transport_charges"),
    ("freight", "freight_transport_local"),
    ("other charges", "other_charges_non_gst"),
)

# GL ledger fields lifted off the header row (everything except the identity
# and line-level columns).
_GL_FIELDS: tuple[str, ...] = (
    "gross_total", "sgst_amount", "cgst_amount", "igst_amount", "round_off",
    "apmc_tax", "packing_charges", "freight_transport_local",
    "freight_transport_charges", "loading_unloading_charges",
    "other_charges_non_gst",
)

# Fallback when no header row can be found: the historical Jan-Mar layout.
_DEFAULT_COL_MAP: dict[str, int] = {
    "date": 0, "particulars": 1, "voucher_type": 2, "po_number": 3,
    "order_reference_no": 4, "narration": 5,
    "quantity": 9, "alt_units": 10, "rate": 11, "value": 12,
    "gross_total": 13, "sgst_amount": 15, "cgst_amount": 16, "round_off": 17,
    "igst_amount": 20, "packing_charges": 23, "freight_transport_local": 26,
    "apmc_tax": 27, "other_charges_non_gst": 28,
    "freight_transport_charges": 40, "loading_unloading_charges": 42,
}
_DEFAULT_DATA_ROW = 13


def _detect_columns(ws) -> tuple[dict[str, int], int]:
    """Find the header row and build {field: column index}.

    Returns (col_map, data_start_row) with data_start_row 1-indexed.
    """
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        joined = " ".join(cells)
        # A real header row names several of these; the letterhead names none.
        if sum(kw in joined for kw in
               ("date", "particulars", "voucher", "quantity", "rate", "value")) < 3:
            continue

        col_map: dict[str, int] = {}
        for col_idx, text in enumerate(cells):
            if not text:
                continue
            for alias, field in _HEADER_ALIASES:
                if field not in col_map and alias in text:
                    col_map[field] = col_idx
                    break

        # A PARTIAL map is worse than no map: mixing detected identity columns
        # with fallback line columns would read Quantity/Rate/Value from
        # positions this workbook never agreed to. Require the two fields that
        # define a line row before trusting the row as a header.
        if "particulars" in col_map and "quantity" in col_map:
            logger.info("Detected PO Book headers at row %d: %s", row_idx, col_map)
            return col_map, row_idx + 1

    logger.warning(
        "Could not detect PO Book headers — falling back to the Jan-Mar column "
        "layout. If this workbook uses a different layout, every line field "
        "from Quantity rightwards will be read from the wrong column."
    )
    return dict(_DEFAULT_COL_MAP), _DEFAULT_DATA_ROW


def _get(row: tuple, col_map: dict[str, int], field: str):
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _is_header_row(row: tuple, col_map: dict[str, int]) -> bool:
    """Header rows carry the voucher date; line rows leave it blank."""
    val = _get(row, col_map, "date")
    return val is not None and str(val).strip() != ""


def _is_line_row(row: tuple, col_map: dict[str, int]) -> bool:
    """Line rows have no date, but do have Particulars and a Quantity."""
    if _is_header_row(row, col_map):
        return False
    particulars = _get(row, col_map, "particulars")
    qty = _get(row, col_map, "quantity")
    return (particulars is not None and str(particulars).strip() != ""
            and qty is not None and str(qty).strip() != "")


def _is_grand_total_row(row: tuple, col_map: dict[str, int]) -> bool:
    particulars = _safe_str(_get(row, col_map, "particulars"))
    return particulars is not None and particulars.lower().startswith("grand total")


def parse_po_book(file_bytes: bytes) -> list[dict]:
    """
    Parse a Purchase Order Book Excel file.
    Returns a list of PO dicts, each with a 'lines' array.
    Uses state-machine approach: header row starts a new PO, line rows add articles.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    current_po = None
    all_orders = []

    col_map, data_start_row = _detect_columns(ws)

    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        row_vals = tuple(row)

        if _is_grand_total_row(row_vals, col_map):
            continue

        if _is_header_row(row_vals, col_map):
            # Flush previous PO
            if current_po is not None:
                all_orders.append(current_po)

            current_po = {
                "po_date": _parse_date(_get(row_vals, col_map, "date")),
                "vendor_supplier_name": _safe_str(_get(row_vals, col_map, "particulars")),
                "voucher_type": _safe_str(_get(row_vals, col_map, "voucher_type")),
                "po_number": _safe_str(_get(row_vals, col_map, "po_number")),
                "order_reference_no": _safe_str(_get(row_vals, col_map, "order_reference_no")),
                "narration": _safe_str(_get(row_vals, col_map, "narration")),
                # The voucher's own Value column — NOT whatever sits at a fixed
                # index, which on the Aug layout is SGST (ITC).
                "total_amount": _safe_float_or_none(_get(row_vals, col_map, "value")),
                "lines": [],
            }

            # GL ledger amounts, also off the header row. Tally writes a column
            # for every ledger in the period but leaves it blank/0 on vouchers
            # that do not use it, so only non-zero values are recorded.
            for field_name in _GL_FIELDS:
                val = _safe_float_or_none(_get(row_vals, col_map, field_name))
                if val:
                    current_po[field_name] = val

        elif _is_line_row(row_vals, col_map):
            if current_po is None:
                continue

            # "Alt. Units" is Tally's secondary *quantity*, not a unit weight.
            # Kept for reference but must NOT feed `uom`: uom is the per-unit
            # weight sourced from all_sku (FRONTEND_API_DOC.md, "What gets
            # matched from all_sku"), and a stray Alt.-Units value here blocks
            # the master fill and poisons po_weight.
            alt_units = _safe_float_or_none(_get(row_vals, col_map, "alt_units"))

            current_po["lines"].append({
                "sku_name": _safe_str(_get(row_vals, col_map, "particulars")),
                "pack_count": _safe_float_or_none(_get(row_vals, col_map, "quantity")),
                "uom": None,          # filled from all_sku during enrichment
                "alt_units": alt_units,
                # None-returning, NOT safe_float_zero: a blank or unit-bearing
                # Rate/Value cell must stay NULL. Coercing it to 0.0 renders as
                # "Rs.0.00" instead of an em dash, makes po_diff emit a bogus
                # 100% rate-variance warning, and erases the difference between
                # "the export had no rate" and "the supplier charged nothing".
                "rate": _safe_float_or_none(_get(row_vals, col_map, "rate")),
                "amount": _safe_float_or_none(_get(row_vals, col_map, "value")),
            })

    # Flush last PO
    if current_po is not None:
        all_orders.append(current_po)

    wb.close()

    # Add line numbers
    for po in all_orders:
        for i, line in enumerate(po["lines"], start=1):
            line["line_number"] = i

    logger.info(
        "Parsed PO Book: %d POs, %d total lines",
        len(all_orders), sum(len(po["lines"]) for po in all_orders),
    )

    return all_orders
