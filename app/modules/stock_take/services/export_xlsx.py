"""Excel builder for the stock adjustment ledger (no DB access).

Pure openpyxl over rows the service already fetched, returning a BytesIO —
the same shape as customer_returns/services/export_xlsx.py.

The active filters are stamped into the sheet header. A spreadsheet outlives the
screen it was exported from, so a figure with no scope on it is unattributable
the moment it is emailed on; the Stock Take app does the same thing when it
writes "N floor(s) excluded" into its export.
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# (key, header, width). Order is the sheet's column order.
COLUMNS: tuple[tuple[str, str, int], ...] = (
    # The 8-digit YYMMDD+NN reference, not the internal txn_id — this is the
    # number on screen, and a spreadsheet that used a different one could not be
    # cross-referenced against it.
    ("txn_code",         "Txn no.",        11),
    ("created_at",       "Date / time",    19),
    ("warehouse",        "Warehouse",      12),
    ("location",         "Floor",          20),
    ("item_name",        "Article",        44),
    ("material_type",    "Material type",  14),
    ("item_category",    "Category",       22),
    ("item_subcategory", "Sub category",   22),
    ("stock_type",       "Stock type",     18),
    ("operation",        "Operation",      13),
    ("units",            "Units",          10),
    ("qty_kg",           "Qty (kg)",       12),
    ("signed_kg",        "Signed (kg)",    12),
    ("reason",           "Reason",         40),
    ("created_by",       "Recorded by",    20),
    ("sku_id",           "SKU ID",         10),
    ("is_new_article",   "New article",    12),
    ("is_reversal",      "Reversal",       10),
    ("reverses_txn_code", "Reverses txn",  13),
)

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="29417A", end_color="29417A", fill_type="solid")
_TITLE_FONT = Font(bold=True, size=13)
_META_FONT = Font(size=9, color="666666")
_ADD_FILL = PatternFill(start_color="EAF6EC", end_color="EAF6EC", fill_type="solid")
_SUB_FILL = PatternFill(start_color="FDF0E6", end_color="FDF0E6", fill_type="solid")
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _describe(filters: dict[str, Any]) -> str:
    if not filters:
        return "No filters — every recorded transaction."
    parts = []
    for key, label in (("warehouse", "Warehouse"), ("location", "Floor"), ("itemName", "Article"),
                       ("operation", "Operation"), ("date", "Date"),
                       ("dateFrom", "From"), ("dateTo", "To")):
        if filters.get(key):
            parts.append(f"{label}: {filters[key]}")
    return "Filters — " + "; ".join(parts) if parts else "No filters — every recorded transaction."


def build_ledger_workbook(rows: list[dict[str, Any]], filters: dict[str, Any],
                          generated_by: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Transactions"

    ws.cell(row=1, column=1, value="Stock adjustment ledger").font = _TITLE_FONT
    ws.cell(row=2, column=1, value=_describe(filters)).font = _META_FONT
    ws.cell(row=3, column=1,
            value=(f"{len(rows)} transaction(s) · exported "
                   f"{datetime.now().strftime('%Y-%m-%d %H:%M')} by {generated_by}")).font = _META_FONT

    head = 5
    for i, (_key, header, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=head, column=i, value=header)
        c.font, c.fill, c.border = _HEADER_FONT, _HEADER_FILL, _BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width

    for r, row in enumerate(rows, start=head + 1):
        add = row.get("operation") == "ADDITION"
        for i, (key, _header, _w) in enumerate(COLUMNS, start=1):
            if key == "signed_kg":
                # Direction is stored in `operation` and magnitudes are always
                # positive, so a signed column is derived here purely so the
                # spreadsheet can SUM() straight down to the net movement.
                v = (row.get("qty_kg") or 0) * (1 if add else -1)
            elif key == "created_at":
                v = str(row.get("created_at") or "").replace("T", " ")[:19]
            elif key in ("is_new_article", "is_reversal"):
                v = "Yes" if row.get(key) else ""
            else:
                v = row.get(key)
            c = ws.cell(row=r, column=i, value=v)
            c.border = _BORDER
            c.fill = _ADD_FILL if add else _SUB_FILL
            if key in ("units", "qty_kg", "signed_kg"):
                c.number_format = "#,##0.00" if key != "units" else "#,##0.000"
                c.alignment = Alignment(horizontal="right")

    ws.freeze_panes = ws.cell(row=head + 1, column=1)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
