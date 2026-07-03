"""Pure openpyxl builder for the customer-returns Excel export (no DB access).
Header row + styling + edited-cell highlighting; returns a BytesIO stream."""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.modules.customer_returns.services.query_service import EXPORT_COLUMNS

# DB field_name -> export header for the 4 highlightable box columns.
FIELD_TO_HEADER = {
    "net_weight": "Box Net Weight",
    "gross_weight": "Box Gross Weight",
    "lot_number": "Box Lot Number",
    "count": "Box Count",
}

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="29417A", end_color="29417A", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
_EDITED_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def build_export_workbook(rows: list, edited_cells: set) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer Returns"

    # header row
    for col_idx, name in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _BORDER

    # data rows
    for row_idx, row in enumerate(rows, start=2):
        box_id = row.get("Box ID") or ""
        for col_idx, name in enumerate(EXPORT_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(name, ""))
            cell.border = _BORDER
        if box_id:
            for field_name, header in FIELD_TO_HEADER.items():
                if (box_id, field_name) in edited_cells:
                    ws.cell(row=row_idx, column=EXPORT_COLUMNS.index(header) + 1).fill = _EDITED_FILL

    # column widths
    for col_idx, name in enumerate(EXPORT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(name) + 4, 14)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
