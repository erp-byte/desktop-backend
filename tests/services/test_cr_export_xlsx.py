"""Pure test: build the export workbook and read it back. Run:
    PYTHONPATH=. python tests/services/test_cr_export_xlsx.py
"""
from openpyxl import load_workbook
from app.modules.customer_returns.services import export_xlsx
from app.modules.customer_returns.services.query_service import EXPORT_COLUMNS


def _row(**over):
    base = {c: "" for c in EXPORT_COLUMNS}
    base.update({"RTV ID": "CR-1", "Box ID": "50123456-1",
                 "Box Net Weight": 25.0, "Box Lot Number": "L1"})
    base.update(over)
    return base


def main() -> None:
    rows = [_row()]
    edited = {("50123456-1", "net_weight")}   # highlight Box Net Weight only
    buf = export_xlsx.build_export_workbook(rows, edited)
    wb = load_workbook(buf)
    ws = wb.active
    assert ws.title == "Customer Returns"
    header = [c.value for c in ws[1]]
    assert header == EXPORT_COLUMNS and len(header) == 33

    nw_col = EXPORT_COLUMNS.index("Box Net Weight") + 1
    lot_col = EXPORT_COLUMNS.index("Box Lot Number") + 1
    nw_cell = ws.cell(row=2, column=nw_col)
    lot_cell = ws.cell(row=2, column=lot_col)
    assert nw_cell.value == 25.0
    # highlighted cell has the light-red fill; unedited box field does not
    assert nw_cell.fill.start_color.rgb.endswith("FEE2E2"), nw_cell.fill.start_color.rgb
    assert not (lot_cell.fill.fill_type == "solid" and lot_cell.fill.start_color.rgb.endswith("FEE2E2"))

    # empty export still writes a valid header-only sheet
    buf2 = export_xlsx.build_export_workbook([], set())
    ws2 = load_workbook(buf2).active
    assert [c.value for c in ws2[1]] == EXPORT_COLUMNS
    print("ASSERTIONS PASSED")


if __name__ == "__main__":
    main()
