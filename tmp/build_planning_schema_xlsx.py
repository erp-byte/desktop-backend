"""One-shot: build docs/db/production_planning_schema.xlsx from production_schema.sql definitions."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUT = Path(__file__).resolve().parents[1] / "docs" / "db" / "production_planning_schema.xlsx"

# (table, category, purpose, columns)
# columns: (name, type, nullable, default, key, references, notes)
TABLES = [
    {
        "name": "machine",
        "category": "Master Data",
        "purpose": "Physical equipment on the factory floor.",
        "columns": [
            ("machine_id",     "SERIAL",         "NO",  "auto", "PK",  "", ""),
            ("machine_name",   "TEXT",           "NO",  "",     "",    "", ""),
            ("machine_type",   "TEXT",           "YES", "",     "",    "", "sorting_table, sealer, scale, metal_detector"),
            ("category",       "TEXT",           "YES", "",     "",    "", "processing, packaging, quality"),
            ("capable_stages", "TEXT[]",         "YES", "",     "",    "", "{sorting, grading, weighing}"),
            ("floor",          "TEXT",           "YES", "",     "",    "", "production floor name"),
            ("factory",        "TEXT",           "YES", "",     "",    "", "W-202, A-185"),
            ("status",         "TEXT",           "NO",  "'active'", "", "", "active, maintenance, retired"),
            ("entity",         "TEXT",           "YES", "",     "",    "", "CHECK in (cfpl, cdpl)"),
            ("created_at",     "TIMESTAMPTZ",    "NO",  "NOW()", "",   "", ""),
        ],
    },
    {
        "name": "machine_capacity",
        "category": "Master Data",
        "purpose": "Throughput per machine per stage per product group (kg/hr).",
        "columns": [
            ("capacity_id",        "SERIAL",        "NO", "auto",  "PK", "", ""),
            ("machine_id",         "INT",           "NO", "",      "FK", "machine(machine_id)", ""),
            ("stage",              "TEXT",          "NO", "",      "",   "", "sorting, weighing, sealing, metal_detection"),
            ("item_group",         "TEXT",          "NO", "",      "",   "", "cashew, almond, dates, seeds, raisin"),
            ("capacity_kg_per_hr", "NUMERIC(15,3)", "NO", "",      "",   "", ""),
            ("created_at",         "TIMESTAMPTZ",   "NO", "NOW()", "",   "", ""),
            ("UNIQUE",             "constraint",    "",   "",      "UQ", "", "(machine_id, stage, item_group)"),
        ],
    },
    {
        "name": "bom_header",
        "category": "Master Data",
        "purpose": "BOM header per FG / customer / pack combination.",
        "columns": [
            ("bom_id",         "SERIAL",        "NO",  "auto",  "PK", "", ""),
            ("fg_sku_name",    "TEXT",          "NO",  "",      "",   "", "finished good name"),
            ("customer_name",  "TEXT",          "YES", "",      "",   "", "NULL = generic BOM"),
            ("pack_size_kg",   "NUMERIC(15,3)", "YES", "",      "",   "", "0.25, 0.5, 1.0..."),
            ("version",        "INT",           "NO",  "1",     "",   "", ""),
            ("is_active",      "BOOLEAN",       "NO",  "TRUE",  "",   "", ""),
            ("effective_from", "DATE",          "YES", "",      "",   "", ""),
            ("effective_to",   "DATE",          "YES", "",      "",   "", "NULL = no expiry"),
            ("item_group",     "TEXT",          "YES", "",      "",   "", "matched from all_sku"),
            ("entity",         "TEXT",          "YES", "",      "",   "", "CHECK in (cfpl, cdpl)"),
            ("notes",          "TEXT",          "YES", "",      "",   "", ""),
            ("created_at",     "TIMESTAMPTZ",   "NO",  "NOW()", "",   "", ""),
        ],
    },
    {
        "name": "bom_line",
        "category": "Master Data",
        "purpose": "BOM material lines (RM + PM).",
        "columns": [
            ("bom_line_id",       "SERIAL",        "NO",  "auto",  "PK", "", ""),
            ("bom_id",            "INT",           "NO",  "",      "FK", "bom_header(bom_id)", ""),
            ("line_number",       "INT",           "NO",  "",      "",   "", ""),
            ("material_sku_name", "TEXT",          "NO",  "",      "",   "", ""),
            ("item_type",         "TEXT",          "NO",  "",      "",   "", "rm, pm"),
            ("quantity_per_unit", "NUMERIC(15,3)", "NO",  "",      "",   "", "kg or pcs per 1 unit FG"),
            ("uom",               "TEXT",          "YES", "",      "",   "", "kg, pcs"),
            ("loss_pct",          "NUMERIC(5,3)",  "YES", "0",     "",   "", "expected loss %"),
            ("godown",            "TEXT",          "YES", "",      "",   "", "RM Store, PM Store"),
            ("can_use_offgrade", "BOOLEAN",       "YES", "FALSE", "",   "", ""),
            ("offgrade_max_pct", "NUMERIC(5,3)",  "YES", "0",     "",   "", "max substitution %"),
            ("created_at",        "TIMESTAMPTZ",   "NO",  "NOW()", "",   "", ""),
            ("UNIQUE",            "constraint",    "",    "",      "UQ", "", "(bom_id, line_number)"),
        ],
    },
    {
        "name": "bom_process_route",
        "category": "Master Data",
        "purpose": "Sequential manufacturing steps for a BOM.",
        "columns": [
            ("route_id",     "SERIAL",        "NO",  "auto",  "PK", "", ""),
            ("bom_id",       "INT",           "NO",  "",      "FK", "bom_header(bom_id)", ""),
            ("step_number",  "INT",           "NO",  "",      "",   "", ""),
            ("process_name", "TEXT",          "NO",  "",      "",   "", "Sorting, Weighing, Sealing, Metal Detection"),
            ("stage",        "TEXT",          "NO",  "",      "",   "", "sorting, weighing, sealing, metal_detection"),
            ("std_time_min", "NUMERIC(10,2)", "YES", "",      "",   "", "standard time in minutes"),
            ("loss_pct",     "NUMERIC(5,3)",  "YES", "0",     "",   "", ""),
            ("qc_check",     "TEXT",          "YES", "",      "",   "", "visual+FM, net weight +/-2g, etc."),
            ("machine_type", "TEXT",          "YES", "",      "",   "", "type of machine needed"),
            ("created_at",   "TIMESTAMPTZ",   "NO",  "NOW()", "",   "", ""),
            ("UNIQUE",       "constraint",    "",    "",      "UQ", "", "(bom_id, step_number)"),
        ],
    },
    {
        "name": "so_fulfillment",
        "category": "Demand Input",
        "purpose": "Bridges SO module to production - pending qty per SO line per FY.",
        "columns": [
            ("fulfillment_id",       "SERIAL",        "NO",  "auto",   "PK", "", ""),
            ("so_line_id",           "INT",           "NO",  "",       "",   "", "FK to so_line.so_line_id (logical)"),
            ("so_id",                "INT",           "YES", "",       "",   "", "FK to so_header.so_id (logical)"),
            ("financial_year",       "TEXT",          "NO",  "",       "",   "", "'2025-26'"),
            ("fg_sku_name",          "TEXT",          "NO",  "",       "",   "", ""),
            ("customer_name",        "TEXT",          "YES", "",       "",   "", ""),
            ("original_qty_kg",      "NUMERIC(15,3)", "NO",  "",       "",   "", ""),
            ("revised_qty_kg",       "NUMERIC(15,3)", "YES", "",       "",   "", ""),
            ("pending_qty_kg",       "NUMERIC(15,3)", "NO",  "",       "",   "", ""),
            ("produced_qty_kg",      "NUMERIC(15,3)", "YES", "0",      "",   "", ""),
            ("dispatched_qty_kg",    "NUMERIC(15,3)", "YES", "0",      "",   "", ""),
            ("order_status",         "TEXT",          "NO",  "'open'", "",   "", "open, partial, fulfilled, carryforward, cancelled"),
            ("delivery_deadline",    "DATE",          "YES", "",       "",   "", ""),
            ("priority",             "INT",           "YES", "5",      "",   "", "1=highest, 10=lowest"),
            ("carryforward_from_id", "INT",           "YES", "",       "FK", "so_fulfillment(fulfillment_id)", "self-ref"),
            ("entity",               "TEXT",          "YES", "",       "",   "", "CHECK in (cfpl, cdpl)"),
            ("created_at",           "TIMESTAMPTZ",   "NO",  "NOW()",  "",   "", ""),
            ("updated_at",           "TIMESTAMPTZ",   "NO",  "NOW()",  "",   "", ""),
            ("UNIQUE",               "constraint",    "",    "",       "UQ", "", "(so_line_id, financial_year)"),
        ],
    },
    {
        "name": "so_revision_log",
        "category": "Demand Input",
        "purpose": "Audit log of SO fulfillment revisions.",
        "columns": [
            ("revision_id",    "SERIAL",      "NO", "auto",  "PK", "", ""),
            ("fulfillment_id", "INT",         "NO", "",      "FK", "so_fulfillment(fulfillment_id)", ""),
            ("revision_type",  "TEXT",        "NO", "",      "",   "", "qty_change, date_change, carryforward, cancel"),
            ("old_value",      "TEXT",        "YES","",      "",   "", ""),
            ("new_value",      "TEXT",        "YES","",      "",   "", ""),
            ("reason",         "TEXT",        "YES","",      "",   "", ""),
            ("revised_by",     "TEXT",        "YES","",      "",   "", ""),
            ("revised_at",     "TIMESTAMPTZ", "NO", "NOW()", "",   "", ""),
        ],
    },
    {
        "name": "production_plan",
        "category": "Core Planning",
        "purpose": "Plan header - one per planning cycle / entity.",
        "columns": [
            ("plan_id",          "SERIAL",      "NO",  "auto",    "PK", "", ""),
            ("plan_name",        "TEXT",        "YES", "",        "",   "", ""),
            ("entity",           "TEXT",        "YES", "",        "",   "", "CHECK in (cfpl, cdpl)"),
            ("plan_type",        "TEXT",        "NO",  "'daily'", "",   "", "daily, weekly, full"),
            ("plan_date",        "DATE",        "NO",  "",        "",   "", "day this plan was created"),
            ("date_from",        "DATE",        "NO",  "",        "",   "", "plan covers from"),
            ("date_to",          "DATE",        "NO",  "",        "",   "", "plan covers to"),
            ("status",           "TEXT",        "NO",  "'draft'", "",   "", "draft, approved, executed, cancelled"),
            ("ai_generated",     "BOOLEAN",     "YES", "FALSE",   "",   "", ""),
            ("ai_analysis_json", "JSONB",       "YES", "",        "",   "", "full Claude response"),
            ("revision_number",  "INT",         "YES", "1",       "",   "", ""),
            ("previous_plan_id", "INT",         "YES", "",        "FK", "production_plan(plan_id)", "self-ref (revision chain)"),
            ("approved_by",      "TEXT",        "YES", "",        "",   "", ""),
            ("approved_at",      "TIMESTAMPTZ", "YES", "",        "",   "", ""),
            ("created_at",       "TIMESTAMPTZ", "NO",  "NOW()",   "",   "", ""),
        ],
    },
    {
        "name": "production_plan_line",
        "category": "Core Planning",
        "purpose": "Per-SKU plan lines (qty, machine, stage sequence, linked SO lines).",
        "columns": [
            ("plan_line_id",              "SERIAL",        "NO",  "auto",      "PK", "", ""),
            ("plan_id",                   "INT",           "NO",  "",          "FK", "production_plan(plan_id)", ""),
            ("fg_sku_name",               "TEXT",          "NO",  "",          "",   "", ""),
            ("customer_name",             "TEXT",          "YES", "",          "",   "", ""),
            ("bom_id",                    "INT",           "YES", "",          "FK", "bom_header(bom_id)", ""),
            ("planned_qty_kg",            "NUMERIC(15,3)", "NO",  "",          "",   "", ""),
            ("planned_qty_units",         "INT",           "YES", "",          "",   "", ""),
            ("machine_id",                "INT",           "YES", "",          "FK", "machine(machine_id)", ""),
            ("priority",                  "INT",           "YES", "5",         "",   "", ""),
            ("shift",                     "TEXT",          "YES", "",          "",   "", "day, night"),
            ("stage_sequence",            "TEXT[]",        "YES", "",          "",   "", "{sorting,weighing,sealing,metal_detection}"),
            ("estimated_hours",           "NUMERIC(10,2)", "YES", "",          "",   "", ""),
            ("linked_so_fulfillment_ids", "INT[]",         "YES", "",          "",   "", "array of so_fulfillment.fulfillment_id (no DB FK)"),
            ("reasoning",                 "TEXT",          "YES", "",          "",   "", "Claude's reasoning"),
            ("status",                    "TEXT",          "NO",  "'planned'", "",   "", "planned, in_progress, completed, cancelled"),
            ("created_at",                "TIMESTAMPTZ",   "NO",  "NOW()",     "",   "", ""),
        ],
    },
    {
        "name": "production_order",
        "category": "Downstream Output",
        "purpose": "Production order created from an approved plan line.",
        "columns": [
            ("prod_order_id",     "SERIAL",        "NO",  "auto",      "PK", "", ""),
            ("prod_order_number", "TEXT",          "NO",  "",          "UQ", "", "PO-2026-0042"),
            ("plan_line_id",      "INT",           "YES", "",          "FK", "production_plan_line(plan_line_id)", ""),
            ("bom_id",            "INT",           "YES", "",          "FK", "bom_header(bom_id)", ""),
            ("fg_sku_name",       "TEXT",          "NO",  "",          "",   "", ""),
            ("customer_name",     "TEXT",          "YES", "",          "",   "", ""),
            ("batch_number",      "TEXT",          "NO",  "",          "UQ", "", "B2026-042"),
            ("batch_size_kg",     "NUMERIC(15,3)", "NO",  "",          "",   "", ""),
            ("net_wt_per_unit",   "NUMERIC(15,3)", "YES", "",          "",   "", "pack size in kg"),
            ("best_before",       "DATE",          "YES", "",          "",   "", ""),
            ("total_stages",      "INT",           "NO",  "1",         "",   "", ""),
            ("status",            "TEXT",          "NO",  "'created'", "",   "", "created, job_cards_issued, in_progress, completed, cancelled"),
            ("entity",            "TEXT",          "YES", "",          "",   "", "CHECK in (cfpl, cdpl)"),
            ("factory",           "TEXT",          "YES", "",          "",   "", ""),
            ("floor",             "TEXT",          "YES", "",          "",   "", ""),
            ("created_at",        "TIMESTAMPTZ",   "NO",  "NOW()",     "",   "", ""),
        ],
    },
    {
        "name": "purchase_indent",
        "category": "Downstream Output",
        "purpose": "Indent raised when MRP detects material shortage for a plan line.",
        "columns": [
            ("indent_id",         "SERIAL",        "NO",  "auto",     "PK", "", ""),
            ("indent_number",     "TEXT",          "NO",  "",         "UQ", "", "IND-20260401-001"),
            ("material_sku_name", "TEXT",          "NO",  "",         "",   "", ""),
            ("required_qty_kg",   "NUMERIC(15,3)", "NO",  "",         "",   "", ""),
            ("required_by_date",  "DATE",          "YES", "",         "",   "", ""),
            ("priority",          "INT",           "YES", "5",        "",   "", ""),
            ("plan_line_id",      "INT",           "YES", "",         "FK", "production_plan_line(plan_line_id)", ""),
            ("po_reference",      "TEXT",          "YES", "",         "",   "", "linked PO number when created"),
            ("status",            "TEXT",          "NO",  "'raised'", "",   "", "raised, acknowledged, po_created, received, cancelled"),
            ("acknowledged_by",   "TEXT",          "YES", "",         "",   "", ""),
            ("acknowledged_at",   "TIMESTAMPTZ",   "YES", "",         "",   "", ""),
            ("entity",            "TEXT",          "YES", "",         "",   "", "CHECK in (cfpl, cdpl)"),
            ("created_at",        "TIMESTAMPTZ",   "NO",  "NOW()",    "",   "", ""),
        ],
    },
    {
        "name": "ai_recommendation",
        "category": "Core Planning",
        "purpose": "Log of Claude interactions tied to a plan.",
        "columns": [
            ("recommendation_id",   "SERIAL",      "NO",  "auto",        "PK", "", ""),
            ("recommendation_type", "TEXT",        "NO",  "",            "",   "", "daily_plan, weekly_plan, full_plan, revision, loss_anomaly, offgrade_reuse, forecast"),
            ("entity",              "TEXT",        "YES", "",            "",   "", "CHECK in (cfpl, cdpl)"),
            ("prompt_text",         "TEXT",        "YES", "",            "",   "", ""),
            ("response_text",       "TEXT",        "YES", "",            "",   "", ""),
            ("response_json",       "JSONB",       "YES", "",            "",   "", ""),
            ("tokens_used",         "INT",         "YES", "",            "",   "", ""),
            ("latency_ms",          "INT",         "YES", "",            "",   "", ""),
            ("model_used",          "TEXT",        "YES", "",            "",   "", ""),
            ("status",              "TEXT",        "NO",  "'generated'", "",   "", "generated, accepted, rejected"),
            ("feedback",            "TEXT",        "YES", "",            "",   "", ""),
            ("plan_id",             "INT",         "YES", "",            "FK", "production_plan(plan_id)", ""),
            ("created_at",          "TIMESTAMPTZ", "NO",  "NOW()",       "",   "", ""),
        ],
    },
]

RELATIONSHIPS = [
    # (child_table, child_col, parent_table, parent_col, notes)
    ("machine_capacity",     "machine_id",              "machine",              "machine_id",     ""),
    ("bom_line",             "bom_id",                  "bom_header",           "bom_id",         ""),
    ("bom_process_route",    "bom_id",                  "bom_header",           "bom_id",         ""),
    ("so_fulfillment",       "carryforward_from_id",    "so_fulfillment",       "fulfillment_id", "self-ref"),
    ("so_revision_log",      "fulfillment_id",          "so_fulfillment",       "fulfillment_id", ""),
    ("production_plan",      "previous_plan_id",        "production_plan",      "plan_id",        "self-ref (revision chain)"),
    ("production_plan_line", "plan_id",                 "production_plan",      "plan_id",        ""),
    ("production_plan_line", "bom_id",                  "bom_header",           "bom_id",         ""),
    ("production_plan_line", "machine_id",              "machine",              "machine_id",     ""),
    ("production_plan_line", "linked_so_fulfillment_ids","so_fulfillment",      "fulfillment_id", "INT[] array, app-managed, no DB FK"),
    ("production_order",     "plan_line_id",            "production_plan_line", "plan_line_id",   ""),
    ("production_order",     "bom_id",                  "bom_header",           "bom_id",         ""),
    ("purchase_indent",      "plan_line_id",            "production_plan_line", "plan_line_id",   ""),
    ("ai_recommendation",    "plan_id",                 "production_plan",      "plan_id",        ""),
]

# ---------- styling ----------
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CAT_FILLS = {
    "Master Data":       PatternFill("solid", fgColor="DDEBF7"),
    "Demand Input":      PatternFill("solid", fgColor="FFF2CC"),
    "Core Planning":     PatternFill("solid", fgColor="E2EFDA"),
    "Downstream Output": PatternFill("solid", fgColor="FCE4D6"),
}
PK_FONT = Font(bold=True, color="9C0006")
FK_FONT = Font(italic=True, color="1F4E79")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# ---------- Overview ----------
ws = wb.active
ws.title = "Overview"
ws["A1"] = "Production Planning - Schema Overview"
ws["A1"].font = Font(bold=True, size=14)
ws.merge_cells("A1:D1")

ws["A2"] = "Source: app/db/production_schema.sql"
ws["A2"].font = Font(italic=True, color="595959")
ws.merge_cells("A2:D2")

header = ["#", "Table", "Category", "Purpose"]
for c, h in enumerate(header, start=1):
    ws.cell(row=4, column=c, value=h)
style_header_row(ws, 4, len(header))

for i, t in enumerate(TABLES, start=1):
    r = 4 + i
    ws.cell(row=r, column=1, value=i)
    ws.cell(row=r, column=2, value=t["name"]).font = Font(bold=True)
    ws.cell(row=r, column=3, value=t["category"])
    ws.cell(row=r, column=4, value=t["purpose"])
    fill = CAT_FILLS.get(t["category"])
    if fill:
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).border = BORDER

autosize(ws, [5, 26, 20, 80])
ws.freeze_panes = "A5"

# ---------- One sheet per table ----------
col_headers = ["Column", "Type", "Nullable", "Default", "Key", "References", "Notes"]
col_widths  = [30, 18, 10, 14, 8, 38, 60]

for t in TABLES:
    sheet = wb.create_sheet(title=t["name"][:31])
    sheet["A1"] = t["name"]
    sheet["A1"].font = Font(bold=True, size=13)
    sheet.merge_cells("A1:G1")

    sheet["A2"] = f"{t['category']}  -  {t['purpose']}"
    sheet["A2"].font = Font(italic=True, color="595959")
    sheet.merge_cells("A2:G2")

    for c, h in enumerate(col_headers, start=1):
        sheet.cell(row=4, column=c, value=h)
    style_header_row(sheet, 4, len(col_headers))

    for i, col in enumerate(t["columns"], start=1):
        r = 4 + i
        for c, v in enumerate(col, start=1):
            cell = sheet.cell(row=r, column=c, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        key = col[4]
        if key == "PK":
            sheet.cell(row=r, column=1).font = PK_FONT
        elif key == "FK":
            sheet.cell(row=r, column=1).font = FK_FONT
        elif key == "UQ":
            sheet.cell(row=r, column=1).font = Font(italic=True, color="595959")

    autosize(sheet, col_widths)
    sheet.freeze_panes = "A5"

# ---------- Relationships ----------
rel = wb.create_sheet(title="Relationships")
rel["A1"] = "Foreign Keys & Logical Links"
rel["A1"].font = Font(bold=True, size=13)
rel.merge_cells("A1:E1")

rel_headers = ["Child Table", "Child Column", "Parent Table", "Parent Column", "Notes"]
for c, h in enumerate(rel_headers, start=1):
    rel.cell(row=3, column=c, value=h)
style_header_row(rel, 3, len(rel_headers))

for i, (ct, cc, pt, pc, n) in enumerate(RELATIONSHIPS, start=1):
    r = 3 + i
    for c, v in enumerate([ct, cc, pt, pc, n], start=1):
        cell = rel.cell(row=r, column=c, value=v)
        cell.border = BORDER

autosize(rel, [24, 28, 24, 18, 50])
rel.freeze_panes = "A4"

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"Wrote: {OUT}")
print(f"Sheets: {wb.sheetnames}")
