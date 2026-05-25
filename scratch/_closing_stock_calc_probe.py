"""Basic-level closing-stock calculation per entity (CFPL/CDPL) + company total.

Formula (per entity, per item_group):
    closing = opening
            + inward                  (PO receipts)
            - transfer_out + transfer_in  (cross-entity only; intra-entity nets to 0)
            - bom_consumption         (RM & PM issued to job cards / Tally voucher RM)
            + cn_inward               (credit notes — customer returns coming back in)
            - outward                 (sales dispatches)

Sources:
    Opening                  Excel  Candor Physical Stock Compilation 31-03-2026.xlsx / 'Compiled'
    Inward                   DB     cfpl_boxes_v2 / cdpl_boxes_v2 ⨝ cfpl_transactions_v2 / cdpl_transactions_v2
                                    (filter: rtv=false, service=false, entry_date >= 2026-04-01)
    Transfers                DB     interunit_transfers_lines ⨝ interunit_transfers_header
                                    (entity derived from from_site/to_site mapping)
    BOM Consumption          Excel  RM_Consumption_Apr2026_CFPL.xlsx / 'Group Summary' (Net Issued KG)
                                    — CFPL only; CDPL = 0 (no file provided)
    CN Inward (RTV from CN)  Excel  Sales Register …/ 'Cancel Inv' sheets (both files)
    Outward                  Excel  Sales Register …/ 'Sales Report' sheets (both files)

Article→Group resolution uses all_sku.particulars (case-insensitive); unmatched rows
are bucketed into '_UNMAPPED' so they remain visible.

Window: 2026-04-01 → 2026-05-12 (last sales register date). Sales register only goes to
May 12, so closing stock is effective May 12 — not today (May 18).

Usage:
    cd d:/Consumption/New/Backend
    python _closing_stock_calc_probe.py
"""
import asyncio
import os
import re
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path

import asyncpg
import openpyxl
from dotenv import load_dotenv

# Force UTF-8 on Windows console so dashes/etc print
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

load_dotenv()
DB_URL = os.environ["DATABASE_URL"]

XLSX_DIR = Path(r"C:\Users\cando\Downloads\Inventory calc")
OPENING_XLSX = XLSX_DIR / "Candor Physical Stock Compilation 31-03-2026.xlsx"
RMCONS_XLSX = XLSX_DIR / "RM_Consumption_Apr2026_CFPL.xlsx"
SALES_XLSX_APR = XLSX_DIR / "Sales Register 30th April 2026.xlsx"
SALES_XLSX_MAY = XLSX_DIR / "Sales Register 12th May 2026.xlsx"

WINDOW_START = date(2026, 4, 1)

# Site → entity mapping (for inter-unit transfer entity derivation)
CFPL_SITES = {"W202", "A68", "F53", "RISHI", "A101"}
CDPL_SITES = {"A185", "COLD STORAGE", "SAVLA D-39", "SAVLA D-514", "PAWANE"}


def site_to_entity(site: str | None) -> str | None:
    if not site:
        return None
    s = str(site).strip().upper()
    if s in CFPL_SITES:
        return "CFPL"
    if s in CDPL_SITES:
        return "CDPL"
    return None  # unknown — log


def norm_group(g: str | None) -> str:
    if not g:
        return "_UNMAPPED"
    return str(g).strip().lower()


def norm_sku(s: str | None) -> str:
    if not s:
        return ""
    # Lowercase, collapse whitespace, strip non-alnum from edges
    return re.sub(r"\s+", " ", str(s).strip().lower())


# Storage type: defaultdict(lambda: defaultdict(float))
# bucket[entity][group] = qty_kg
def new_bucket():
    return defaultdict(lambda: defaultdict(float))


# ──────────────────────────────────────────────────────────────────────────
# 1. OPENING STOCK (Excel)
# ──────────────────────────────────────────────────────────────────────────

def load_opening() -> dict:
    wb = openpyxl.load_workbook(OPENING_XLSX, read_only=True, data_only=True)
    ws = wb["Compiled"]
    bucket = new_bucket()
    for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True)):
        company = row[0]
        group = row[4]
        total_kg = row[14]
        if not company or not group or total_kg is None:
            continue
        try:
            kg = float(total_kg)
        except (TypeError, ValueError):
            continue
        entity = "CFPL" if str(company).strip().upper() == "CFPL" else "CDPL"
        bucket[entity][norm_group(group)] += kg
    wb.close()
    return bucket


# ──────────────────────────────────────────────────────────────────────────
# 2. INWARD (DB)
# ──────────────────────────────────────────────────────────────────────────

async def load_inward(conn, sku_to_group: dict) -> tuple[dict, dict]:
    bucket = new_bucket()
    unmapped = defaultdict(float)
    for entity, prefix in [("CFPL", "cfpl"), ("CDPL", "cdpl")]:
        rows = await conn.fetch(
            f"""
            SELECT b.article_description AS sku, COALESCE(b.net_weight, 0) AS kg
            FROM {prefix}_boxes_v2 b
            JOIN {prefix}_transactions_v2 t ON b.transaction_no = t.transaction_no
            WHERE t.entry_date >= $1::date
              AND COALESCE(t.rtv, false) = false
              AND COALESCE(t.service, false) = false
            """,
            WINDOW_START,
        )
        for r in rows:
            kg = float(r["kg"] or 0)
            if kg <= 0:
                continue
            g = sku_to_group.get(norm_sku(r["sku"]))
            if g is None:
                unmapped[r["sku"]] += kg
                g = "_UNMAPPED"
            bucket[entity][g] += kg
    return bucket, dict(unmapped)


# ──────────────────────────────────────────────────────────────────────────
# 3a. TRANSFER OUT (DB)  — uses interunit_transfers_lines
# 3b. TRANSFER IN  (DB)  — same source: every cross-entity OUT = IN for receiver
# Only cross-entity transfers move stock at the entity level. Intra-entity nets to 0.
# ──────────────────────────────────────────────────────────────────────────

async def load_transfers(conn, sku_to_group: dict) -> tuple[dict, dict, list]:
    out_b = new_bucket()
    in_b = new_bucket()
    unknown_sites = []

    rows = await conn.fetch(
        """
        SELECT h.from_site, h.to_site, h.stock_trf_date,
               l.item_category, l.item_desc_raw, COALESCE(l.net_weight, 0) AS kg
        FROM interunit_transfers_lines l
        JOIN interunit_transfers_header h ON l.header_id = h.id
        WHERE h.stock_trf_date >= $1::date
        """,
        WINDOW_START,
    )

    for r in rows:
        kg = float(r["kg"] or 0)
        if kg <= 0:
            continue
        from_e = site_to_entity(r["from_site"])
        to_e = site_to_entity(r["to_site"])
        if from_e is None or to_e is None:
            unknown_sites.append((r["from_site"], r["to_site"]))
            continue
        if from_e == to_e:
            continue  # intra-entity → net 0

        # Resolve group: prefer item_category (it's the actual group name like 'ALMOND'),
        # fall back to all_sku lookup by item_desc_raw
        g = norm_group(r["item_category"]) if r["item_category"] else None
        if not g or g == "_UNMAPPED":
            g = sku_to_group.get(norm_sku(r["item_desc_raw"]), "_UNMAPPED")

        out_b[from_e][g] += kg
        in_b[to_e][g] += kg

    return out_b, in_b, unknown_sites


# ──────────────────────────────────────────────────────────────────────────
# 4. BOM CONSUMPTION (Excel — Tally voucher rollup)
# ──────────────────────────────────────────────────────────────────────────

def load_bom_consumption() -> dict:
    bucket = new_bucket()
    wb = openpyxl.load_workbook(RMCONS_XLSX, read_only=True, data_only=True)
    ws = wb["Group Summary"]
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        group = row[0]
        net_issued = row[3]
        if not group or net_issued is None:
            continue
        # Skip footer/total rows that appear at bottom of the sheet
        g_lower = str(group).strip().lower()
        if g_lower.startswith("grand total") or g_lower.startswith("total"):
            continue
        try:
            kg = float(net_issued)
        except (TypeError, ValueError):
            continue
        bucket["CFPL"][norm_group(group)] += kg
    wb.close()
    return bucket


# ──────────────────────────────────────────────────────────────────────────
# 5+6+7. SALES REGISTER — Sales Report (outward) + Cancel Inv (CN inward)
# ──────────────────────────────────────────────────────────────────────────

def _read_sales_sheet(path: Path, sheet: str, sku_to_group: dict,
                       bucket: dict, unmapped: dict, header_row: int = 2):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return
    ws = wb[sheet]
    # Discover columns from header row (row index 1 = second row, since rows start at 0)
    # Cols of interest (1-indexed in our exploration): Date(0), Article(4), Company01(10), In Kg(17)
    # We'll just read by fixed index per the structure seen.
    for i, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True)):
        if not row or len(row) < 18:
            continue
        date = row[0]
        article = row[4]
        company = row[10]
        in_kg = row[17]
        if date is None or article is None or in_kg is None:
            continue
        comp_s = str(company or "").strip().upper()
        entity = "CFPL" if comp_s == "CFPL" else ("CDPL" if comp_s == "CDPL" else None)
        if entity is None:
            continue  # skip junk rows with invoice no in Company01
        try:
            kg = float(in_kg)
        except (TypeError, ValueError):
            continue
        g = sku_to_group.get(norm_sku(article))
        if g is None:
            unmapped[article] += abs(kg)
            g = "_UNMAPPED"
        bucket[entity][g] += abs(kg)
    wb.close()


def load_outward(sku_to_group: dict) -> tuple[dict, dict]:
    bucket = new_bucket()
    unmapped: dict = defaultdict(float)
    for p in [SALES_XLSX_APR, SALES_XLSX_MAY]:
        _read_sales_sheet(p, "Sales Report", sku_to_group, bucket, unmapped)
    return bucket, dict(unmapped)


def load_cn_inward(sku_to_group: dict) -> tuple[dict, dict]:
    bucket = new_bucket()
    unmapped: dict = defaultdict(float)
    for p in [SALES_XLSX_APR, SALES_XLSX_MAY]:
        _read_sales_sheet(p, "Cancel Inv", sku_to_group, bucket, unmapped)
    return bucket, dict(unmapped)


# ──────────────────────────────────────────────────────────────────────────
# all_sku lookup
# ──────────────────────────────────────────────────────────────────────────

async def load_sku_lookup(conn) -> dict:
    rows = await conn.fetch("SELECT particulars, item_group FROM all_sku")
    return {norm_sku(r["particulars"]): norm_group(r["item_group"]) for r in rows if r["particulars"]}


# ──────────────────────────────────────────────────────────────────────────
# Output
# ──────────────────────────────────────────────────────────────────────────

def collect_groups(*buckets) -> list[str]:
    s: set = set()
    for b in buckets:
        for ent in b:
            for g in b[ent]:
                s.add(g)
    return sorted(s)


def fmt(v: float) -> str:
    if v == 0:
        return "0"
    return f"{v:,.0f}"


def print_entity_table(entity: str, groups: list[str], opening, inward, t_out, t_in,
                       bom, cn_in, outward, closing) -> None:
    print(f"\n{'═'*120}")
    print(f"  {entity}  —  Stock movement Apr 1, 2026 → May 12, 2026 (all qty in KG)")
    print(f"{'═'*120}")
    hdr = (
        f"{'GROUP':<22} {'Opening':>12} {'Inward':>12} {'-TxOut':>10} {'+TxIn':>10} "
        f"{'-BOM':>12} {'+CN-In':>10} {'-Outward':>12} {'= CLOSING':>14}"
    )
    print(hdr)
    print("-" * len(hdr))
    totals = [0.0] * 8
    for g in groups:
        op = opening[entity].get(g, 0)
        iw = inward[entity].get(g, 0)
        to = t_out[entity].get(g, 0)
        ti = t_in[entity].get(g, 0)
        bm = bom[entity].get(g, 0)
        cn = cn_in[entity].get(g, 0)
        ow = outward[entity].get(g, 0)
        cl = closing[entity].get(g, 0)
        if op == iw == to == ti == bm == cn == ow == cl == 0:
            continue
        print(f"{g[:22]:<22} {fmt(op):>12} {fmt(iw):>12} {fmt(to):>10} {fmt(ti):>10} "
              f"{fmt(bm):>12} {fmt(cn):>10} {fmt(ow):>12} {fmt(cl):>14}")
        for idx, val in enumerate([op, iw, to, ti, bm, cn, ow, cl]):
            totals[idx] += val
    print("-" * len(hdr))
    print(f"{'TOTAL':<22} {fmt(totals[0]):>12} {fmt(totals[1]):>12} {fmt(totals[2]):>10} "
          f"{fmt(totals[3]):>10} {fmt(totals[4]):>12} {fmt(totals[5]):>10} "
          f"{fmt(totals[6]):>12} {fmt(totals[7]):>14}")


def print_company_table(groups: list[str], opening, inward, t_out, t_in, bom, cn_in,
                         outward, closing) -> None:
    print(f"\n{'═'*120}")
    print(f"  COMPANY TOTAL (CFPL + CDPL)  —  Apr 1, 2026 → May 12, 2026 (all qty in KG)")
    print(f"{'═'*120}")
    hdr = (
        f"{'GROUP':<22} {'Opening':>12} {'Inward':>12} {'NetTxfr':>10} "
        f"{'-BOM':>12} {'+CN-In':>10} {'-Outward':>12} {'= CLOSING':>14}"
    )
    print(hdr)
    print("-" * len(hdr))
    totals = [0.0] * 7
    for g in groups:
        op = sum(opening[e].get(g, 0) for e in ["CFPL", "CDPL"])
        iw = sum(inward[e].get(g, 0) for e in ["CFPL", "CDPL"])
        # Net transfer at company level: CFPL→CDPL and CDPL→CFPL cancel in aggregate
        net_t = sum(t_in[e].get(g, 0) - t_out[e].get(g, 0) for e in ["CFPL", "CDPL"])
        bm = sum(bom[e].get(g, 0) for e in ["CFPL", "CDPL"])
        cn = sum(cn_in[e].get(g, 0) for e in ["CFPL", "CDPL"])
        ow = sum(outward[e].get(g, 0) for e in ["CFPL", "CDPL"])
        cl = sum(closing[e].get(g, 0) for e in ["CFPL", "CDPL"])
        if op == iw == net_t == bm == cn == ow == cl == 0:
            continue
        print(f"{g[:22]:<22} {fmt(op):>12} {fmt(iw):>12} {fmt(net_t):>10} "
              f"{fmt(bm):>12} {fmt(cn):>10} {fmt(ow):>12} {fmt(cl):>14}")
        for idx, val in enumerate([op, iw, net_t, bm, cn, ow, cl]):
            totals[idx] += val
    print("-" * len(hdr))
    print(f"{'TOTAL':<22} {fmt(totals[0]):>12} {fmt(totals[1]):>12} {fmt(totals[2]):>10} "
          f"{fmt(totals[3]):>12} {fmt(totals[4]):>10} {fmt(totals[5]):>12} "
          f"{fmt(totals[6]):>14}")


# ──────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────

async def main() -> None:
    print("Loading opening stock (Excel)...")
    opening = load_opening()
    print(f"  CFPL opening groups: {len(opening['CFPL'])} | CDPL: {len(opening['CDPL'])}")

    print("Loading BOM consumption (Excel — CFPL only)...")
    bom = load_bom_consumption()
    print(f"  CFPL BOM groups: {len(bom['CFPL'])}")

    pool = await asyncpg.create_pool(DB_URL, min_size=1, max_size=2)
    try:
        async with pool.acquire() as c:
            print("Loading all_sku lookup...")
            sku_to_group = await load_sku_lookup(c)
            print(f"  SKUs indexed: {len(sku_to_group)}")

            print("Loading inward (DB)...")
            inward, inward_unmapped = await load_inward(c, sku_to_group)
            print(f"  CFPL inward groups: {len(inward['CFPL'])} | "
                  f"CDPL: {len(inward['CDPL'])} | unmapped SKUs: {len(inward_unmapped)}")

            print("Loading transfers (DB)...")
            t_out, t_in, unknown_sites = await load_transfers(c, sku_to_group)
            print(f"  cross-entity OUT groups CFPL: {len(t_out['CFPL'])}, CDPL: {len(t_out['CDPL'])} | "
                  f"unknown_site rows: {len(unknown_sites)}")
    finally:
        await pool.close()

    print("Loading sales register (Excel)...")
    outward, outward_unmapped = load_outward(sku_to_group)
    cn_in, cn_unmapped = load_cn_inward(sku_to_group)
    print(f"  Outward groups CFPL: {len(outward['CFPL'])}, CDPL: {len(outward['CDPL'])} | "
          f"unmapped SKUs: {len(outward_unmapped)}")
    print(f"  CN-IN groups CFPL: {len(cn_in['CFPL'])}, CDPL: {len(cn_in['CDPL'])}")

    # Compute closing per entity per group
    closing = new_bucket()
    for entity in ["CFPL", "CDPL"]:
        all_g = set()
        for b in [opening, inward, t_out, t_in, bom, cn_in, outward]:
            all_g.update(b[entity].keys())
        for g in all_g:
            closing[entity][g] = (
                opening[entity].get(g, 0)
                + inward[entity].get(g, 0)
                - t_out[entity].get(g, 0)
                + t_in[entity].get(g, 0)
                - bom[entity].get(g, 0)
                + cn_in[entity].get(g, 0)
                - outward[entity].get(g, 0)
            )

    groups = collect_groups(opening, inward, t_out, t_in, bom, cn_in, outward, closing)

    print_entity_table("CFPL", groups, opening, inward, t_out, t_in, bom, cn_in, outward, closing)
    print_entity_table("CDPL", groups, opening, inward, t_out, t_in, bom, cn_in, outward, closing)
    print_company_table(groups, opening, inward, t_out, t_in, bom, cn_in, outward, closing)

    # Top unmapped SKUs (so user can see which Articles fell through all_sku lookup)
    def top_unmapped(label: str, d: dict, n: int = 10):
        if not d:
            return
        items = sorted(d.items(), key=lambda x: x[1], reverse=True)[:n]
        print(f"\n  Top {n} unmapped SKUs in {label}:")
        for sku, kg in items:
            print(f"    {fmt(kg):>10} kg   {sku}")

    print("\n" + "─" * 60)
    print("  DATA-QUALITY NOTES")
    print("─" * 60)
    top_unmapped("inward", inward_unmapped)
    top_unmapped("outward", outward_unmapped)
    top_unmapped("CN-IN", cn_unmapped)
    if unknown_sites:
        unk = defaultdict(int)
        for fs, ts in unknown_sites:
            unk[(fs, ts)] += 1
        print(f"\n  Transfer rows with unknown site mapping: {len(unknown_sites)}")
        for (fs, ts), cnt in sorted(unk.items(), key=lambda x: -x[1])[:10]:
            print(f"    {cnt:>4}  from={fs!r:<25}  to={ts!r}")


if __name__ == "__main__":
    asyncio.run(main())
