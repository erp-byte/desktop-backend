"""Convert scratch/_closing_v5_report.html → supporting Excel workbook.

Sheets produced:
  - Summary            (per-view opening/inward/outward/closing totals)
  - Consolidated, W-202, A-185, F-53, A-68, Rishi, Savla, Other
                       (one sheet per view — closing-stock table + TOTAL row)
  - Sale_Classification (pill counts from the bottom card)
  - Intercompany_Digest (group × entity digest of stock-transfer invoices)
  - Intercompany_Lines  (individual stock-transfer invoice lines)

All numeric cells are parsed back to numbers (with proper signs / negatives).
The HTML tooltips (title="...") are dropped — the formula header text is enough.
"""
from __future__ import annotations

import html
import re
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
SRC = HERE / "_closing_v5_report.html"
OUT = HERE / "_closing_v5_report.xlsx"

# View id → friendly title
VIEW_TITLES = {
    "total": "Consolidated",
    "w202":  "W-202",
    "a185":  "A-185",
    "f53":   "F-53 (APMC)",
    "a68":   "A-68",
    "rishi": "Rishi",
    "savla": "Savla",
    "other": "Other",
}

CLOSING_COLS = [
    "Group", "Opening", "+ Inward", "+ Inf-In", "+ Xfr-In", "− Xfr-Out",
    "+ CN", "− Outward", "+ Adj-In", "− Adj-Out", "± Rearr", "= Closing",
]


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

NUM_RE = re.compile(r"^-?[\d,]+$")


def to_number(s: str):
    """Convert '12,930' / '-11,398' / '0' → int; everything else returned as-is."""
    s = s.strip()
    if not s:
        return None
    if NUM_RE.match(s):
        try:
            return int(s.replace(",", ""))
        except ValueError:
            pass
    return s


# ---------------------------------------------------------------------------
# Closing-stock table parser — one row per view, tracks the current <div view>
# ---------------------------------------------------------------------------

class ClosingParser(HTMLParser):
    """Extract per-view closing-stock tables.

    Result: self.views = {view_id: [ [group, opening, ...closing], ... ]}
    The last row in each list is the TOTAL row.
    """

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.views: dict[str, list[list]] = {}
        self.cur_view: str | None = None
        self.in_view_table = False
        self.in_thead = False
        self.in_tbody = False
        self.in_tr = False
        self.in_td = False
        self.in_th = False
        self.cur_row: list[str] = []
        self.cur_cell: list[str] = []

    def handle_starttag(self, tag, attrs):
        a = dict(attrs)
        if tag == "div":
            # match <div class="view ..." id='view-xxx'>
            cls = a.get("class") or ""
            view_id = a.get("id") or ""
            if "view" in cls.split() and view_id.startswith("view-"):
                vid = view_id.removeprefix("view-")
                if vid in VIEW_TITLES:
                    self.cur_view = vid
                    self.views.setdefault(vid, [])
        elif tag == "table" and self.cur_view:
            self.in_view_table = True
        elif tag == "thead" and self.in_view_table:
            self.in_thead = True
        elif tag == "tbody" and self.in_view_table:
            self.in_tbody = True
        elif tag == "tr" and self.in_view_table and self.in_tbody:
            self.in_tr = True
            self.cur_row = []
        elif tag == "td" and self.in_tr:
            self.in_td = True
            self.cur_cell = []
        elif tag == "th" and self.in_thead:
            self.in_th = True

    def handle_endtag(self, tag):
        if tag == "td" and self.in_td:
            # The first-line text of the cell is the number we want
            # (the cell may contain a Top-N tooltip — but that's in title=,
            #  which is an attribute, so handle_data only sees the visible text).
            txt = "".join(self.cur_cell).strip()
            # Visible text is just the number; collapse whitespace
            txt = re.sub(r"\s+", " ", txt).strip()
            self.cur_row.append(txt)
            self.in_td = False
            self.cur_cell = []
        elif tag == "th" and self.in_th:
            self.in_th = False
        elif tag == "tr" and self.in_tr:
            if self.cur_row:
                self.views[self.cur_view].append(self.cur_row)
            self.in_tr = False
            self.cur_row = []
        elif tag == "thead":
            self.in_thead = False
        elif tag == "tbody":
            self.in_tbody = False
        elif tag == "table" and self.in_view_table:
            self.in_view_table = False
        elif tag == "div" and self.cur_view and not self.in_view_table:
            # Leaving the view container — only reset when we left its table.
            # HTML structure: <div view><table>...</table></div>; once </table>
            # has fired we treat the next </div> as the view boundary.
            self.cur_view = None

    def handle_data(self, data):
        if self.in_td:
            self.cur_cell.append(data)


# ---------------------------------------------------------------------------
# Top-level summary parsed from inline JS map
# ---------------------------------------------------------------------------

def parse_summaries(text: str) -> dict[str, dict[str, str]]:
    m = re.search(r"const\s+summaries\s*=\s*(\{[^;]+\});", text)
    if not m:
        return {}
    blob = m.group(1)
    # Each entry: key: {closing:'…', opening:'…', inward:'…', outward:'…'}
    out: dict[str, dict[str, str]] = {}
    for key, body in re.findall(r"(\w+)\s*:\s*\{([^}]+)\}", blob):
        d: dict[str, str] = {}
        for k, v in re.findall(r"(\w+)\s*:\s*'([^']*)'", body):
            d[k] = v
        if d:
            out[key] = d
    return out


# ---------------------------------------------------------------------------
# Sale-classification pills + footer counters
# ---------------------------------------------------------------------------

def parse_sale_class(text: str) -> list[tuple[str, str]]:
    """Return list of (label, count_or_text) tuples from the bottom card."""
    rows: list[tuple[str, str]] = []
    pill_re = re.compile(
        r"<span class='pill'[^>]*>([^<]*?):\s*([\d,]+)</span>"
    )
    for label, count in pill_re.findall(text):
        rows.append((html.unescape(label.strip()), count.strip()))
    # Sales / CN counters
    footer_re = re.compile(
        r"Sales lines:\s*([\d,]+).*?CN from Excel:\s*([\d,]+).*?CN from PDFs:\s*([\d,]+)",
        re.DOTALL,
    )
    fm = footer_re.search(text)
    if fm:
        rows.append(("Sales lines total", fm.group(1)))
        rows.append(("CN from Excel", fm.group(2)))
        rows.append(("CN from PDFs (deduped by doc_no+product)", fm.group(3)))
    return rows


# ---------------------------------------------------------------------------
# Inter-company section
# ---------------------------------------------------------------------------

class SectionTableParser(HTMLParser):
    """Pull every <table>…</table> inside a single <section class='card'> block,
    flattened to a list of (header_row, data_rows)."""

    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.tables: list[tuple[list[str], list[list[str]]]] = []
        self.in_table = False
        self.in_thead = False
        self.in_tbody = False
        self.in_tr = False
        self.in_td_or_th = False
        self.header: list[str] = []
        self.rows: list[list[str]] = []
        self.cur_row: list[str] = []
        self.cell: list[str] = []

    def handle_starttag(self, tag, attrs):
        if tag == "table":
            self.in_table = True
            self.header = []
            self.rows = []
        elif tag == "thead" and self.in_table:
            self.in_thead = True
        elif tag == "tbody" and self.in_table:
            self.in_tbody = True
        elif tag == "tr" and self.in_table:
            self.in_tr = True
            self.cur_row = []
        elif tag in ("td", "th") and self.in_tr:
            self.in_td_or_th = True
            self.cell = []

    def handle_endtag(self, tag):
        if tag in ("td", "th") and self.in_td_or_th:
            txt = "".join(self.cell).strip()
            txt = re.sub(r"\s+", " ", txt)
            if self.in_thead:
                self.header.append(txt)
            else:
                self.cur_row.append(txt)
            self.in_td_or_th = False
            self.cell = []
        elif tag == "tr" and self.in_tr:
            if self.cur_row and self.in_tbody:
                self.rows.append(self.cur_row)
            self.in_tr = False
            self.cur_row = []
        elif tag == "thead":
            self.in_thead = False
        elif tag == "tbody":
            self.in_tbody = False
        elif tag == "table" and self.in_table:
            self.tables.append((self.header, self.rows))
            self.in_table = False

    def handle_data(self, data):
        if self.in_td_or_th:
            self.cell.append(data)


def extract_intercompany_section(text: str) -> str:
    """Slice out the inter-company <section> block."""
    m = re.search(
        r"<section[^>]*>\s*<h2>Inter-company.*?</section>",
        text,
        re.DOTALL,
    )
    return m.group(0) if m else ""


# ---------------------------------------------------------------------------
# Header counters from intercompany hint paragraphs
# ---------------------------------------------------------------------------

def parse_intercompany_meta(text: str) -> dict[str, str]:
    m = re.search(r"<b>(\d+) line items</b>, total <b>([\d,]+) kg</b>", text)
    if not m:
        return {}
    return {"line_items": m.group(1), "total_kg": m.group(2)}


# ===========================================================================
# Workbook build
# ===========================================================================

HDR_FILL    = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HDR_FONT    = Font(bold=True, color="FFFFFF")
TOTAL_FILL  = PatternFill(start_color="EEF3F0", end_color="EEF3F0", fill_type="solid")
TOTAL_FONT  = Font(bold=True, color="16513F")
ADJ_FILL    = PatternFill(start_color="FDF6E3", end_color="FDF6E3", fill_type="solid")
NEG_FONT    = Font(color="B6260D")


def style_header(ws, n_cols: int, row: int = 1):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_closing_sheet(wb: Workbook, view_id: str, rows: list[list[str]],
                          summary: dict[str, str] | None):
    title = VIEW_TITLES[view_id]
    # Excel sheet name limit 31
    sheet_name = title[:31]
    ws = wb.create_sheet(title=sheet_name)
    # Title block
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color="16513F")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(CLOSING_COLS))
    if summary:
        ws["A2"] = (
            f"Closing {summary.get('closing','-')} kg  |  "
            f"Opening {summary.get('opening','-')} kg  |  "
            f"Inward {summary.get('inward','-')} kg  |  "
            f"Outward {summary.get('outward','-')} kg"
        )
        ws["A2"].font = Font(italic=True, color="666666")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(CLOSING_COLS))
    # Header row at row 4
    for i, h in enumerate(CLOSING_COLS, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, len(CLOSING_COLS), row=4)
    # Data rows from row 5
    adj_col_idx = {CLOSING_COLS.index(c) + 1 for c in ("+ Adj-In", "− Adj-Out", "± Rearr")}
    closing_col_idx = len(CLOSING_COLS)  # last column
    r = 5
    for raw_row in rows:
        # Each raw_row has 12 strings; first is group, last is closing
        if len(raw_row) != len(CLOSING_COLS):
            # Skip malformed; should not happen
            continue
        is_total = raw_row[0].strip().upper() == "TOTAL"
        for ci, val in enumerate(raw_row, start=1):
            num = to_number(val) if ci > 1 else val.strip()
            cell = ws.cell(row=r, column=ci, value=num)
            if isinstance(num, (int, float)) and num != 0:
                cell.number_format = "#,##0;-#,##0;0"
            if ci in adj_col_idx and not is_total:
                cell.fill = ADJ_FILL
            if isinstance(num, (int, float)) and num < 0:
                cell.font = NEG_FONT
            if is_total:
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL
            cell.alignment = Alignment(
                horizontal="left" if ci == 1 else "right",
                vertical="center",
            )
        r += 1
    # Widths
    widths = [22] + [13] * (len(CLOSING_COLS) - 1)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B5"


def write_summary_sheet(wb: Workbook, summaries: dict[str, dict[str, str]]):
    ws = wb.create_sheet(title="Summary", index=0)
    ws["A1"] = "Closing Stock — Window 2026-04-01 → 2026-05-12"
    ws["A1"].font = Font(bold=True, size=14, color="16513F")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws["A2"] = "All quantities in KG. Source: scratch/_closing_v5_report.html"
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

    headers = ["View", "Opening", "+ Inward", "− Outward", "= Closing"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, len(headers), row=4)
    r = 5
    for view_id, title in VIEW_TITLES.items():
        s = summaries.get(view_id, {})
        ws.cell(row=r, column=1, value=title)
        for ci, key in enumerate(("opening", "inward", "outward", "closing"), start=2):
            v = to_number(s.get(key, ""))
            cell = ws.cell(row=r, column=ci, value=v)
            if isinstance(v, (int, float)):
                cell.number_format = "#,##0;-#,##0;0"
        if view_id == "total":
            for ci in range(1, len(headers) + 1):
                ws.cell(row=r, column=ci).font = TOTAL_FONT
                ws.cell(row=r, column=ci).fill = TOTAL_FILL
        r += 1
    for i, w in enumerate([26, 14, 14, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    # Formula explainer
    ws.cell(row=r + 1, column=1, value="Formula").font = Font(bold=True, color="16513F")
    ws.cell(
        row=r + 2, column=1,
        value=(
            "Closing[group, unit] = Opening + Inward + Inf-In + Xfr-In − Xfr-Out + CN "
            "− Outward + Adj-In − Adj-Out ± Rearr"
        ),
    )
    ws.merge_cells(start_row=r + 2, start_column=1, end_row=r + 2, end_column=5)
    ws.cell(
        row=r + 3, column=1,
        value=(
            "Outward is deducted from the article's own group (FG sale → FG group). "
            "Adj-In cancels the FG outward and Adj-Out deducts the BOM-derived "
            "ingredients from each source RM group. ± Rearr auto-rebalances cells "
            "that would close negative against units with stock for the same group "
            "(sums to zero per group across units)."
        ),
    )
    ws.merge_cells(start_row=r + 3, start_column=1, end_row=r + 3, end_column=5)
    ws.cell(row=r + 3, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r + 3].height = 48


def write_classification_sheet(wb: Workbook, rows: list[tuple[str, str]]):
    ws = wb.create_sheet(title="Sale_Classification")
    ws["A1"] = "Sale-line classification"
    ws["A1"].font = Font(bold=True, size=14, color="16513F")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws.cell(row=3, column=1, value="Category")
    ws.cell(row=3, column=2, value="Count")
    style_header(ws, 2, row=3)
    for i, (label, count) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=label)
        v = to_number(count)
        cell = ws.cell(row=i, column=2, value=v)
        if isinstance(v, (int, float)):
            cell.number_format = "#,##0"
    ws.column_dimensions["A"].width = 56
    ws.column_dimensions["B"].width = 14


def write_intercompany_sheets(wb: Workbook, html_text: str):
    section = extract_intercompany_section(html_text)
    if not section:
        return
    meta = parse_intercompany_meta(section)
    p = SectionTableParser()
    p.feed(section)
    # First table: digest (Group, Entity, KG)
    # Second table: line items (Date, Doc No, Entity, Buyer, Unit, Group, Article, KG)
    if not p.tables:
        return

    digest_hdr, digest_rows = p.tables[0]
    ws = wb.create_sheet(title="Intercompany_Digest")
    ws["A1"] = "Inter-company Stock-Transfer Invoices — Digest"
    ws["A1"].font = Font(bold=True, size=14, color="16513F")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(digest_hdr))
    if meta:
        ws["A2"] = (
            f"{meta.get('line_items', '?')} line items · total "
            f"{meta.get('total_kg', '?')} kg "
            "(excluded from Outward / Closing)"
        )
        ws["A2"].font = Font(italic=True, color="666666")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(digest_hdr))
    for i, h in enumerate(digest_hdr, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, len(digest_hdr), row=4)
    for r_idx, row in enumerate(digest_rows, start=5):
        for c_idx, v in enumerate(row, start=1):
            num = to_number(v)
            cell = ws.cell(row=r_idx, column=c_idx, value=num)
            if isinstance(num, (int, float)):
                cell.number_format = "#,##0"
    for i, w in enumerate([20, 14, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    if len(p.tables) < 2:
        return
    lines_hdr, lines_rows = p.tables[1]
    ws2 = wb.create_sheet(title="Intercompany_Lines")
    ws2["A1"] = "Inter-company Stock-Transfer Invoices — Line items"
    ws2["A1"].font = Font(bold=True, size=14, color="16513F")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(lines_hdr))
    for i, h in enumerate(lines_hdr, start=1):
        ws2.cell(row=3, column=i, value=h)
    style_header(ws2, len(lines_hdr), row=3)
    for r_idx, row in enumerate(lines_rows, start=4):
        for c_idx, v in enumerate(row, start=1):
            num = to_number(v) if lines_hdr[c_idx - 1].strip().upper() in ("KG",) else v
            cell = ws2.cell(row=r_idx, column=c_idx, value=num)
            if isinstance(num, (int, float)):
                cell.number_format = "#,##0"
    widths = [13, 22, 9, 32, 10, 12, 36, 12]
    for i in range(1, len(lines_hdr) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = widths[i - 1] if i - 1 < len(widths) else 16
    ws2.freeze_panes = "A4"


# ===========================================================================
# Main
# ===========================================================================

def main() -> None:
    text = SRC.read_text(encoding="utf-8")

    parser = ClosingParser()
    parser.feed(text)

    summaries = parse_summaries(text)
    classification = parse_sale_class(text)

    wb = Workbook()
    # Drop default sheet — we'll create Summary at index 0 below
    default = wb.active
    wb.remove(default)

    write_summary_sheet(wb, summaries)
    for view_id in VIEW_TITLES:
        rows = parser.views.get(view_id, [])
        if rows:
            write_closing_sheet(wb, view_id, rows, summaries.get(view_id))

    write_classification_sheet(wb, classification)
    write_intercompany_sheets(wb, text)

    wb.save(OUT)
    total_rows = sum(len(v) for v in parser.views.values())
    print(
        f"wrote {OUT}  ({len(parser.views)} views, {total_rows} data rows, "
        f"{len(wb.sheetnames)} sheets)"
    )


if __name__ == "__main__":
    main()
