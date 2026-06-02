"""Inspect the new FG_Master_Completion (1).xlsx structure vs what sync_fg_master.py expects."""
import openpyxl
from pathlib import Path

paths = {
    "NEW (data)": Path("data/FG_Master_Completion (1).xlsx"),
    "OLD (data)": Path("data/FG_Master_Completion.xlsx"),
}

for label, p in paths.items():
    print("=" * 80)
    print(label, "->", p, "exists:", p.exists())
    if not p.exists():
        continue
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    print("  sheets:", wb.sheetnames)
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"  [{sn}] max_row={ws.max_row} max_col={ws.max_column}")
    # focus on FG_Master_Fill if present
    target = "FG_Master_Fill" if "FG_Master_Fill" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[target]
    print(f"  --- first 5 rows of [{target}] ---")
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i > 5:
            break
        vals = list(row)
        print(f"  row{i}: {vals}")
    # count non-empty FG name rows (col index 1), skipping rows 1-2 and 'Particluars'
    cnt = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= 2:
            continue
        vals = list(row)
        while len(vals) < 17:
            vals.append(None)
        fg = vals[1]
        fg = str(fg).strip() if fg is not None else None
        if fg and fg != "Particluars":
            cnt += 1
    print(f"  data FG rows (parser would read): {cnt}")
    wb.close()
