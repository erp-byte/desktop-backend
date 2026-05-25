"""Quick inspector for BOM Excel files - dumps sheet names, headers, samples, and formulas."""
import sys
import os
from openpyxl import load_workbook

FILES = [
    r"C:\Users\cando\Downloads\Inventory calc\BOM 5 JUN CFPL working for Conv - Hamper.xlsx",
    r"C:\Users\cando\Downloads\Inventory calc\BOM 5 JUN CFPL working for Conv.xlsx",
    r"C:\Users\cando\Downloads\Inventory calc\BOM Details CFPL as on 16-05-26 (1).xlsx",
    r"C:\Users\cando\Downloads\Inventory calc\BOM Item Wise CFPL April 13.xlsx",
    r"C:\Users\cando\Downloads\Inventory calc\BOM 18-05 CDPL.xlsx",
]

OUT = r"d:\Consumption\New\Backend\_bom_working_formulas.txt"


def cell_repr(v):
    if v is None:
        return ""
    s = str(v)
    if len(s) > 80:
        s = s[:77] + "..."
    return s


def inspect(path, out):
    out.write("\n" + "=" * 100 + "\n")
    out.write(f"FILE: {os.path.basename(path)}\n")
    out.write("=" * 100 + "\n")
    try:
        wb_f = load_workbook(path, data_only=False)
        wb_v = load_workbook(path, data_only=True)
    except Exception as e:
        out.write(f"ERROR loading: {e}\n")
        return

    for sname in wb_f.sheetnames:
        sh_f = wb_f[sname]
        sh_v = wb_v[sname]
        out.write(f"\n--- Sheet: {sname!r}  rows={sh_f.max_row} cols={sh_f.max_column} state={sh_f.sheet_state} ---\n")
        if sh_f.max_row == 0 or sh_f.max_column == 0:
            continue

        # Header row(s) - dump first 3 rows raw values
        out.write("First 3 rows (values):\n")
        for r in range(1, min(4, sh_f.max_row + 1)):
            row = []
            for c in range(1, min(sh_f.max_column + 1, 30)):
                v = sh_v.cell(r, c).value
                row.append(cell_repr(v))
            out.write(f"  R{r}: " + " | ".join(row) + "\n")

        # Rows 2-6 values
        if sh_f.max_row >= 2:
            out.write("Sample data rows 2-6 (values):\n")
            for r in range(2, min(7, sh_f.max_row + 1)):
                row = []
                for c in range(1, min(sh_f.max_column + 1, 30)):
                    v = sh_v.cell(r, c).value
                    row.append(cell_repr(v))
                out.write(f"  R{r}: " + " | ".join(row) + "\n")

        # Formula dump - scan first 30 rows for any cell containing a formula
        out.write("Formulas (first 30 rows, any column with =):\n")
        any_formula = False
        for r in range(1, min(31, sh_f.max_row + 1)):
            for c in range(1, sh_f.max_column + 1):
                v = sh_f.cell(r, c).value
                if isinstance(v, str) and v.startswith("="):
                    any_formula = True
                    coord = sh_f.cell(r, c).coordinate
                    out.write(f"  {coord}: {v}\n")
        if not any_formula:
            out.write("  (no formulas in first 30 rows)\n")

        # Also: scan for distinct formula PATTERNS across whole sheet (sample one per column)
        out.write("Formula pattern per column (first occurrence in sheet):\n")
        seen_col = set()
        for r in range(1, sh_f.max_row + 1):
            for c in range(1, sh_f.max_column + 1):
                if c in seen_col:
                    continue
                v = sh_f.cell(r, c).value
                if isinstance(v, str) and v.startswith("="):
                    coord = sh_f.cell(r, c).coordinate
                    out.write(f"  col {c} ({sh_f.cell(1, c).value!r}) first formula at {coord}: {v}\n")
                    seen_col.add(c)
            if len(seen_col) >= sh_f.max_column:
                break

        # Hidden columns / hidden state
        hidden_cols = [k for k, dim in sh_f.column_dimensions.items() if dim.hidden]
        if hidden_cols:
            out.write(f"Hidden columns: {hidden_cols}\n")

    # Defined names
    if wb_f.defined_names:
        out.write("\nDefined names:\n")
        try:
            for dn in wb_f.defined_names:
                try:
                    out.write(f"  {dn}: {wb_f.defined_names[dn].value}\n")
                except Exception:
                    out.write(f"  {dn}\n")
        except Exception as e:
            out.write(f"  (error reading defined names: {e})\n")


def main():
    with open(OUT, "w", encoding="utf-8") as f:
        for path in FILES:
            if not os.path.exists(path):
                f.write(f"\nMISSING: {path}\n")
                continue
            inspect(path, f)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
