"""Generate Excel workbook of all job_card* table fields.

Sources:
  - app/db/production_schema.sql  (base DDL)
  - app/db/production_migrate.sql (ADD/DROP COLUMN migrations applied since)

Each table = one worksheet. An "Overview" sheet lists all tables and a one-line
purpose. Columns dropped via migration are omitted; columns added via migration
are appended at the end with source = 'migrate'.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent / "jobcard_tables_schema.xlsx"

# ---------------------------------------------------------------------------
# Table definitions.  Each row: (column, type, nullable, default/notes, source)
# ---------------------------------------------------------------------------

TABLES: dict[str, dict] = {
    "job_card": {
        "purpose": "Job card header (one per production step / stage).",
        "columns": [
            # from production_schema.sql
            ("job_card_id",             "SERIAL",        "NO",  "PK",                                                  "schema"),
            ("job_card_number",         "TEXT",          "NO",  "UNIQUE  e.g. PO-2026-0042/1",                         "schema"),
            ("prod_order_id",           "INT",           "NO",  "FK production_order(prod_order_id)",                  "schema"),
            ("bom_id",                  "INT",           "YES", "FK bom_header(bom_id)",                               "schema"),
            ("step_number",             "INT",           "NO",  "1, 2, 3...",                                          "schema"),
            ("process_name",            "TEXT",          "NO",  "Sorting, Weighing, etc.",                             "schema"),
            ("stage",                   "TEXT",          "NO",  "sorting, weighing, etc.",                             "schema"),
            ("fg_sku_name",             "TEXT",          "NO",  "denormalised from production_order",                  "schema"),
            ("customer_name",           "TEXT",          "YES", "",                                                    "schema"),
            ("batch_number",            "TEXT",          "NO",  "",                                                    "schema"),
            ("batch_size_kg",           "NUMERIC(15,3)", "NO",  "",                                                    "schema"),
            ("machine_id",              "INT",           "YES", "FK machine(machine_id)",                              "schema"),
            ("assigned_to_team_leader", "TEXT",          "YES", "",                                                    "schema"),
            ("team_members",            "TEXT[]",        "YES", "",                                                    "schema"),
            ("is_locked",               "BOOLEAN",       "NO",  "DEFAULT TRUE",                                        "schema"),
            ("locked_reason",           "TEXT",          "YES", "awaiting_previous_stage, material_pending",           "schema"),
            ("force_unlocked",          "BOOLEAN",       "YES", "DEFAULT FALSE",                                       "schema"),
            ("force_unlock_by",         "TEXT",          "YES", "",                                                    "schema"),
            ("force_unlock_reason",     "TEXT",          "YES", "",                                                    "schema"),
            ("force_unlock_at",         "TIMESTAMPTZ",   "YES", "",                                                    "schema"),
            ("status",                  "TEXT",          "NO",  "DEFAULT 'locked'  (locked/unlocked/assigned/material_received/in_progress/completed/closed)", "schema"),
            ("start_time",              "TIMESTAMPTZ",   "YES", "",                                                    "schema"),
            ("end_time",                "TIMESTAMPTZ",   "YES", "",                                                    "schema"),
            ("total_time_min",          "NUMERIC(10,2)", "YES", "",                                                    "schema"),
            ("factory",                 "TEXT",          "YES", "",                                                    "schema"),
            ("floor",                   "TEXT",          "YES", "",                                                    "schema"),
            ("entity",                  "TEXT",          "YES", "CHECK IN ('cfpl','cdpl')",                            "schema"),
            ("next_job_card_id",        "INT",           "YES", "FK job_card(job_card_id)  - chaining",                "schema"),
            ("prev_job_card_id",        "INT",           "YES", "FK job_card(job_card_id)  - chaining",                "schema"),
            ("carried_qty_kg",          "NUMERIC(15,3)", "NO",  "DEFAULT 0  (received from prev stage)",               "schema"),
            ("dispatched_to_next_kg",   "NUMERIC(15,3)", "NO",  "DEFAULT 0  (pushed to next stage)",                   "schema"),
            ("updated_at",              "TIMESTAMPTZ",   "YES", "",                                                    "schema"),
            ("updated_by",              "TEXT",          "YES", "",                                                    "schema"),
            ("deleted_at",              "TIMESTAMPTZ",   "YES", "soft-delete",                                         "schema"),
            ("deleted_by",              "TEXT",          "YES", "",                                                    "schema"),
            ("cancellation_reason",     "TEXT",          "YES", "",                                                    "schema"),
            ("created_at",              "TIMESTAMPTZ",   "NO",  "DEFAULT NOW()",                                       "schema"),
            # added via production_migrate.sql (lines 47-56)
            ("sales_order_ref",         "TEXT",          "YES", "",                                                    "migrate"),
            ("article_code",            "TEXT",          "YES", "",                                                    "migrate"),
            ("mrp",                     "NUMERIC(15,3)", "YES", "",                                                    "migrate"),
            ("ean",                     "TEXT",          "YES", "",                                                    "migrate"),
            ("bu",                      "TEXT",          "YES", "",                                                    "migrate"),
            ("fumigation",              "BOOLEAN",       "YES", "DEFAULT FALSE",                                       "migrate"),
            ("metal_detector_used",     "BOOLEAN",       "YES", "DEFAULT FALSE",                                       "migrate"),
            ("roasting_pasteurization", "BOOLEAN",       "YES", "DEFAULT FALSE",                                       "migrate"),
            ("control_sample_gm",       "NUMERIC(10,2)", "YES", "",                                                    "migrate"),
            ("magnets_used",            "BOOLEAN",       "YES", "DEFAULT FALSE",                                       "migrate"),
            # added via production_migrate.sql line 199
            ("store_allocation_status", "TEXT",          "YES", "DEFAULT 'pending'",                                   "migrate"),
        ],
    },

    "job_card_partial_dispatch": {
        "purpose": "Chunked handoffs between chained job cards.",
        "columns": [
            ("dispatch_id",      "SERIAL",        "NO",  "PK",                                  "schema"),
            ("from_job_card_id", "INT",           "NO",  "FK job_card(job_card_id)",            "schema"),
            ("to_job_card_id",   "INT",           "NO",  "FK job_card(job_card_id)",            "schema"),
            ("qty_kg",           "NUMERIC(15,3)", "NO",  "CHECK > 0",                           "schema"),
            ("dispatched_at",    "TIMESTAMPTZ",   "NO",  "DEFAULT NOW()",                       "schema"),
            ("dispatched_by",    "TEXT",          "YES", "",                                    "schema"),
        ],
    },

    "job_card_rm_indent": {
        "purpose": "Raw-material requirement lines for a job card.",
        "columns": [
            ("rm_indent_id",       "SERIAL",        "NO",  "PK",                                              "schema"),
            ("job_card_id",        "INT",           "NO",  "FK job_card(job_card_id)",                        "schema"),
            ("material_sku_name",  "TEXT",          "NO",  "",                                                "schema"),
            ("uom",                "TEXT",          "YES", "kg",                                              "schema"),
            ("reqd_qty",           "NUMERIC(15,3)", "NO",  "",                                                "schema"),
            ("loss_pct",           "NUMERIC(5,3)",  "YES", "DEFAULT 0",                                       "schema"),
            ("gross_qty",          "NUMERIC(15,3)", "NO",  "reqd / (1 - loss%)",                              "schema"),
            ("issued_qty",         "NUMERIC(15,3)", "YES", "DEFAULT 0",                                       "schema"),
            ("batch_no",           "TEXT",          "YES", "lot number from po_box",                          "schema"),
            ("godown",             "TEXT",          "YES", "",                                                "schema"),
            ("scanned_box_ids",    "TEXT[]",        "YES", "array of po_box.box_id",                          "schema"),
            ("variance",           "NUMERIC(15,3)", "YES", "issued - gross",                                  "schema"),
            ("status",             "TEXT",          "NO",  "DEFAULT 'pending'  (pending/partial/fulfilled)",  "schema"),
            ("created_at",         "TIMESTAMPTZ",   "NO",  "DEFAULT NOW()",                                   "schema"),
            # migrations 184-189
            ("store_decision",     "TEXT",          "YES", "DEFAULT 'pending'",                               "migrate"),
            ("store_approved_qty", "NUMERIC(15,3)", "YES", "DEFAULT 0",                                       "migrate"),
            ("store_decided_by",   "TEXT",          "YES", "",                                                "migrate"),
            ("store_decided_at",   "TIMESTAMPTZ",   "YES", "",                                                "migrate"),
            ("source_location",    "TEXT",          "YES", "",                                                "migrate"),
            ("quality_grade",      "TEXT",          "YES", "",                                                "migrate"),
            # migrations 207-208
            ("manual_ack_by",      "TEXT",          "YES", "",                                                "migrate"),
            ("manual_ack_at",      "TIMESTAMPTZ",   "YES", "",                                                "migrate"),
            # migrations 533-534
            ("bom_line_id",        "INT",           "YES", "links indent row to bom_line",                    "migrate"),
            ("consumed_qty",       "NUMERIC(15,3)", "YES", "per-line actual consumed",                        "migrate"),
        ],
    },

    "job_card_pm_indent": {
        "purpose": "Packaging-material requirement lines for a job card.",
        "columns": [
            ("pm_indent_id",       "SERIAL",        "NO",  "PK",                                              "schema"),
            ("job_card_id",        "INT",           "NO",  "FK job_card(job_card_id)",                        "schema"),
            ("material_sku_name",  "TEXT",          "NO",  "",                                                "schema"),
            ("uom",                "TEXT",          "YES", "pcs",                                             "schema"),
            ("reqd_qty",           "NUMERIC(15,3)", "NO",  "",                                                "schema"),
            ("loss_pct",           "NUMERIC(5,3)",  "YES", "DEFAULT 0",                                       "schema"),
            ("gross_qty",          "NUMERIC(15,3)", "NO",  "",                                                "schema"),
            ("issued_qty",         "NUMERIC(15,3)", "YES", "DEFAULT 0",                                       "schema"),
            ("batch_no",           "TEXT",          "YES", "",                                                "schema"),
            ("godown",             "TEXT",          "YES", "",                                                "schema"),
            ("scanned_box_ids",    "TEXT[]",        "YES", "",                                                "schema"),
            ("variance",           "NUMERIC(15,3)", "YES", "",                                                "schema"),
            ("status",             "TEXT",          "NO",  "DEFAULT 'pending'",                               "schema"),
            ("created_at",         "TIMESTAMPTZ",   "NO",  "DEFAULT NOW()",                                   "schema"),
            ("store_decision",     "TEXT",          "YES", "DEFAULT 'pending'",                               "migrate"),
            ("store_approved_qty", "NUMERIC(15,3)", "YES", "DEFAULT 0",                                       "migrate"),
            ("store_decided_by",   "TEXT",          "YES", "",                                                "migrate"),
            ("store_decided_at",   "TIMESTAMPTZ",   "YES", "",                                                "migrate"),
            ("source_location",    "TEXT",          "YES", "",                                                "migrate"),
            ("quality_grade",      "TEXT",          "YES", "",                                                "migrate"),
            ("manual_ack_by",      "TEXT",          "YES", "",                                                "migrate"),
            ("manual_ack_at",      "TIMESTAMPTZ",   "YES", "",                                                "migrate"),
            ("bom_line_id",        "INT",           "YES", "links indent row to bom_line",                    "migrate"),
            ("consumed_qty",       "NUMERIC(15,3)", "YES", "per-line actual consumed",                        "migrate"),
        ],
    },

    "job_card_process_step": {
        "purpose": "Process-step rows within a single job card.",
        "columns": [
            ("step_id",          "SERIAL",        "NO",  "PK",                                "schema"),
            ("job_card_id",      "INT",           "NO",  "FK job_card(job_card_id)",          "schema"),
            ("step_number",      "INT",           "NO",  "UNIQUE per (job_card_id, step_number)", "schema"),
            ("process_name",     "TEXT",          "NO",  "",                                  "schema"),
            ("machine_name",     "TEXT",          "YES", "",                                  "schema"),
            ("std_time_min",     "NUMERIC(10,2)", "YES", "",                                  "schema"),
            ("qc_check",         "TEXT",          "YES", "",                                  "schema"),
            ("loss_pct",         "NUMERIC(5,3)",  "YES", "DEFAULT 0",                         "schema"),
            ("operator_name",    "TEXT",          "YES", "",                                  "schema"),
            ("operator_sign_at", "TIMESTAMPTZ",   "YES", "",                                  "schema"),
            ("qc_sign_at",       "TIMESTAMPTZ",   "YES", "",                                  "schema"),
            ("time_done",        "TIMESTAMPTZ",   "YES", "",                                  "schema"),
            ("status",           "TEXT",          "NO",  "DEFAULT 'pending'  (pending/in_progress/completed)", "schema"),
            ("created_at",       "TIMESTAMPTZ",   "NO",  "DEFAULT NOW()",                     "schema"),
        ],
    },

    "job_card_output": {
        "purpose": "FG output / RM consumed / losses for the job card (Section 5).",
        "columns": [
            ("output_id",         "SERIAL",        "NO",  "PK",                                "schema"),
            ("job_card_id",       "INT",           "NO",  "UNIQUE  FK job_card(job_card_id)",  "schema"),
            ("fg_expected_units", "INT",           "YES", "",                                  "schema"),
            ("fg_actual_units",   "INT",           "YES", "",                                  "schema"),
            ("fg_expected_kg",    "NUMERIC(15,3)", "YES", "",                                  "schema"),
            ("fg_actual_kg",      "NUMERIC(15,3)", "YES", "",                                  "schema"),
            ("rm_consumed_kg",    "NUMERIC(15,3)", "YES", "",                                  "schema"),
            ("process_loss_kg",   "NUMERIC(15,3)", "YES", "DEFAULT 0",                         "schema"),
            ("net_output_kg",     "NUMERIC(15,3)", "YES", "DEFAULT 0  (fg_actual_kg + byproduct_total_kg)", "schema"),
            ("yield_pct",         "NUMERIC(8,3)",  "YES", "DEFAULT 0  (net / rm_consumed * 100)", "schema"),
            ("created_at",        "TIMESTAMPTZ",   "NO",  "DEFAULT NOW()",                     "schema"),
            # NOTE: extra_give_away_kg, balance_material_kg, wastage_kg, material_return_kg,
            # rejection_kg, rejection_reason, process_loss_pct, offgrade_kg, offgrade_category,
            # dispatch_qty were all DROPPED by production_migrate.sql lines 514-523.
        ],
    },

    "job_card_environment": {
        "purpose": "Environmental parameter readings (Annexure C).",
        "columns": [
            ("env_id",         "SERIAL",      "NO",  "PK",                                          "schema"),
            ("job_card_id",    "INT",         "NO",  "FK job_card(job_card_id)",                    "schema"),
            ("parameter_name", "TEXT",        "NO",  "brine_salinity, temp, humidity, fan_pct, rpm, gas, magnet", "schema"),
            ("value",          "TEXT",        "YES", "",                                            "schema"),
            ("updated_at",     "TIMESTAMPTZ", "YES", "",                                            "schema"),
            ("updated_by",     "TEXT",        "YES", "",                                            "schema"),
            ("deleted_at",     "TIMESTAMPTZ", "YES", "soft-delete",                                 "schema"),
            ("deleted_by",     "TEXT",        "YES", "",                                            "schema"),
            ("recorded_at",    "TIMESTAMPTZ", "NO",  "DEFAULT NOW()",                               "schema"),
        ],
    },

    "job_card_metal_detection": {
        "purpose": "Metal-detection / seal / weight / oven readings (Annexure A/B).",
        "columns": [
            ("detection_id",      "SERIAL",        "NO",  "PK",                                       "schema"),
            ("job_card_id",       "INT",           "NO",  "FK job_card(job_card_id)",                 "schema"),
            ("check_type",        "TEXT",          "NO",  "pre_packaging / post_packaging",           "schema"),
            ("fe_pass",           "BOOLEAN",       "YES", "",                                         "schema"),
            ("nfe_pass",          "BOOLEAN",       "YES", "",                                         "schema"),
            ("ss_pass",           "BOOLEAN",       "YES", "",                                         "schema"),
            ("failed_units",      "INT",           "YES", "DEFAULT 0",                                "schema"),
            ("remarks",           "TEXT",          "YES", "",                                         "schema"),
            ("updated_at",        "TIMESTAMPTZ",   "YES", "",                                         "schema"),
            ("updated_by",        "TEXT",          "YES", "",                                         "schema"),
            ("deleted_at",        "TIMESTAMPTZ",   "YES", "soft-delete",                              "schema"),
            ("deleted_by",        "TEXT",          "YES", "",                                         "schema"),
            ("recorded_at",       "TIMESTAMPTZ",   "NO",  "DEFAULT NOW()",                            "schema"),
            ("seal_check",        "BOOLEAN",       "YES", "",                                         "migrate"),
            ("seal_failed_units", "INT",           "YES", "DEFAULT 0",                                "migrate"),
            ("wt_check",          "BOOLEAN",       "YES", "",                                         "migrate"),
            ("wt_failed_units",   "INT",           "YES", "DEFAULT 0",                                "migrate"),
            ("dough_temp_c",      "NUMERIC(10,2)", "YES", "",                                         "migrate"),
            ("oven_temp_c",       "NUMERIC(10,2)", "YES", "",                                         "migrate"),
            ("baking_temp_c",     "NUMERIC(10,2)", "YES", "",                                         "migrate"),
        ],
    },

    "job_card_weight_check": {
        "purpose": "20-sample weight/leak checks (Annexure B).",
        "columns": [
            ("check_id",          "SERIAL",        "NO",  "PK",                            "schema"),
            ("job_card_id",       "INT",           "NO",  "FK job_card(job_card_id)",      "schema"),
            ("sample_number",     "INT",           "NO",  "",                              "schema"),
            ("net_weight",        "NUMERIC(15,3)", "YES", "",                              "schema"),
            ("gross_weight",      "NUMERIC(15,3)", "YES", "",                              "schema"),
            ("leak_test_pass",    "BOOLEAN",       "YES", "",                              "schema"),
            ("updated_at",        "TIMESTAMPTZ",   "YES", "",                              "schema"),
            ("updated_by",        "TEXT",          "YES", "",                              "schema"),
            ("deleted_at",        "TIMESTAMPTZ",   "YES", "soft-delete",                   "schema"),
            ("deleted_by",        "TEXT",          "YES", "",                              "schema"),
            ("recorded_at",       "TIMESTAMPTZ",   "NO",  "DEFAULT NOW()",                 "schema"),
            ("target_wt_g",       "NUMERIC(10,2)", "YES", "",                              "migrate"),
            ("tolerance_g",       "NUMERIC(10,2)", "YES", "",                              "migrate"),
            ("accept_range_min",  "NUMERIC(10,2)", "YES", "",                              "migrate"),
            ("accept_range_max",  "NUMERIC(10,2)", "YES", "",                              "migrate"),
        ],
    },

    "job_card_loss_reconciliation": {
        "purpose": "Loss reconciliation per category (Annexure D).",
        "columns": [
            ("recon_id",          "SERIAL",        "NO",  "PK",                                     "schema"),
            ("job_card_id",       "INT",           "NO",  "FK job_card(job_card_id)",               "schema"),
            ("loss_category",     "TEXT",          "NO",  "sorting_rejection, roasting_loss, packaging_rejection, metal_detector, spillage, qc_sample", "schema"),
            ("budgeted_loss_pct", "NUMERIC(5,3)",  "YES", "",                                       "schema"),
            ("budgeted_loss_kg",  "NUMERIC(15,3)", "YES", "",                                       "schema"),
            ("actual_loss_kg",    "NUMERIC(15,3)", "YES", "",                                       "schema"),
            ("variance_kg",       "NUMERIC(15,3)", "YES", "",                                       "schema"),
            ("remarks",           "TEXT",          "YES", "",                                       "schema"),
            ("updated_at",        "TIMESTAMPTZ",   "YES", "",                                       "schema"),
            ("updated_by",        "TEXT",          "YES", "",                                       "schema"),
            ("deleted_at",        "TIMESTAMPTZ",   "YES", "soft-delete",                            "schema"),
            ("deleted_by",        "TEXT",          "YES", "",                                       "schema"),
            ("created_at",        "TIMESTAMPTZ",   "NO",  "DEFAULT NOW()",                          "schema"),
        ],
    },

    "job_card_remarks": {
        "purpose": "Observation / deviation / corrective-action notes (Annexure E).",
        "columns": [
            ("remark_id",    "SERIAL",      "NO",  "PK",                                                  "schema"),
            ("job_card_id",  "INT",         "NO",  "FK job_card(job_card_id)",                            "schema"),
            ("remark_type",  "TEXT",        "NO",  "observation / deviation / corrective_action",         "schema"),
            ("content",      "TEXT",        "YES", "",                                                    "schema"),
            ("recorded_by",  "TEXT",        "YES", "",                                                    "schema"),
            ("updated_at",   "TIMESTAMPTZ", "YES", "",                                                    "schema"),
            ("updated_by",   "TEXT",        "YES", "",                                                    "schema"),
            ("deleted_at",   "TIMESTAMPTZ", "YES", "soft-delete",                                         "schema"),
            ("deleted_by",   "TEXT",        "YES", "",                                                    "schema"),
            ("recorded_at",  "TIMESTAMPTZ", "NO",  "DEFAULT NOW()",                                       "schema"),
        ],
    },

    "job_card_sign_off": {
        "purpose": "Section 6 / Annexure page 4 sign-offs (one row per role).",
        "columns": [
            ("sign_off_id",   "SERIAL",      "NO",  "PK",                                          "migrate"),
            ("job_card_id",   "INT",         "NO",  "FK job_card(job_card_id)",                    "migrate"),
            ("sign_off_type", "TEXT",        "NO",  "UNIQUE per (job_card_id, sign_off_type)",     "migrate"),
            ("name",          "TEXT",        "YES", "",                                            "migrate"),
            ("signed_at",     "TIMESTAMPTZ", "YES", "",                                            "migrate"),
        ],
    },

    "job_card_material_consumption": {
        "purpose": "Per-BOM-line actual consumption tracking for job-card close.",
        "columns": [
            ("consumption_id",      "SERIAL",        "NO",  "PK",                                                "migrate"),
            ("job_card_id",         "INT",           "NO",  "FK job_card(job_card_id)",                          "migrate"),
            ("rm_indent_id",        "INT",           "YES", "",                                                  "migrate"),
            ("material_sku_name",   "TEXT",          "NO",  "",                                                  "migrate"),
            ("item_type",           "TEXT",          "NO",  "DEFAULT 'rm'",                                      "migrate"),
            ("uom",                 "TEXT",          "YES", "",                                                  "migrate"),
            ("bom_reqd_qty",        "NUMERIC(15,3)", "NO",  "DEFAULT 0",                                         "migrate"),
            ("issued_qty",          "NUMERIC(15,3)", "YES", "DEFAULT 0",                                         "migrate"),
            ("actual_consumed_qty", "NUMERIC(15,3)", "YES", "",                                                  "migrate"),
            ("return_qty",          "NUMERIC(15,3)", "YES", "DEFAULT 0",                                         "migrate"),
            ("variance_qty",        "NUMERIC(15,3)", "YES", "",                                                  "migrate"),
            ("remarks",             "TEXT",          "YES", "",                                                  "migrate"),
            ("created_at",          "TIMESTAMPTZ",   "NO",  "DEFAULT NOW()",                                     "migrate"),
            # UNIQUE (job_card_id, material_sku_name, item_type)
        ],
    },

    "job_card_byproduct": {
        "purpose": "By-product / off-grade output categories (one row per category).",
        "columns": [
            ("byproduct_id", "SERIAL",        "NO",  "PK",                                          "migrate"),
            ("job_card_id",  "INT",           "NO",  "FK job_card(job_card_id)",                    "migrate"),
            ("category",     "TEXT",          "NO",  "UNIQUE per (job_card_id, category)",          "migrate"),
            ("quantity_kg",  "NUMERIC(15,3)", "NO",  "DEFAULT 0",                                   "migrate"),
            ("remarks",      "TEXT",          "YES", "",                                            "migrate"),
            ("created_at",   "TIMESTAMPTZ",   "NO",  "DEFAULT NOW()",                               "migrate"),
            ("uom",          "TEXT",          "YES", "DEFAULT 'kg'",                                "migrate"),
        ],
    },

    "job_card_material_accounting": {
        "purpose": "Material-accounting summary (one row per job card).",
        "columns": [
            ("accounting_id",                "SERIAL",        "NO",  "PK",                              "migrate"),
            ("job_card_id",                  "INT",           "NO",  "UNIQUE  FK job_card(job_card_id)", "migrate"),
            ("total_material_issued_kg",     "NUMERIC(15,3)", "NO",  "DEFAULT 0",                       "migrate"),
            ("fg_output_kg",                 "NUMERIC(15,3)", "NO",  "DEFAULT 0",                       "migrate"),
            ("process_loss_kg",              "NUMERIC(15,3)", "YES", "DEFAULT 0",                       "migrate"),
            ("process_loss_sub_categories",  "JSONB",         "YES", "",                                "migrate"),
            ("extra_give_away_kg",           "NUMERIC(15,3)", "YES", "DEFAULT 0",                       "migrate"),
            ("balance_material_kg",          "NUMERIC(15,3)", "YES", "DEFAULT 0",                       "migrate"),
            ("offgrade_total_kg",            "NUMERIC(15,3)", "YES", "DEFAULT 0",                       "migrate"),
            ("rejection_kg",                 "NUMERIC(15,3)", "YES", "DEFAULT 0",                       "migrate"),
            ("wastage_kg",                   "NUMERIC(15,3)", "YES", "DEFAULT 0",                       "migrate"),
            ("process_loss_pct",             "NUMERIC(8,3)",  "YES", "DEFAULT 0",                       "migrate"),
            ("other_loss_pct",               "NUMERIC(8,3)",  "YES", "DEFAULT 0",                       "migrate"),
            ("total_loss_pct",               "NUMERIC(8,3)",  "YES", "DEFAULT 0",                       "migrate"),
            ("balance_difference_kg",        "NUMERIC(15,3)", "YES", "DEFAULT 0",                       "migrate"),
            ("is_balanced",                  "BOOLEAN",       "YES", "DEFAULT FALSE",                   "migrate"),
            ("balanced_by",                  "TEXT",          "YES", "",                                "migrate"),
            ("balanced_at",                  "TIMESTAMPTZ",   "YES", "",                                "migrate"),
            ("created_at",                   "TIMESTAMPTZ",   "NO",  "DEFAULT NOW()",                   "migrate"),
            ("control_sample_kg",            "NUMERIC(15,3)", "YES", "DEFAULT 0",                       "migrate"),
        ],
    },

    "job_card_balance_material": {
        "purpose": "Per-material extras/returns/wastage/control-sample records.",
        "columns": [
            ("balance_id",    "SERIAL",        "NO",  "PK",                                                          "migrate"),
            ("job_card_id",   "INT",           "NO",  "FK job_card(job_card_id)",                                    "migrate"),
            ("material_id",   "INT",           "YES", "",                                                            "migrate"),
            ("material_name", "TEXT",          "NO",  "",                                                            "migrate"),
            ("balance_type",  "TEXT",          "NO",  "CHECK IN ('extra_given','returned','wastage','control_sample')", "migrate"),
            ("qty_kg",        "NUMERIC(15,3)", "NO",  "DEFAULT 0",                                                   "migrate"),
            ("remarks",       "TEXT",          "YES", "",                                                            "migrate"),
            ("created_at",    "TIMESTAMPTZ",   "NO",  "DEFAULT NOW()",                                               "migrate"),
            ("bom_line_id",   "INT",           "YES", "",                                                            "migrate"),
        ],
    },
}


# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
SOURCE_FILLS = {
    "schema":  PatternFill(start_color="E8F1FA", end_color="E8F1FA", fill_type="solid"),
    "migrate": PatternFill(start_color="FFF4D6", end_color="FFF4D6", fill_type="solid"),
}

wb = Workbook()

# --- Overview sheet ---
ws = wb.active
ws.title = "Overview"
ws.append(["Table", "# columns", "Purpose"])
for cell in ws[1]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = WRAP
for tname, tdef in TABLES.items():
    ws.append([tname, len(tdef["columns"]), tdef["purpose"]])
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 10
ws.column_dimensions["C"].width = 90
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = WRAP

# --- One sheet per table ---
HEADERS = ["#", "column", "type", "nullable", "default / notes", "source"]
WIDTHS  = [4,   30,        18,     10,         70,                10]

for tname, tdef in TABLES.items():
    # Excel sheet-name limit is 31 chars
    sheet_name = tname[:31]
    ws = wb.create_sheet(title=sheet_name)
    # Title row
    ws["A1"] = tname
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    ws["A2"] = tdef["purpose"]
    ws["A2"].alignment = WRAP
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
    # Header row
    ws.append([])  # row 3 blank
    ws.append(HEADERS)
    for cell in ws[4]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
    # Data rows
    for idx, (col, typ, nullable, notes, source) in enumerate(tdef["columns"], start=1):
        row = [idx, col, typ, nullable, notes, source]
        ws.append(row)
        fill = SOURCE_FILLS[source]
        for cell in ws[ws.max_row]:
            cell.alignment = WRAP
            cell.fill = fill
    # Column widths
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    # Freeze header
    ws.freeze_panes = "A5"

wb.save(OUT)
print(f"wrote {OUT}  ({sum(len(t['columns']) for t in TABLES.values())} columns across {len(TABLES)} tables)")
