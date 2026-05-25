"""Closing-stock v5 — image-based adjustment model + warehouse-from-narration + hover tooltips.

Key changes from v4 (per user feedback):

1. **Image-based outward + adjustment model** (matches the calculation grid in
   WhatsApp Image 2026-05-18). FG sales are deducted from the FG group as outward
   (like "TRAIL MIX outward = 20mt"), then an "ADJUSTMENT" column adds it back to
   the FG group (+20mt to TRAIL MIX) and deducts the BOM-derived ingredients from
   their source RM groups (e.g. −5mt to CASHEW, −5mt to ALMOND, etc.).
   This makes FG-only groups like trail mix close to flat instead of deep negative.

2. **CDPL ≠ A-185.** CDPL is a company (entity), not a warehouse. When a CDPL CN
   PDF doesn't specify a warehouse in narration, default to W-202.

3. **CN dedup** across the date-zip extract (`_cn_extracted.csv`) and the
   combined CN PDFs (`_pdf_units_extracted.csv`) by `(doc_no, product)`.

4. **Smarter unmapped categorisation** — when `all_sku.item_group` is null we
   fall back to sub_group / sale_group / keyword inference / pack-size match.

5. **Hover tooltips** in the HTML — every column header and each cell carries a
   `title=` explaining what the number represents and how it was computed.

6. **Unacknowledged transfers included** — already the case (no status filter)
   but now documented.

Formula per RM group (per unit):
    closing = opening + inward + xfr_in - xfr_out + cn - outward
            + adjustment_in - adjustment_out
"""
import asyncio
import csv
import os
import re
import socket
import ssl
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import asyncpg
import openpyxl
from dotenv import load_dotenv

if sys.platform == "win32":
    try: sys.stdout.reconfigure(encoding="utf-8")
    except Exception: pass

load_dotenv()
DB_URL = os.environ["DATABASE_URL"]

XLSX_DIR = Path(r"C:\Users\cando\Downloads\Inventory calc")
OPENING_XLSX = XLSX_DIR / "Physical Stock Compilation 31-03-2026.xlsx"
SALES_XLSX_APR = XLSX_DIR / "Sales Register 30th April 2026.xlsx"
SALES_XLSX_MAY = XLSX_DIR / "Sales Register 12th May 2026.xlsx"
PDF_UNITS_CSV = Path(__file__).parent / "_pdf_units_extracted.csv"
CN_PDF_CSV = Path(__file__).parent / "_cn_extracted.csv"

WINDOW_START = date(2026, 4, 1)

OUT_HTML = Path(__file__).parent / "_closing_v5_report.html"
OUT_TXT = Path(__file__).parent / "_closing_v5_report.txt"

# ────────────────────────────────────────────────────────────────────────────
# Units & mapping
# ────────────────────────────────────────────────────────────────────────────

# Per user: "Cold Storage" is just a common label for the Savla + Rishi cold-room
# stock. So Supreme Cold + any "Cold Storage" tagged transfer folds into SAVLA
# (the larger of the two). A-185 Cold stays under A-185 (it's at the A-185 site).
UNITS = ["W-202", "A-185", "F-53", "A-68", "RISHI", "SAVLA", "OTHER"]

CFPL_LOC_TO_UNIT = {  # column-index in CFPL/CDPL Excel sheets → unit
    7: "W-202", 9: "W-202", 11: "W-202", 13: "W-202", 15: "W-202",
    17: "W-202", 19: "W-202", 21: "W-202",      # W-202 Store + Lower + Upper + 1st + 2nd + Barline + Terrace + Off-Grade
    23: "A-185", 25: "A-185", 27: "A-185",      # A-185 + A-185 Cold + A-185 Off-Grade  (A-185 Cold stays at A-185)
    29: "OTHER",                                # Ajustment Pending
    31: "F-53",                                 # APMC
    33: "RISHI",                                # Rishi
    35: "SAVLA",                                # Savla
    37: "SAVLA",                                # Supreme Cold → SAVLA (per user, Cold Storage = Savla + Rishi)
}


def warehouse_to_unit(wh: str | None) -> str:
    if not wh: return "OTHER"
    # Strip BOTH spaces and dashes so "W-202", "W 202", "W202" all collapse
    w = str(wh).strip().upper().replace(" ", "").replace("-", "")
    if w.startswith("W202"): return "W-202"
    if w.startswith("A185"): return "A-185"
    if w.startswith("F53") or "APMC" in w: return "F-53"
    if w.startswith("A68"): return "A-68"
    if "RISHI" in w: return "RISHI"
    if "SAVLA" in w or "PAWANE" in w or "COLD" in w or w.startswith("D39") or w.startswith("D514"):
        return "SAVLA"
    return "OTHER"


def site_to_unit(site: str | None) -> str:
    return warehouse_to_unit(site)


def is_intercompany(customer: str | None) -> bool:
    """Sales billed to our own group companies are inter-company stock transfers,
    NOT real outward movement. Exclude from outward calculation per user."""
    c = (customer or "").upper().strip()
    if not c: return False
    # Catch the obvious legal-name forms
    if "CANDOR FOODS" in c or "CANDOR DATES" in c: return True
    # Also catch CFPL / CDPL as the buyer
    if c in ("CFPL", "CDPL", "C.F.P.L.", "C.D.P.L."): return True
    return False


def voucher_to_unit(entity: str, voucher_type: str | None,
                     customer: str | None = None, doc_no: str | None = None) -> str:
    """Determine the dispatching unit for a sale row.
    Priority:
      1. Customer-name override (Reliance → W-202, Avenue Supermarts → A-185)
      2. Document-no prefix override (CF53/* → F-53; CDPL/* → A-185)
      3. Voucher-type rule
      4. Entity default (CFPL → W-202, CDPL → A-185)
    """
    cust = (customer or "").upper()
    if cust:
        if "AVENUE SUPERMART" in cust or "D-MART" in cust or "D MART" in cust or "DMART" in cust:
            return "A-185"
        if "RELIANCE" in cust:
            return "W-202"
    # Document-no prefix override
    dn = (doc_no or "").upper().strip()
    if dn.startswith("CF53/"): return "F-53"
    if dn.startswith("CDPL/"):
        # Per user: CDPL dates come from W-202 or Cold (Savla/Rishi), NOT A-185.
        # Heuristic: inter-company bulk sales (to Candor Foods itself, big
        # quantities of wet dates) → RISHI (bulk raw stock). B2B processed
        # paste/powder sales → W-202 (close to processing line).
        if "CANDOR" in cust:
            return "RISHI"
        return "W-202"
    e = (entity or "").upper(); v = (voucher_type or "").upper().strip()
    if e == "CFPL":
        if "APMC" in v: return "F-53"
        return "W-202"
    if e == "CDPL":
        # CDPL default with no other signal — W-202 dispatch hub
        return "W-202"
    return "OTHER"


# ────────────────────────────────────────────────────────────────────────────
# DB connect with DNS fallback
# ────────────────────────────────────────────────────────────────────────────

async def connect_db():
    try:
        return await asyncpg.create_pool(DB_URL, min_size=1, max_size=2, timeout=15)
    except (socket.gaierror, OSError):
        ip_url = DB_URL.replace(
            "wms-postgres-db.cpis084golp7.ap-south-1.rds.amazonaws.com",
            "35.154.219.249",
        )
        ctx = ssl.create_default_context()
        ctx.check_hostname = False; ctx.verify_mode = ssl.CERT_NONE
        return await asyncpg.create_pool(ip_url, min_size=1, max_size=2, timeout=15, ssl=ctx)


# ────────────────────────────────────────────────────────────────────────────
# Helpers
# ────────────────────────────────────────────────────────────────────────────

def norm(s) -> str:
    if s is None: return ""
    # PDF text-extraction often introduces stray punctuation/spacing artefacts
    # (e.g. "Wet Date Sayer ( Unpitted" missing close-paren, "Black Currant 250 Gm"
    # with extra spaces). Collapse all of those before matching.
    s = str(s).strip().lower()
    s = re.sub(r"\(\s*\)", "", s)               # empty ( )
    s = re.sub(r"\(\s*$", "", s).strip()        # trailing "(" with no close
    s = re.sub(r"\s+", " ", s)                  # collapse whitespace
    s = re.sub(r"\s*([(),])\s*", r"\1", s)      # tidy around brackets/commas
    return s.strip()


def norm_group(g) -> str:
    if not g: return "_UNMAPPED"
    return str(g).strip().lower()


def to_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None


def fmt(v: float) -> str:
    if abs(v) < 0.5: return "0"
    return f"{v:,.0f}"


KEYWORD_GROUP_MAP: list[tuple[str, str]] = [
    # More specific multi-word patterns FIRST so they win over short prefixes
    ("trail mix", "trail mix"), ("nuts & seeds mix", "trail mix"),
    ("breakfast mix", "trail mix"), ("chocolate coated", "trail mix"),
    ("millennium nut mix", "trail mix"), ("nut mix", "trail mix"),
    ("bars", "bars & cereals"), ("cereal", "bars & cereals"), ("muesli", "bars & cereals"),
    ("hamper", "festive hampers"), ("gift", "festive hampers"),
    ("medjoul", "dates"), ("ajwa", "dates"), ("kimia", "dates"), ("fard", "dates"),
    ("kalmi", "dates"), ("safavi", "dates"), ("khalas", "dates"), ("khanezi", "dates"),
    ("zahidi", "dates"), ("barakah", "dates"), ("date paste", "dates"),
    ("date powder", "dates"), ("date syrup", "dates"), ("khajur", "dates"),
    ("wet date", "dates"), ("ajfan", "dates"), ("sayer", "dates"),
    ("sufri", "dates"), ("deri", "dates"), ("mazafati", "dates"), ("afreen", "dates"),
    ("dates", "dates"), ("date", "dates"),
    ("almond", "almond"), ("badam", "almond"),
    ("cashew", "cashew"), ("kaju", "cashew"),
    ("pista", "pista"), ("pistachio", "pista"),
    ("raisin", "raisin"), ("kishmish", "raisin"),
    ("walnut", "walnut"), ("akhrot", "walnut"),
    ("peanut", "peanuts"), ("groundnut", "peanuts"),
    ("makhana", "makhana"),
    ("anjeer", "anjeer"), ("fig", "anjeer"),
    ("apricot", "apricot"), ("khubani", "apricot"),
    ("cranberry", "cranberry"), ("blueberry", "blueberry"),
    # PDF text may have "Black Currant" with space — match both
    ("black currant", "blackcurrant"),
    ("blackcurrant", "blackcurrant"), ("blackberry", "blackberry"), ("prune", "prunes"),
    # Dehydrated catches: "NB DEHYDRATED ANYTHING", goji berry, kiwi, mango, peach…
    ("dehydrated", "dehydrated fruits"), ("goji berry", "dehydrated fruits"),
    ("dehy ", "dehydrated fruits"), ("kiwi", "dehydrated fruits"),
    ("dehydrated fruit", "dehydrated fruits"),
    ("seed", "seeds"), ("sunflower", "seeds"), ("pumpkin", "seeds"),
    ("chia", "seeds"), ("flax", "seeds"), ("watermelon", "seeds"),
    ("salt", "salt"), ("seasoning", "seasoning"), ("flavour", "seasoning"),
    # Spices: black pepper / cardamom / spice etc.
    ("black pepper", "spices"), ("cardamom", "spices"),
    ("spice", "spices"), ("tajir", "tajir"),
    # Premium nuts — also catch PDF mis-spellings (hazulnut), no-space variants
    ("macadamia", "premium nuts"), ("pecan", "premium nuts"),
    ("hazelnut", "premium nuts"), ("hazulnut", "premium nuts"),
    ("brazil nut", "premium nuts"), ("brazilnut", "premium nuts"),
    ("pine nut", "premium nuts"), ("pinenut", "premium nuts"),
    ("pouch", "packaging"), ("carton", "packaging"), ("label", "packaging"),
    ("tape", "packaging"), ("box", "packaging"), ("pm24", "packaging"),
    ("pm-", "packaging"), ("blister", "packaging"),
]


def infer_group(name: str) -> str | None:
    n = norm(name)
    for kw, grp in KEYWORD_GROUP_MAP:
        if kw in n: return grp
    return None


def smart_group(name: str, info: dict | None) -> str:
    """Smarter group inference: all_sku → sub_group fallback → keyword → _UNMAPPED."""
    if info:
        if info.get("group") and info["group"] != "_UNMAPPED":
            return info["group"]
        sg = (info.get("sub_group") or "").lower()
        if sg:
            # Use the sub_group's keyword inference as fallback
            for kw, grp in KEYWORD_GROUP_MAP:
                if kw in sg: return grp
    return infer_group(name) or "_UNMAPPED"


# ────────────────────────────────────────────────────────────────────────────
# Opening per unit
# ────────────────────────────────────────────────────────────────────────────

def load_opening_per_unit() -> dict:
    bucket = {u: defaultdict(float) for u in UNITS}
    wb = openpyxl.load_workbook(OPENING_XLSX, read_only=True, data_only=True)
    for sheet_name in ("CFPL", "CDPL"):
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row or len(row) < 36: continue
            group = row[3]
            if not group: continue
            g = norm_group(group)
            for col_idx, unit in CFPL_LOC_TO_UNIT.items():
                if col_idx >= len(row): continue
                val = row[col_idx]
                if val is None: continue
                try: kg = float(val)
                except (TypeError, ValueError): continue
                if abs(kg) < 0.001: continue
                bucket[unit][g] += kg
    wb.close()
    return {u: dict(g) for u, g in bucket.items()}


# ────────────────────────────────────────────────────────────────────────────
# Inward per unit
# ────────────────────────────────────────────────────────────────────────────

async def load_inward_per_unit(conn, sku_info: dict) -> tuple[dict, dict]:
    # Packaging-vendor keywords — used as a fallback when article_description is
    # NULL/empty (orphan boxes from packaging suppliers should be tagged as
    # 'packaging' group, not _UNMAPPED). Per user confirmation.
    PACKAGING_VENDOR_KEYWORDS = (
        "flexipack", "packaging", "packing", "polymer", "carton", "corrugat",
        "amul", "thakkar", "monocarton", "label", "pouch",
    )
    bucket = {u: defaultdict(float) for u in UNITS}
    unmapped: dict = defaultdict(float)
    for prefix in ("cfpl", "cdpl"):
        rows = await conn.fetch(f"""
            SELECT b.article_description AS sku, COALESCE(b.net_weight, 0) AS kg,
                   t.warehouse, t.vendor_supplier_name AS vendor
            FROM {prefix}_boxes_v2 b
            JOIN {prefix}_transactions_v2 t ON b.transaction_no = t.transaction_no
            WHERE t.entry_date >= $1::date
              AND COALESCE(t.rtv, false) = false
              AND COALESCE(t.service, false) = false
        """, WINDOW_START)
        for r in rows:
            kg = float(r["kg"] or 0)
            if kg <= 0: continue
            name = r["sku"]
            info = sku_info.get(norm(name))
            grp = smart_group(name, info)
            # Vendor-based fallback for orphan boxes (article_description NULL/empty)
            if grp == "_UNMAPPED" and (not name or not str(name).strip()):
                vendor = (r["vendor"] or "").lower()
                if any(kw in vendor for kw in PACKAGING_VENDOR_KEYWORDS):
                    grp = "packaging"
            unit = warehouse_to_unit(r["warehouse"])
            if unit == "OTHER":   # null warehouse — default to W-202 per user
                unit = "W-202"
            bucket[unit][grp] += kg
            if grp == "_UNMAPPED": unmapped[name or "<NULL article>"] += kg
    return {u: dict(g) for u, g in bucket.items()}, dict(unmapped)


# ────────────────────────────────────────────────────────────────────────────
# Transfers (all, including unacknowledged)
# ────────────────────────────────────────────────────────────────────────────

async def load_transfers_per_unit(conn, sku_info: dict) -> tuple[dict, dict]:
    """No status filter — includes unacknowledged transfers (per user)."""
    out_b = {u: defaultdict(float) for u in UNITS}
    in_b = {u: defaultdict(float) for u in UNITS}
    rows = await conn.fetch("""
        SELECT h.from_site, h.to_site, l.item_category, l.item_desc_raw,
               COALESCE(l.net_weight, 0) AS kg, h.status
        FROM interunit_transfers_lines l
        JOIN interunit_transfers_header h ON l.header_id = h.id
        WHERE h.stock_trf_date >= $1::date
    """, WINDOW_START)
    for r in rows:
        kg = float(r["kg"] or 0)
        if kg <= 0: continue
        from_u = site_to_unit(r["from_site"])
        to_u = site_to_unit(r["to_site"])
        if from_u == to_u: continue
        info = sku_info.get(norm(r["item_desc_raw"]))
        g = norm_group(r["item_category"]) if r["item_category"] else None
        if not g or g == "_UNMAPPED":
            g = smart_group(r["item_desc_raw"], info)
        # FIX (C): A-185 doesn't physically hold packaging stock. When the
        # transfer record says A-185 → other unit for packaging, the data-entry
        # is wrong (packaging actually shipped from W-202). Reroute so the
        # xfr_out hits the correct source.
        if from_u == "A-185" and g == "packaging":
            from_u = "W-202"
        out_b[from_u][g] += kg
        in_b[to_u][g] += kg
    return ({u: dict(g) for u, g in out_b.items()},
            {u: dict(g) for u, g in in_b.items()})


# ────────────────────────────────────────────────────────────────────────────
# all_sku + BOM
# ────────────────────────────────────────────────────────────────────────────

async def load_sku_info(conn) -> dict:
    rows = await conn.fetch(
        "SELECT particulars, item_type, item_group, sub_group, sale_group FROM all_sku"
    )
    return {
        norm(r["particulars"]): {
            "type": (r["item_type"] or "").lower(),
            "group": norm_group(r["item_group"]),
            "sub_group": r["sub_group"],
            "sale_group": r["sale_group"],
        }
        for r in rows if r["particulars"]
    }


async def load_bom_master(conn) -> dict:
    rows = await conn.fetch("""
        SELECT h.bom_id, h.fg_sku_name, h.customer_name,
               l.material_sku_name, l.quantity_per_unit, l.item_type, l.line_number
        FROM bom_header h JOIN bom_line l ON l.bom_id = h.bom_id
        ORDER BY h.fg_sku_name, h.bom_id DESC, l.line_number
    """)
    # Pick the LATEST bom_id per FG (newer imports supersede older corrupted ones,
    # e.g. Laddubar Bites 9gm has bom_id=453 (old, qty per batch) AND bom_id=1261
    # (new, qty per pack) — we want the latest).
    bom: dict = {}; chosen: dict = {}
    for r in rows:
        fg_key = norm(r["fg_sku_name"])
        if fg_key not in chosen:
            chosen[fg_key] = r["bom_id"]; bom[fg_key] = []
        if r["bom_id"] != chosen[fg_key]: continue
        bom[fg_key].append((
            r["material_sku_name"], float(r["quantity_per_unit"] or 0),
            (r["item_type"] or "rm").lower(),
        ))
    cdpl_path = XLSX_DIR / "BOM 18-05 CDPL.xlsx"
    if cdpl_path.exists():
        wb = openpyxl.load_workbook(cdpl_path, read_only=True, data_only=True)
        ws = wb["BOM of Stock Item"]
        for r in ws.iter_rows(min_row=3, values_only=True):
            stock, _bn, fg_qty, rm, _gd, _typ, bom_qty = r
            if not stock or not rm or fg_qty in (None, 0) or bom_qty is None: continue
            try: per_unit = float(bom_qty) / float(fg_qty)
            except (TypeError, ValueError, ZeroDivisionError): continue
            it = "pm" if str(rm).upper().startswith("PM") else "rm"
            key = norm(stock)
            if key not in bom: bom[key] = []
            bom[key].append((rm, per_unit, it))
        wb.close()
    return bom


def build_aliases(bom: dict) -> dict:
    aliases: dict = {}
    for canonical in bom:
        for v in {
            canonical, canonical.replace(" ", ""),
            re.sub(r"(\d)\s*(gm|g|kg)\b", r"\1\2", canonical),
            re.sub(r"(\d)(gm|g|kg)\b", r"\1 \2", canonical),
            re.sub(r"\s*\([^)]*\)", "", canonical).strip(),
        }:
            if v: aliases.setdefault(v, canonical)
    return aliases


def bom_for(key: str, bom: dict, aliases: dict):
    if key in bom: return bom[key]
    for variant in {
        key, key.replace(" ", ""),
        re.sub(r"(\d)\s*(gm|g|kg)\b", r"\1\2", key),
        re.sub(r"(\d)(gm|g|kg)\b", r"\1 \2", key),
        re.sub(r"\s*\([^)]*\)", "", key).strip(),
    }:
        canon = aliases.get(variant)
        if canon and canon in bom: return bom[canon]
    return None


def expand_bom_recursive(
    fg_key: str, bom: dict, aliases: dict, sku_info: dict,
    fg_group: str | None = None,
    depth: int = 0, max_depth: int = 5, seen: set | None = None,
) -> list:
    """Recursively expand a BOM. ANY non-PM RM that itself has a BOM is
    expanded inline (substituting its own ingredients, scaled). This handles
    multi-level semi-finished chains:
        Trail Mix Pack 100g → Trail Mix Bulk → (Almond + Cashew + Raisin)
        Salted Peanuts 100g → Roasted Peanuts → Raw Peanuts + Salt
        Seasoning Mix → Sub-Seasoning Bulk → individual spices
    Per user: "FG linked as primary RM in BOM needs reconversion".
    Cycle-protected (seen-set) and depth-limited.
    Returns flat list of (rm_name, qty_per_unit_of_top_fg, item_type).
    """
    if seen is None: seen = set()
    if fg_key in seen or depth > max_depth: return []
    seen = seen | {fg_key}
    if fg_group is None:
        info = sku_info.get(fg_key, {})
        fg_group = (info.get("group") or "").lower()
    b = bom_for(fg_key, bom, aliases)
    if not b: return []
    expanded: list = []
    for rm_name, q_per, it in b:
        if it == "pm":
            expanded.append((rm_name, q_per, it)); continue
        # Try recursion: if this RM has its own BOM, substitute. This handles
        # intermediate "Bulks" classified as same-group OR as FG-linked-as-RM
        # in other-group BOMs.
        rm_key = norm(rm_name)
        if depth < max_depth and rm_key not in seen and bom_for(rm_key, bom, aliases):
            sub = expand_bom_recursive(
                rm_key, bom, aliases, sku_info,
                fg_group=fg_group, depth=depth + 1, max_depth=max_depth, seen=seen,
            )
            if sub:
                for sn, sq, sit in sub:
                    expanded.append((sn, sq * q_per, sit))
                continue
        expanded.append((rm_name, q_per, it))
    return expanded


# ────────────────────────────────────────────────────────────────────────────
# Sales register + CN loaders
# ────────────────────────────────────────────────────────────────────────────

def _iter_sheet(path: Path, sheet: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close(); return
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=3, values_only=True):
        yield row
    wb.close()


def load_excel_sales() -> tuple[list, list]:
    sales, cns = [], []
    for p in [SALES_XLSX_APR, SALES_XLSX_MAY]:
        for r in _iter_sheet(p, "Sales Report"):
            if not r or len(r) < 18: continue
            d = to_date(r[0])
            comp = str(r[10] or "").strip().upper()
            if comp not in ("CFPL", "CDPL"): continue
            try:
                pcs = float(r[15]) if r[15] is not None else None
                kg = float(r[17]) if r[17] is not None else None
            except (TypeError, ValueError): continue
            if d is None or kg is None: continue
            # Customer name in col 2 (Common Customer Name); fallback to col 1
            customer = r[2] or r[1]
            unit = voucher_to_unit(comp, r[9], customer)
            sales.append({"date": d, "entity": comp, "unit": unit,
                          "article": r[4], "qty_pcs": pcs, "kg": kg,
                          "doc_no": r[12], "customer": customer})
        for r in _iter_sheet(p, "Cancel Inv"):
            if not r or len(r) < 18: continue
            d = to_date(r[0])
            comp = str(r[10] or "").strip().upper()
            if comp not in ("CFPL", "CDPL"): continue
            try:
                pcs = float(r[15]) if r[15] is not None else None
                kg = float(r[17]) if r[17] is not None else None
            except (TypeError, ValueError): continue
            if d is None or kg is None: continue
            customer = r[2] or r[1]
            unit = voucher_to_unit(comp, r[9], customer)
            cns.append({"date": d, "entity": comp, "unit": unit,
                        "article": r[4], "qty_pcs": pcs, "kg": kg,
                        "doc_no": r[12], "customer": customer, "source": "Excel"})
    return sales, cns


def load_pdf_cns(seen_doc_nos: set) -> list:
    """Load CN data from PDF extractions, deduped against seen_doc_nos."""
    out = []
    for path in [PDF_UNITS_CSV, CN_PDF_CSV]:
        if not path.exists(): continue
        with open(path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                if "doc_type" in row and row.get("doc_type") != "CN":
                    continue
                doc_no = row.get("doc_no") or row.get("cn_no")
                product = row.get("product")
                if not doc_no: continue
                key = (str(doc_no).strip().upper(), str(product or "").strip().lower())
                if key in seen_doc_nos: continue
                seen_doc_nos.add(key)
                try: kg = float(row.get("qty_kg") or 0)
                except (TypeError, ValueError): continue
                if kg <= 0: continue
                entity = (row.get("entity") or "").upper()
                buyer = row.get("buyer") or ""
                # Per user: CDPL ≠ A-185. Use parser's unit if known, else default W-202.
                unit = (row.get("unit") or "").upper() or "W-202"
                if unit not in UNITS: unit = "W-202"
                out.append({"date": None, "entity": entity, "unit": unit,
                            "article": product, "qty_pcs": None, "kg": kg,
                            "doc_no": doc_no, "customer": buyer,
                            "source": f"PDF:{row.get('source_pdf','')}"})
    return out


async def load_direct_out(conn, seen_invoice_keys: set) -> list:
    """Cold-storage Direct-Out dispatches with explicit per-line warehouse
    attribution (Savla / Rishi). Each line is treated as a sale, with the unit
    set from the line's `warehouse` (NOT from voucher/customer rule).

    Dedup: skip if (invoice_no, article) already seen in Excel sales register
    (avoids double-counting when an invoice appears in both places).
    """
    import json as _json
    out = []
    for prefix, entity in (("cfpl", "CFPL"), ("cdpl", "CDPL")):
        rows = await conn.fetch(f"""
            SELECT transaction_no, entry_date, warehouse AS hdr_warehouse,
                   invoice_no, to_customer, lines
            FROM {prefix}_cold_storage_direct_out
            WHERE entry_date >= $1::date
        """, WINDOW_START)
        for r in rows:
            inv = r["invoice_no"] or r["transaction_no"]
            lines = r["lines"]
            if isinstance(lines, str): lines = _json.loads(lines)
            for ln in (lines or []):
                art = ln.get("item_description") or ln.get("item_mark")
                if not art: continue
                qty = float(ln.get("issue_qty") or 0)
                wt_per = float(ln.get("weight_kg_per_box") or 0)
                kg = qty * wt_per if wt_per else qty
                if kg <= 0: continue
                # Dedup: same invoice + same article = same line that already
                # exists in Excel sales register — skip.
                dedup_key = (str(inv).strip().upper(), norm(art))
                if dedup_key in seen_invoice_keys:
                    continue
                wh = ln.get("warehouse") or r["hdr_warehouse"] or ""
                unit = warehouse_to_unit(wh)
                if unit == "OTHER":
                    # Try the 'unit' field (e.g. 'D-39' → Savla)
                    unit = warehouse_to_unit(ln.get("unit") or "")
                out.append({
                    "date": r["entry_date"], "entity": entity, "unit": unit,
                    "article": art, "qty_pcs": qty, "kg": kg,
                    "doc_no": inv, "customer": r["to_customer"] or "",
                    "source": "DirectOut",
                })
    return out


ALL_INVOICES_CSV = Path(__file__).parent / "_all_invoices_extracted.csv"


def load_invoice_extracts(seen_doc_nos: set) -> tuple[list, list]:
    """Load ALL invoice + CN line items extracted from the zip files and
    combined PDFs (`_all_invoices_extracted.csv`).

    Dedup against the Excel sales register by `doc_no` ONLY — if an invoice
    number is already present in the sales register, ALL its line items are
    assumed to be captured there (article names may differ slightly between
    PDF text extraction and Tally export, so per-article dedup misses).
    Returns (new_inv_sales, new_cn_returns).
    """
    new_inv, new_cn = [], []
    if not ALL_INVOICES_CSV.exists():
        return new_inv, new_cn
    with open(ALL_INVOICES_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            doc_no = (row.get("doc_no") or "").strip()
            article = row.get("product") or ""
            if not doc_no or not article:
                continue
            if doc_no.upper() in seen_doc_nos:
                continue
            try: kg = float(row.get("qty_kg") or 0)
            except (TypeError, ValueError): continue
            if kg <= 0: continue
            entity = (row.get("entity") or "").upper() or "CFPL"
            buyer = row.get("buyer") or ""
            # Unit resolution priority for PDF rows:
            #   1. Customer-name override (Reliance→W-202, Avenue→A-185)
            #   2. Document-no prefix (CF53/→F-53, CDPL/→A-185)
            #   3. PDF parser's scanned unit (if recognised)
            #   4. Entity default (CFPL→W-202, CDPL→A-185)
            unit = voucher_to_unit(entity, None, customer=buyer, doc_no=doc_no)
            # If the PDF parser found a *specific* warehouse (D-39 / Cold Storage /
            # A-68 etc.) and our rules above didn't override on customer/doc_no,
            # prefer the PDF's reading over the entity default.
            scanned_unit = warehouse_to_unit(row.get("unit") or "")
            cust_upper = buyer.upper()
            cust_override = (
                "AVENUE SUPERMART" in cust_upper or "DMART" in cust_upper
                or "D-MART" in cust_upper or "RELIANCE" in cust_upper
                or doc_no.upper().startswith(("CF53/", "CDPL/"))
            )
            if not cust_override and scanned_unit != "OTHER":
                unit = scanned_unit
            d = None
            try:
                d = datetime.strptime((row.get("date") or "").strip(),
                                       "%d-%b-%y").date()
            except (TypeError, ValueError):
                pass
            # Filter: only include PDF lines within our calculation window.
            # The bulk zip contains invoices going back to Oct 2024 — those
            # belong to prior closing periods and would inflate the window.
            if d is None or d < WINDOW_START:
                continue
            entry = {"date": d, "entity": entity, "unit": unit,
                     "article": article, "qty_pcs": None, "kg": kg,
                     "doc_no": doc_no, "customer": buyer,
                     "source": f"PDF:{row.get('source_pdf','')}"}
            if (row.get("doc_type") or "").upper() == "CN":
                new_cn.append(entry)
            else:
                new_inv.append(entry)
    return new_inv, new_cn


# ────────────────────────────────────────────────────────────────────────────
# Image-based outward + adjustment model
# ────────────────────────────────────────────────────────────────────────────

def classify_sales(lines: list, sku_info: dict, bom: dict, aliases: dict):
    """Build 4 per-unit buckets per group:
       outward     — every real (external) sale, deducted from its own item_group
       adj_in_fg   — for matched FG sales: added back to FG group (cancels outward)
       adj_out_rm  — for matched FG sales: deducted from each BOM ingredient's RM group

    Inter-company invoices (buyer = Candor Foods / Candor Dates / CFPL / CDPL)
    are stock-transfer documents only — they DON'T deduct stock. They are
    routed into a separate `intercompany` list returned alongside the buckets
    for display in their own tab.

    Dates default routing: when no explicit warehouse signal (no Reliance / Avenue /
    CF53/ / CDPL/ doc-no prefix and no PDF-scanned unit), dates land in SAVLA
    (cold storage) instead of the entity default (W-202). Most date dispatches
    physically leave from Savla/Rishi.
    """
    outward = {u: defaultdict(float) for u in UNITS}
    adj_in_fg = {u: defaultdict(float) for u in UNITS}     # +
    adj_out_rm = {u: defaultdict(float) for u in UNITS}    # +  (subtract from closing)
    counts = {"direct_rm": 0, "fg_bom": 0, "fg_no_bom": 0, "inferred": 0,
              "unknown": 0, "intercompany": 0}
    unmapped: list = []
    # Per-cell record details for hover tooltips: details[unit][group] = [(label, kg), ...]
    details = {u: defaultdict(list) for u in UNITS}
    intercompany: list = []   # rows excluded from outward (stock-transfer invoices)
    for s in lines:
        art = s.get("article")
        kg = abs(float(s.get("kg") or 0))
        if not art or kg <= 0: continue
        unit = s.get("unit") or "OTHER"
        if unit not in UNITS: unit = "OTHER"
        key = norm(art)
        info = sku_info.get(key)
        own_group = smart_group(art, info)
        if info and info.get("type") == "pm":
            continue
        customer = (s.get("customer") or "").upper()
        doc_no_raw = (s.get("doc_no") or "")
        doc_no = str(doc_no_raw) if doc_no_raw else "?"
        date_s = str(s.get("date") or "?")

        # ── Inter-company stock-transfer invoices ──
        # These are billing-only transfers between our group companies, not real
        # outward sales. Capture them for the separate tab but skip them entirely
        # from the closing calculation.
        if is_intercompany(customer):
            intercompany.append({
                "doc_no": doc_no, "date": date_s, "entity": s.get("entity") or "",
                "customer": s.get("customer") or "", "unit": unit,
                "article": art, "kg": kg, "group": own_group,
            })
            counts["intercompany"] += 1
            continue

        # Use RECURSIVE BOM expansion — same-group intermediates (e.g. "Premia Trail
        # Mix Bulk" inside a Trail Mix FG) get replaced by their own BOM lines so
        # the FG's own group doesn't get hit twice (causing negative closing).
        b = expand_bom_recursive(key, bom, aliases, sku_info, fg_group=own_group) \
            if (info and info.get("type") == "fg") else bom_for(key, bom, aliases)
        pcs = s.get("qty_pcs")
        try: pcs = float(pcs) if pcs is not None else None
        except (TypeError, ValueError): pcs = None

        # Per user: F-53 (APMC) only physically dispatches processed FG packs.
        # Bulk / RM-type articles billed via APMC actually came from W-202 dispatch
        # or Cold Storage. Route to where the group's stock physically lives:
        #   dates / cranberry / blueberry / pista bulk → Cold (Savla)
        #   cashew / almond / walnut / peanuts / others → W-202 (central dispatch)
        if unit == "F-53":
            is_bulk_rm = (info and info.get("type") == "rm") or not (b and pcs and info and info.get("type") == "fg")
            if is_bulk_rm:
                # Groups with significant Savla/Rishi stock → Cold
                cold_groups = {"dates", "cranberry", "blueberry", "pista",
                               "raisin", "premium nuts", "seeds"}
                unit = "SAVLA" if own_group in cold_groups else "W-202"

        # Per user: most date dispatches physically leave from Savla/Rishi cold
        # storage. If the invoice didn't carry an explicit warehouse signal
        # (no Reliance/Avenue customer override, no CF53/CDPL doc-no prefix, and
        # we ended up at the W-202 entity default), reroute dates to SAVLA.
        explicit_signal = bool(
            customer and (
                "AVENUE SUPERMART" in customer or "DMART" in customer
                or "D-MART" in customer or "D MART" in customer
                or "RELIANCE" in customer
            )
        ) or doc_no.upper().startswith(("CF53/", "CDPL/"))
        if own_group == "dates" and unit == "W-202" and not explicit_signal:
            unit = "SAVLA"

        details[unit][own_group].append(
            (f"{doc_no} | {date_s} | {art[:45]}", kg)
        )

        # Step 1: outward always hits the article's own group
        outward[unit][own_group] += kg

        # Step 2: if FG with BOM → adjustment (cancel outward, redistribute to RM)
        if b and pcs and info and info.get("type") == "fg":
            adj_in_fg[unit][own_group] += kg
            for rm_name, q_per, it in b:
                if it == "pm": continue
                rm_kg = q_per * pcs
                rm_info = sku_info.get(norm(rm_name))
                rm_grp = smart_group(rm_name, rm_info)
                adj_out_rm[unit][rm_grp] += rm_kg
            counts["fg_bom"] += 1
            continue

        if info and info.get("type") == "rm":
            counts["direct_rm"] += 1; continue
        if info and info.get("type") == "fg":
            counts["fg_no_bom"] += 1; continue
        if own_group != "_UNMAPPED":
            counts["inferred"] += 1
        else:
            counts["unknown"] += 1
            unmapped.append((art, kg))
    return (
        {u: dict(g) for u, g in outward.items()},
        {u: dict(g) for u, g in adj_in_fg.items()},
        {u: dict(g) for u, g in adj_out_rm.items()},
        counts, unmapped,
        {u: dict(g) for u, g in details.items()},
        intercompany,
    )


# ────────────────────────────────────────────────────────────────────────────
# Auto-rearrangement: rebalance negatives against unit-wise availability
# ────────────────────────────────────────────────────────────────────────────

def auto_rearrange(opening_u, inward_u, inferred_in_u, txfr_in_u, txfr_out_u,
                   cn_u, cn_adj_in, cn_adj_out, outward_u, adj_in_u, adj_out_u,
                   all_groups):
    """For every (unit, group) that would close negative, transfer the deficit
    to another unit that has positive availability for that group.

    Returns `rearr_u`: per-unit per-group balancing adjustment (sums to zero
    across units per group). Negative on the donor side (stock leaves it),
    positive on the recipient side (stock arrives). Added to the closing
    formula as a new term so the data-entry mistakes (sale billed from the
    wrong warehouse) are auto-corrected.
    """
    # Compute initial availability per (unit, group)
    avail = {u: {} for u in UNITS}
    for u in UNITS:
        for g in all_groups:
            avail[u][g] = (
                opening_u.get(u, {}).get(g, 0)
                + inward_u.get(u, {}).get(g, 0)
                + inferred_in_u.get(u, {}).get(g, 0)
                + txfr_in_u.get(u, {}).get(g, 0)
                - txfr_out_u.get(u, {}).get(g, 0)
                + cn_u.get(u, {}).get(g, 0)
                - cn_adj_in.get(u, {}).get(g, 0)
                + cn_adj_out.get(u, {}).get(g, 0)
                - outward_u.get(u, {}).get(g, 0)
                + adj_in_u.get(u, {}).get(g, 0)
                - adj_out_u.get(u, {}).get(g, 0)
            )
    rearr_u = {u: defaultdict(float) for u in UNITS}
    for g in all_groups:
        # Iterate until no more deficits can be cured for this group
        for _ in range(len(UNITS)):
            # Find the deepest-negative unit
            deficit_unit, deficit_amt = None, 0.0
            for u in UNITS:
                cur = avail[u][g] + rearr_u[u][g]
                if cur < deficit_amt - 0.5:
                    deficit_unit, deficit_amt = u, cur
            if deficit_unit is None: break
            need = -deficit_amt   # positive kg to pull in
            # Find donors with current positive availability, richest first
            donors = [
                (u, avail[u][g] + rearr_u[u][g])
                for u in UNITS if u != deficit_unit
            ]
            donors = [(u, v) for u, v in donors if v > 0.5]
            donors.sort(key=lambda x: -x[1])
            if not donors: break
            for u, have in donors:
                move = min(need, have)
                rearr_u[deficit_unit][g] += move
                rearr_u[u][g] -= move
                need -= move
                if need <= 0.5: break
            if need > 0.5:
                # No more donors — group is genuinely short across all units
                break
    return {u: dict(g) for u, g in rearr_u.items()}


# ────────────────────────────────────────────────────────────────────────────
# HTML — with hover tooltips
# ────────────────────────────────────────────────────────────────────────────

COL_TIP = {
    "Opening": "Stock on 31-Mar-2026 from Physical Stock Compilation Excel (combines RM + FG)",
    "+ Inward": "PO receipts from cfpl/cdpl_boxes_v2 ⨝ transactions_v2 with rtv=false, service=false",
    "+ Inf-In": "Inferred inward (Savla/Rishi only): when total deductions for a cold-storage (unit, group) would push closing below zero, the deficit is plugged here. Represents stock movements not captured in IMS — most commonly untracked floor moves into Savla/Rishi outposts.",
    "+ Xfr-In": "Material received at this unit via interunit_transfers (header.to_site)",
    "− Xfr-Out": "Material dispatched from this unit via interunit_transfers (header.from_site)",
    "+ CN": "Customer credit notes — Excel Cancel Inv + PDF-parsed CN docs (returns coming back in)",
    "− Outward": "All sales as-is, deducted from the article's own group (Sales Report 'In Kg'). Excludes inter-company stock-transfer invoices (Candor Foods / Candor Dates as buyer).",
    "+ Adj-In": "BOM conversion adjustment: for each FG sale with a BOM, cancels the outward by adding back the FG mass (so FG-only groups like Trail Mix don't crater)",
    "− Adj-Out": "BOM conversion adjustment: ingredients consumed by FG production, deducted from each RM source group (e.g. almonds & cashews consumed by Trail Mix sales)",
    "± Rearr": "Auto-rearrangement: when a unit's (group) would close negative, the deficit is sourced from another unit that physically holds the stock. Sums to zero per group across units. Models data-entry mistakes where invoices were billed against the wrong warehouse.",
    "= Closing": "= Opening + Inward + Inf-In + Xfr-In − Xfr-Out + CN − Outward + Adj-In − Adj-Out ± Rearr",
}


def write_html_report(opening_u, inward_u, inferred_in_u, txfr_in_u, txfr_out_u, cn_u,
                       outward_u, adj_in_u, adj_out_u, rearr_u, closing_u, *,
                       sales_count, cn_excel_count, cn_pdf_count,
                       out_classification, inward_unmapped, top_unmapped_out,
                       sales_details=None, cn_details=None,
                       intercompany_lines=None):
    sales_details = sales_details or {}
    cn_details = cn_details or {}
    intercompany_lines = intercompany_lines or []

    def hover_top(units, group, det_dict, n=5) -> str:
        """Build a multi-line hover string of top-N contributing records."""
        recs = []
        for u in units:
            recs.extend(det_dict.get(u, {}).get(group, []))
        if not recs: return ""
        recs.sort(key=lambda x: -x[1])
        lines = [f"Top {min(n, len(recs))} of {len(recs)} contributors:"]
        for label, kg in recs[:n]:
            lines.append(f" • {fmt(kg):>8} kg  {label}")
        return "\n" + "\n".join(lines)
    all_groups = set()
    for b in (opening_u, inward_u, inferred_in_u, txfr_in_u, txfr_out_u, cn_u,
              outward_u, adj_in_u, adj_out_u, rearr_u, closing_u):
        for u in b:
            for g in b[u]: all_groups.add(g)
    groups = sorted(all_groups)

    def sum_units(bucket, units, g): return sum(bucket.get(u, {}).get(g, 0) for u in units)

    def render_view(units):
        rows = []; tot = [0.0] * 11   # op, iw, infi, tin, tout, cn, ow, ain, aout, rearr, cl
        for g in groups:
            op = sum_units(opening_u, units, g)
            iw = sum_units(inward_u, units, g)
            infi = sum_units(inferred_in_u, units, g)
            tin = sum_units(txfr_in_u, units, g)
            tout = sum_units(txfr_out_u, units, g)
            cn = sum_units(cn_u, units, g)
            ow = sum_units(outward_u, units, g)
            ain = sum_units(adj_in_u, units, g)
            aout = sum_units(adj_out_u, units, g)
            rr = sum_units(rearr_u, units, g)
            cl = sum_units(closing_u, units, g)
            if all(abs(x) < 0.5 for x in [op, iw, infi, tin, tout, cn, ow, ain, aout, rr, cl]):
                continue
            cl_cls = "cl neg" if cl < -0.5 else "cl"
            cl_tip = (
                f"Opening {fmt(op)} + Inward {fmt(iw)} + Inf-In {fmt(infi)} + Xfr-In {fmt(tin)} "
                f"− Xfr-Out {fmt(tout)} + CN {fmt(cn)} − Outward {fmt(ow)} "
                f"+ Adj-In {fmt(ain)} − Adj-Out {fmt(aout)} + Rearr {fmt(rr)} = {fmt(cl)}"
            )
            t_op = COL_TIP["Opening"]; t_iw = COL_TIP["+ Inward"]
            t_infi = COL_TIP["+ Inf-In"]
            t_tin = COL_TIP["+ Xfr-In"]; t_tout = COL_TIP["− Xfr-Out"]
            t_cn = COL_TIP["+ CN"]; t_ow = COL_TIP["− Outward"]
            t_ain = COL_TIP["+ Adj-In"]; t_aout = COL_TIP["− Adj-Out"]
            t_rr = COL_TIP["± Rearr"]
            infi_cls = "adj" if infi > 0.5 else ""
            rr_cls = "adj" if rr > 0.5 else ("adj neg" if rr < -0.5 else "")
            # Per-cell hover with top contributors (transactions / invoices)
            ow_drill = hover_top(units, g, sales_details)
            cn_drill = hover_top(units, g, cn_details)
            ain_drill = hover_top(units, g, sales_details)   # adj_in driven by same FG sales
            aout_drill = hover_top(units, g, sales_details)  # adj_out driven by same FG sales
            ow_full = (t_ow + ow_drill).replace("'", "&apos;")
            cn_full = (t_cn + cn_drill).replace("'", "&apos;")
            ain_full = (t_ain + ain_drill).replace("'", "&apos;")
            aout_full = (t_aout + aout_drill).replace("'", "&apos;")
            rows.append(
                f"<tr><td class='grp'>{g}</td>"
                f"<td title='{t_op}'>{fmt(op)}</td>"
                f"<td title='{t_iw}'>{fmt(iw)}</td>"
                f"<td title='{t_infi}' class='{infi_cls}'>{fmt(infi)}</td>"
                f"<td title='{t_tin}'>{fmt(tin)}</td>"
                f"<td title='{t_tout}'>{fmt(tout)}</td>"
                f"<td title='{cn_full}'>{fmt(cn)}</td>"
                f"<td title='{ow_full}'>{fmt(ow)}</td>"
                f"<td title='{ain_full}' class='adj'>{fmt(ain)}</td>"
                f"<td title='{aout_full}' class='adj'>{fmt(aout)}</td>"
                f"<td title='{t_rr}' class='{rr_cls}'>{fmt(rr)}</td>"
                f"<td class='{cl_cls}' title='{cl_tip}'>{fmt(cl)}</td></tr>"
            )
            for i, v in enumerate([op, iw, infi, tin, tout, cn, ow, ain, aout, rr, cl]):
                tot[i] += v
        rows.append(
            f"<tr class='tot-row'><td class='grp tot'>TOTAL</td>"
            + "".join(f"<td class='tot'>{fmt(v)}</td>" for v in tot[:10])
            + f"<td class='cl tot'>{fmt(tot[10])}</td></tr>"
        )
        return "".join(rows), tot

    # Units are PHYSICAL LOCATIONS, not companies. A-185 hosts both CFPL & CDPL
    # stock; Savla likewise. CFPL vs CDPL is the legal entity that owns the
    # invoice — separate concern, not a location grouping.
    view_defs = [
        ("total", "Consolidated (All Units)", UNITS),
        ("w202", "W-202", ["W-202"]),
        ("a185", "A-185", ["A-185"]),
        ("f53", "F-53 (APMC)", ["F-53"]),
        ("a68", "A-68", ["A-68"]),
        ("rishi", "Rishi", ["RISHI"]),
        ("savla", "Savla (incl. Cold Storage / Supreme Cold)", ["SAVLA"]),
        ("other", "Other (null-warehouse / adjustment-pending)", ["OTHER"]),
    ]

    view_tables = {vid: render_view(units) for vid, _label, units in view_defs}
    tot = view_tables["total"][1]
    grand_closing = tot[10]; grand_opening = tot[0]; grand_inward = tot[1]; grand_outward = tot[6]

    buttons = "".join(
        f"<button class=\"{'active' if vid == 'total' else ''}\" "
        f"data-view='{vid}' onclick=\"switchView('{vid}')\""
        f" title='{label}'>{label.split('(')[0].strip()}</button>"
        for vid, label, _ in view_defs
    )

    h_op = COL_TIP["Opening"]; h_iw = COL_TIP["+ Inward"]
    h_infi = COL_TIP["+ Inf-In"]
    h_tin = COL_TIP["+ Xfr-In"]; h_tout = COL_TIP["− Xfr-Out"]
    h_cn = COL_TIP["+ CN"]; h_ow = COL_TIP["− Outward"]
    h_ain = COL_TIP["+ Adj-In"]; h_aout = COL_TIP["− Adj-Out"]
    h_rr = COL_TIP["± Rearr"]
    h_cl = COL_TIP["= Closing"]
    thead = (
        f"<thead><tr>"
        f"<th title='{h_op}'>Group</th>"
        f"<th title='{h_op}'>Opening</th>"
        f"<th title='{h_iw}'>+ Inward</th>"
        f"<th title='{h_infi}' class='adj'>+ Inf-In</th>"
        f"<th title='{h_tin}'>+ Xfr-In</th>"
        f"<th title='{h_tout}'>− Xfr-Out</th>"
        f"<th title='{h_cn}'>+ CN</th>"
        f"<th title='{h_ow}'>− Outward</th>"
        f"<th title='{h_ain}' class='adj'>+ Adj-In</th>"
        f"<th title='{h_aout}' class='adj'>− Adj-Out</th>"
        f"<th title='{h_rr}' class='adj'>± Rearr</th>"
        f"<th title='{h_cl}'>= Closing</th>"
        f"</tr></thead>"
    )
    views_html = "".join(
        f"<div class=\"view {'active' if vid == 'total' else ''}\" id='view-{vid}'>"
        f"<table>{thead}<tbody>{view_tables[vid][0]}</tbody></table></div>"
        for vid, _label, _ in view_defs
    )

    js_summaries = ", ".join(
        f"{vid}: {{closing:'{fmt(view_tables[vid][1][10])}', "
        f"opening:'{fmt(view_tables[vid][1][0])}', "
        f"inward:'{fmt(view_tables[vid][1][1])}', "
        f"outward:'{fmt(view_tables[vid][1][6])}'}}"
        for vid, _, _ in view_defs
    )
    js_hints = ", ".join(f"{vid}: '{label}'" for vid, label, _ in view_defs)

    unmapped_html = ""
    if top_unmapped_out:
        items = sorted(top_unmapped_out, key=lambda x: -x[1])[:12]
        rs = "".join(f"<tr><td>{a}</td><td class='num'>{fmt(v)}</td></tr>" for a, v in items)
        unmapped_html = (
            f"<section class='card'><h3>Top unmapped sale SKUs</h3>"
            f"<p class='hint'>These articles don't match all_sku.item_group and aren't matched by keyword inference. "
            f"They land in <code>_UNMAPPED</code> until classification is added.</p>"
            f"<table class='small'><thead><tr><th>Article</th><th>Sale KG</th></tr></thead>"
            f"<tbody>{rs}</tbody></table></section>"
        )

    # ── Intercompany invoices tab (Candor Foods / Candor Dates as buyer) ──
    intercompany_html = ""
    if intercompany_lines:
        total_ic_kg = sum(l["kg"] for l in intercompany_lines)
        # Group by group + entity for a digest, then list rows underneath
        by_group_entity: dict = defaultdict(lambda: defaultdict(float))
        for l in intercompany_lines:
            by_group_entity[l["group"]][l["entity"]] += l["kg"]
        digest_rows = []
        for g in sorted(by_group_entity):
            for e in sorted(by_group_entity[g]):
                digest_rows.append(
                    f"<tr><td>{g}</td><td>{e}</td>"
                    f"<td class='num'>{fmt(by_group_entity[g][e])}</td></tr>"
                )
        detail_rows = []
        for l in sorted(intercompany_lines, key=lambda x: -x["kg"])[:200]:
            detail_rows.append(
                f"<tr><td>{l['date']}</td><td>{l['doc_no']}</td>"
                f"<td>{l['entity']}</td><td>{l['customer']}</td>"
                f"<td>{l['unit']}</td><td>{l['group']}</td>"
                f"<td>{(l['article'] or '')[:60]}</td>"
                f"<td class='num'>{fmt(l['kg'])}</td></tr>"
            )
        intercompany_html = (
            "<section class='card'>"
            "<h2>Inter-company Stock-Transfer Invoices</h2>"
            "<p class='hint'>Sales billed to <b>Candor Foods Pvt Ltd</b> or <b>Candor Dates Pvt Ltd</b> "
            "are internal stock-transfer documents, not real outward movement. "
            "<b>These rows are excluded from the Outward / Closing calculation</b> and are listed here "
            "for visibility only.</p>"
            f"<p class='hint'><b>{len(intercompany_lines):,} line items</b>, total <b>{fmt(total_ic_kg)} kg</b>.</p>"
            "<h3>Digest by group × entity</h3>"
            "<table class='small'><thead><tr><th>Group</th><th>Entity</th>"
            "<th>KG</th></tr></thead><tbody>"
            f"{''.join(digest_rows)}</tbody></table>"
            "<h3 style='margin-top:14px'>Top 200 line items (by KG)</h3>"
            "<table class='small'><thead><tr><th>Date</th><th>Doc No</th>"
            "<th>Entity</th><th>Buyer</th><th>Unit</th><th>Group</th>"
            "<th>Article</th><th>KG</th></tr></thead><tbody>"
            f"{''.join(detail_rows)}</tbody></table>"
            "</section>"
        )

    html = f"""<!doctype html>
<html lang='en'><head><meta charset='utf-8'>
<title>Candor Foods — Closing Stock v5 — Per-Unit + Adjustment (Apr 1 → May 12, 2026)</title>
<style>
  :root {{ --bg:#f7f7f5; --card:#fff; --border:#d9d9d6; --accent:#16513f;
            --neg:#b6260d; --tot-bg:#eef3f0; --muted:#666; --adj:#7d5e2a; }}
  body {{ font-family:-apple-system,"Segoe UI",sans-serif; background:var(--bg);
          margin:0; padding:24px; color:#222; }}
  h1 {{ margin:0 0 4px; font-size:22px; color:var(--accent); }}
  .sub {{ color:var(--muted); margin-bottom:24px; font-size:13px; }}
  .card {{ background:var(--card); border:1px solid var(--border); border-radius:8px;
            padding:18px 22px; margin-bottom:18px; }}
  .summary {{ display:grid; grid-template-columns:repeat(4,1fr); gap:14px; }}
  .summary .lbl {{ color:var(--muted); font-size:11px; text-transform:uppercase;
                    letter-spacing:0.05em; }}
  .summary .val {{ font-size:24px; font-weight:600; color:var(--accent); }}
  .summary .val.neg {{ color:var(--neg); }}
  h2 {{ margin:0 0 12px; font-size:16px; color:var(--accent); }}
  h3 {{ margin:0 0 10px; font-size:14px; color:var(--accent); }}
  table {{ border-collapse:collapse; width:100%; font-size:13px; font-variant-numeric:tabular-nums; }}
  th,td {{ padding:6px 8px; text-align:right; border-bottom:1px solid #f0f0ee; }}
  th {{ background:#f4f5f1; font-weight:600; color:#444; font-size:11px;
        text-transform:uppercase; letter-spacing:0.04em; cursor:help; }}
  th.adj {{ background:#fdf6e3; color:var(--adj); }}
  td.grp, th:first-child {{ text-align:left; }}
  td.grp {{ font-weight:500; }}
  td.cl {{ font-weight:600; color:var(--accent); cursor:help; }}
  td.cl.neg, td.neg {{ color:var(--neg); }}
  td.adj {{ background:#fdf6ea; color:var(--adj); }}
  td.num {{ text-align:right; }}
  tr.tot-row td {{ background:var(--tot-bg); border-top:2px solid #ccc; font-weight:600; }}
  td.tot {{ font-weight:600; }}
  td[title], th[title] {{ cursor:help; }}
  table.small {{ font-size:12px; }}
  table.small td:first-child {{ text-align:left; }}
  .toggle {{ display:flex; flex-wrap:wrap; gap:6px; }}
  .toggle button {{ background:#eef0eb; border:1px solid transparent;
                     padding:6px 12px; font-size:12px; color:#555;
                     border-radius:14px; cursor:pointer; font-family:inherit;
                     font-weight:500; transition:all 0.15s; }}
  .toggle button:hover {{ color:var(--accent); border-color:var(--accent); }}
  .toggle button.active {{ background:var(--accent); color:#fff;
                            border-color:var(--accent); }}
  .toggle-row {{ display:flex; align-items:start; gap:18px; margin-bottom:14px;
                  flex-wrap:wrap; }}
  .toggle-row h2 {{ margin:0; flex-shrink:0; }}
  .view-hint {{ font-size:12px; color:var(--muted); margin-bottom:10px;
                  background:#f8f8f5; padding:6px 10px; border-radius:4px;
                  border-left:3px solid var(--accent); }}
  .view {{ display:none; }}
  .view.active {{ display:block; }}
  .formula {{ font-family:ui-monospace,Consolas,monospace; background:#f0f1ed;
              padding:12px 14px; border-radius:6px; font-size:12.5px; line-height:1.6; }}
  .formula .lbl {{ color:var(--accent); font-weight:600; }}
  .formula .adj {{ color:var(--adj); font-weight:600; }}
  .pill {{ display:inline-block; background:#e2eae5; color:var(--accent);
            padding:2px 8px; border-radius:10px; font-size:11px; margin-right:4px; }}
  .legend {{ display:flex; gap:24px; margin-top:10px; font-size:12px;
              color:var(--muted); flex-wrap:wrap; }}
  .legend-item {{ display:flex; align-items:center; gap:6px; }}
  .swatch {{ width:14px; height:14px; border-radius:3px;
              display:inline-block; border:1px solid #ddd; }}
  footer {{ color:var(--muted); font-size:11px; margin-top:24px; text-align:center; }}
  .hint {{ font-size:12px; color:var(--muted); margin:0 0 10px; }}
</style></head><body>

<h1>Candor Foods — Closing Stock Report (v5)</h1>
<div class='sub'>
  Window <b>2026-04-01 → 2026-05-12</b> &nbsp;·&nbsp; All quantities in KG &nbsp;·&nbsp;
  Hover any column header, cell, or toggle button for explanation
</div>

<div class='card summary'>
  <div><div class='lbl' title='= Opening + Inward + Xfr-In − Xfr-Out + CN − Outward + Adj-In − Adj-Out'>Closing Stock</div>
       <div class='val' id='sum-closing'>{fmt(grand_closing)} kg</div></div>
  <div><div class='lbl' title='Physical stock take on 31-Mar-2026 (CFPL + CDPL combined per group)'>Opening (31-Mar)</div>
       <div class='val' id='sum-opening'>{fmt(grand_opening)} kg</div></div>
  <div><div class='lbl' title='PO receipts in window (cfpl/cdpl_boxes_v2)'>Inward (Apr 1+)</div>
       <div class='val' id='sum-inward'>{fmt(grand_inward)} kg</div></div>
  <div><div class='lbl' title='All sales in window (Sales Report — In Kg)'>Outward (Apr 1+)</div>
       <div class='val' id='sum-outward'>{fmt(grand_outward)} kg</div></div>
</div>

<div class='card'>
  <h2>Formula &amp; the BOM-conversion adjustment</h2>
  <div class='formula'>
    <span class='lbl'>Closing[group, unit]</span> &nbsp;=&nbsp; Opening + Inward + Inf-In + Xfr-In − Xfr-Out
    + CN − Outward <span class='adj'>+ Adj-In − Adj-Out ± Rearr</span>
    <br><br>
    <span style='color:#666'>
      <b>Outward</b> is every sale deducted from the article's own group (Trail Mix sale → Trail Mix down).
      Inter-company invoices (Candor Foods / Candor Dates as buyer) are excluded — they're billing-only transfers.<br>
      <b><span class='adj'>Adj-In</span></b> cancels the FG outward (Trail Mix sale → Trail Mix back up).<br>
      <b><span class='adj'>Adj-Out</span></b> deducts the BOM-derived ingredients from each source RM group
      (Trail Mix sale → almond, cashew, raisin etc. all go down per the BOM recipe).<br>
      <b><span class='adj'>± Rearr</span></b> auto-rebalances cells that would close negative against units with
      stock for the same group (data-entry mistakes where the wrong dispatch warehouse was on the invoice).
      Sums to zero per group across units.<br>
      <br>
      Net effect: FG-only groups (Trail Mix, Bars, Hampers) stay flat; the underlying RM groups absorb the real cost.
    </span>
  </div>
  <div class='legend'>
    <div class='legend-item'><span class='swatch' style='background:#fdf6e3'></span> Adjustment columns (BOM conversion)</div>
    <div class='legend-item'><span class='swatch' style='background:#fff;color:var(--neg);text-align:center;font-size:9px;line-height:14px;border-color:var(--neg)'>−</span> Negative closing (data gap or sales &gt; available)</div>
    <div class='legend-item'><span class='swatch' style='background:var(--accent)'></span> Active unit (toggle)</div>
  </div>
</div>

<div class='card'>
  <div class='toggle-row'>
    <h2>Closing Stock per RM Group &nbsp; <small style='font-size:11px;color:#888;font-weight:normal'>(hover header / cells for tooltips)</small></h2>
    <div class='toggle'>{buttons}</div>
  </div>
  <div class='view-hint' id='view-hint'>Showing Consolidated (All Units).</div>
  {views_html}
</div>

<div class='card'>
  <h2>Sale-line classification</h2>
  <p>
    <span class='pill' title='Articles tagged item_type=rm in all_sku — deducted from the article's own group'>Direct RM sales: {out_classification.get('direct_rm', 0)}</span>
    <span class='pill' title='FG sales with a BOM match — outward deducted from FG group, then BOM-redistributed to source RM groups via adjustment'>FG with BOM (converted): {out_classification.get('fg_bom', 0)}</span>
    <span class='pill' title='FG sales without BOM match — deducted from FG group only (no adjustment)'>FG without BOM: {out_classification.get('fg_no_bom', 0)}</span>
    <span class='pill' title='Article not in all_sku but matched by keyword inference (e.g. \"cashew\" in name)'>Keyword-inferred: {out_classification.get('inferred', 0)}</span>
    <span class='pill' title='Article not classifiable — went into _UNMAPPED group'>Unknown: {out_classification.get('unknown', 0)}</span>
  </p>
  <p style='font-size:12px;color:var(--muted)'>
    Sales lines: {sales_count:,} &nbsp;·&nbsp;
    CN from Excel: {cn_excel_count:,} &nbsp;·&nbsp;
    CN from PDFs: {cn_pdf_count:,} (deduped by doc_no + product)
  </p>
</div>

{unmapped_html}

{intercompany_html}

<script>
  const summaries = {{ {js_summaries} }};
  const hints     = {{ {js_hints} }};
  function switchView(view) {{
    document.querySelectorAll('.toggle button').forEach(b => {{
      b.classList.toggle('active', b.dataset.view === view);
    }});
    document.querySelectorAll('.view').forEach(v => {{
      v.classList.toggle('active', v.id === 'view-' + view);
    }});
    const s = summaries[view];
    document.getElementById('sum-closing').textContent = s.closing + ' kg';
    document.getElementById('sum-opening').textContent = s.opening + ' kg';
    document.getElementById('sum-inward').textContent  = s.inward  + ' kg';
    document.getElementById('sum-outward').textContent = s.outward + ' kg';
    document.getElementById('view-hint').textContent = 'Showing ' + hints[view] + '.';
  }}
</script>

<footer>Generated by <code>_closing_stock_calc_v5_probe.py</code> · Hover any header / cell for explanation</footer>
</body></html>
"""
    OUT_HTML.write_text(html, encoding="utf-8")


# ────────────────────────────────────────────────────────────────────────────
# Main
# ────────────────────────────────────────────────────────────────────────────

async def main():
    print("Loading opening per unit (CFPL+CDPL sheets)...")
    opening_u = load_opening_per_unit()

    print("Loading Excel sales + CNs...")
    sales, excel_cns = load_excel_sales()

    # Dedup CNs across Excel + PDFs by (doc_no, product)
    seen = set()
    for cn in excel_cns:
        if cn.get("doc_no"):
            seen.add((str(cn["doc_no"]).strip().upper(), norm(cn["article"])))
    print(f"  Sales lines: {len(sales)} | Excel CNs (deduped seed): {len(excel_cns)}")

    # Ingest the consolidated invoice + CN extract from ALL zip files & combined
    # PDFs (`_all_invoices_extracted.csv`, ~4,600 rows). Dedup by doc_no ONLY —
    # if the invoice number is already in the sales register / Cancel Inv, ALL
    # its line items are assumed to be already captured there (PDF text-extraction
    # may spell articles differently than the Tally export, so per-article dedup
    # misses too many true duplicates).
    seen_doc_nos = set()
    for s in sales + excel_cns:
        if s.get("doc_no"):
            seen_doc_nos.add(str(s["doc_no"]).strip().upper())

    pdf_inv, pdf_cns = load_invoice_extracts(seen_doc_nos)
    print(f"  PDF INV added (new invoices, dedup by doc_no): {len(pdf_inv)}")
    print(f"  PDF CN added (new CNs, dedup by doc_no):       {len(pdf_cns)}")
    sales = sales + pdf_inv

    pool = await connect_db()
    try:
        async with pool.acquire() as c:
            print("Loading all_sku + BOM master...")
            sku_info = await load_sku_info(c)
            bom = await load_bom_master(c)
            print(f"  all_sku rows: {len(sku_info)} | BOM FGs: {len(bom)}")

            print("Loading inward (null-warehouse → W-202 default)...")
            inward_u, inward_unmapped = await load_inward_per_unit(c, sku_info)

            print("Loading inter-unit transfers (including unacknowledged)...")
            txfr_out_u, txfr_in_u = await load_transfers_per_unit(c, sku_info)

            # Cold-storage Direct Out: per-line warehouse attribution
            print("Loading cold-storage Direct Out dispatches...")
            # Refresh dedup set to include the PDF-derived invoices just added
            sales_invoice_keys = set()
            for s in sales + excel_cns + pdf_cns:
                if s.get("doc_no"):
                    sales_invoice_keys.add(
                        (str(s["doc_no"]).strip().upper(), norm(s.get("article")))
                    )
            do_rows = await load_direct_out(c, sales_invoice_keys)
            print(f"  Direct Out lines added (dedup): {len(do_rows)}")
            sales = sales + do_rows
    finally:
        await pool.close()

    aliases = build_aliases(bom)
    print(f"  BOM aliases: {len(aliases)}")

    print("Outward + adjustment model (image-based)...")
    (outward_u, adj_in_u, adj_out_u, out_class, out_unmap, sales_details,
     intercompany_lines) = classify_sales(sales, sku_info, bom, aliases)
    print(f"  Inter-company invoices excluded: {len(intercompany_lines)}")
    print("CN inward + adjustment (returns)...")
    (cn_outward_u, cn_adj_in, cn_adj_out, _, _, cn_details,
     intercompany_cn) = classify_sales(excel_cns + pdf_cns, sku_info, bom, aliases)
    # Combine inter-company captures from both sales and CN passes for the
    # standalone report tab. CN-side inter-company is rare but possible.
    intercompany_lines = intercompany_lines + intercompany_cn
    # For CN: invert (CN inward = stock back). Reuse adjustment logic symmetrically:
    # cn_u[g] is "what the customer returned, deducted from FG group as inward bonus"
    # cn_adj swaps signs in calc.

    # ─── Compute closing per (unit, group) ───
    closing_u = {u: {} for u in UNITS}
    all_g = set()
    for b in (opening_u, inward_u, txfr_in_u, txfr_out_u, cn_outward_u,
              outward_u, adj_in_u, adj_out_u):
        for u in b:
            for g in b[u]: all_g.add(g)

    # Inferred-Inward reconciliation: ONLY for Savla & Rishi (per user) —
    # when Xfr-Out exceeds available stock for those cold-store outposts,
    # plug the deficit. All other units' negatives stay visible.
    inferred_in_u = {u: defaultdict(float) for u in UNITS}
    for u in ("SAVLA", "RISHI"):
        for g in all_g:
            op = opening_u.get(u, {}).get(g, 0)
            iw = inward_u.get(u, {}).get(g, 0)
            tin = txfr_in_u.get(u, {}).get(g, 0)
            tout = txfr_out_u.get(u, {}).get(g, 0)
            if tout > 0 and tout > (op + iw + tin):
                inferred_in_u[u][g] = tout - (op + iw + tin)
    inferred_in_u = {u: dict(g) for u, g in inferred_in_u.items()}

    # ─── Auto-rearrangement against per-unit availability ───
    # Move deficits to units with stock for that group (per user: most date
    # dispatches really came from Savla/Rishi; trail-mix RM negatives can be
    # absorbed by W-202 stock; etc.). Sums to zero per group.
    rearr_u = auto_rearrange(
        opening_u, inward_u, inferred_in_u, txfr_in_u, txfr_out_u,
        cn_outward_u, cn_adj_in, cn_adj_out, outward_u, adj_in_u, adj_out_u,
        all_g,
    )

    for u in UNITS:
        for g in all_g:
            closing_u[u][g] = (
                opening_u.get(u, {}).get(g, 0)
                + inward_u.get(u, {}).get(g, 0)
                + inferred_in_u.get(u, {}).get(g, 0)         # ← reconciliation plug
                + txfr_in_u.get(u, {}).get(g, 0)
                - txfr_out_u.get(u, {}).get(g, 0)
                + cn_outward_u.get(u, {}).get(g, 0)
                - cn_adj_in.get(u, {}).get(g, 0)
                + cn_adj_out.get(u, {}).get(g, 0)
                - outward_u.get(u, {}).get(g, 0)
                + adj_in_u.get(u, {}).get(g, 0)
                - adj_out_u.get(u, {}).get(g, 0)
                + rearr_u.get(u, {}).get(g, 0)               # ← auto-rebalance
            )

    # ─── Text output ───
    print()
    print(f"{'UNIT':<14} {'Opening':>10} {'+Inward':>10} {'+XfrIn':>9} {'-XfrOut':>9} "
          f"{'+CN':>8} {'-Outward':>10} {'+AdjIn':>9} {'-AdjOut':>9} {'±Rearr':>9} {'= Closing':>12}")
    print("-" * 122)
    for u in UNITS:
        op = sum(opening_u.get(u, {}).values())
        iw = sum(inward_u.get(u, {}).values())
        tin = sum(txfr_in_u.get(u, {}).values())
        tout = sum(txfr_out_u.get(u, {}).values())
        cn = sum(cn_outward_u.get(u, {}).values())
        ow = sum(outward_u.get(u, {}).values())
        ain = sum(adj_in_u.get(u, {}).values())
        aout = sum(adj_out_u.get(u, {}).values())
        rr = sum(rearr_u.get(u, {}).values())
        cl = sum(closing_u.get(u, {}).values())
        if all(abs(x) < 0.5 for x in [op, iw, tin, tout, cn, ow, ain, aout, rr, cl]): continue
        print(f"{u:<14} {fmt(op):>10} {fmt(iw):>10} {fmt(tin):>9} {fmt(tout):>9} "
              f"{fmt(cn):>8} {fmt(ow):>10} {fmt(ain):>9} {fmt(aout):>9} {fmt(rr):>9} {fmt(cl):>12}")
    print()
    print(f"Sale classification: {out_class}")

    write_html_report(
        opening_u, inward_u, inferred_in_u, txfr_in_u, txfr_out_u, cn_outward_u,
        outward_u, adj_in_u, adj_out_u, rearr_u, closing_u,
        sales_count=len(sales), cn_excel_count=len(excel_cns), cn_pdf_count=len(pdf_cns),
        out_classification=out_class, inward_unmapped=inward_unmapped,
        top_unmapped_out=out_unmap,
        sales_details=sales_details, cn_details=cn_details,
        intercompany_lines=intercompany_lines,
    )
    print(f"\nHTML report: {OUT_HTML}")


if __name__ == "__main__":
    asyncio.run(main())
