"""Closing-stock v4 — per-UNIT tracking (W-202, A-185, F-53, Cold Storage, A-68, Rishi, Savla).

Extends v3 by replacing the two-entity (CFPL/CDPL) dimension with finer-grained
warehouse units. Units roll up into entities for the rollup view.

UNIT ALLOCATION RULES
─────────────────────
Opening (CFPL + CDPL sheets, per-warehouse columns):
   W-202 Store, W-202 Lower, W-202 Upper, W-202 1st, W-202 2nd, W-202 Barline,
     W-202 Terrace, W-202 Off-Grade   → W-202
   A-185, A-185 Off-Grade              → A-185
   A-185 Cold, Supreme Cold            → COLD STORAGE
   APMC-F53                            → F-53
   Rishi                               → RISHI
   Savla                               → SAVLA

Inward (DB warehouse column on *_transactions_v2):
   W202 → W-202     A185 → A-185     F53 → F-53     A68 → A-68
   Rishi → RISHI    Savla D-39/D-514 → SAVLA       Cold* → COLD STORAGE

Outward (Sales Register voucher type + entity):
   CFPL + HO Sales        → W-202
   CFPL + APMC Sale       → F-53
   CDPL + SALES GST       → A-185
   Any unrecognised       → OTHER

CN inward:
   From PDF extract CSV (_pdf_units_extracted.csv) — uses per-row 'unit' field
   with override: source_pdf=CDPL_CN forces unit=A-185 (parser bug-fix)
   Excel CN rows fall back to entity-based unit (CFPL→W-202, CDPL→A-185)

The HTML report has a toggle: Consolidated / CFPL / CDPL / W-202 / A-185 / F-53 / COLD STORAGE / OTHER.
"""
import asyncio
import csv
import os
import re
import socket
import ssl
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import asyncpg
import openpyxl
from dotenv import load_dotenv

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

load_dotenv()
DB_URL = os.environ["DATABASE_URL"]

XLSX_DIR = Path(r"C:\Users\cando\Downloads\Inventory calc")
OPENING_XLSX = XLSX_DIR / "Candor Physical Stock Compilation 31-03-2026.xlsx"
SALES_XLSX_APR = XLSX_DIR / "Sales Register 30th April 2026.xlsx"
SALES_XLSX_MAY = XLSX_DIR / "Sales Register 12th May 2026.xlsx"
CN_PDF_CSV = Path(__file__).parent / "_cn_extracted.csv"
PDF_UNITS_CSV = Path(__file__).parent / "_pdf_units_extracted.csv"

WINDOW_START = date(2026, 4, 1)

OUT_HTML = Path(__file__).parent / "_closing_v4_report.html"
OUT_TXT = Path(__file__).parent / "_closing_v4_report.txt"

# ──────────────────────────────────────────────────────────────────────────
# Unit definitions
# ──────────────────────────────────────────────────────────────────────────

UNITS = ["W-202", "A-185", "F-53", "COLD STORAGE", "A-68", "RISHI", "SAVLA", "OTHER"]
ENTITY_OF_UNIT = {
    "W-202": "CFPL", "F-53": "CFPL", "A-68": "CFPL", "RISHI": "CFPL",
    "A-185": "CDPL", "COLD STORAGE": "CDPL", "SAVLA": "CDPL",
    "OTHER": "CFPL",  # default
}

# Opening (CFPL/CDPL sheet) column index → unit
# Columns in CFPL sheet (0-indexed from data row):
#   0=SrNo 1=ITEM 2=FG/RM 3=GROUP 4=SubGroup 5=PackSize
#   per-warehouse (Qty cnt, Qty kg) starting at 6:
#   6/7=W-202 Store          8/9=W-202 Lower         10/11=W-202 Upper
#   12/13=W-202 1st Floor    14/15=W-202 2nd Floor   16/17=W-202 Barline
#   18/19=W-202 Terrace      20/21=W-202 Off-Grade   22/23=A-185
#   24/25=A-185 Cold         26/27=A-185 Off-Grade   28/29=APMC-F53
#   30/31=Rishi              32/33=Savla             34/35=Supreme Cold
#   36/37=Total
CFPL_LOC_TO_UNIT = {
    7: "W-202", 9: "W-202", 11: "W-202", 13: "W-202", 15: "W-202",
    17: "W-202", 19: "W-202", 21: "W-202",
    23: "A-185", 27: "A-185",
    25: "COLD STORAGE", 35: "COLD STORAGE",
    29: "F-53",
    31: "RISHI",
    33: "SAVLA",
}
# CDPL sheet has same column layout
CDPL_LOC_TO_UNIT = CFPL_LOC_TO_UNIT


def warehouse_to_unit(wh: str | None) -> str:
    if not wh:
        return "OTHER"
    w = str(wh).strip().upper().replace(" ", "")
    if w.startswith("W202"): return "W-202"
    if w.startswith("A185"): return "A-185"
    if w.startswith("F53") or w == "F-53": return "F-53"
    if w.startswith("A68"): return "A-68"
    if w.startswith("A101"): return "OTHER"  # very few
    if "COLD" in w: return "COLD STORAGE"
    if "RISHI" in w: return "RISHI"
    if "SAVLA" in w: return "SAVLA"
    if "PAWANE" in w: return "SAVLA"  # Pawane is Savla-area cold storage
    return "OTHER"


def voucher_to_unit(entity: str, voucher_type: str | None) -> str:
    """Sales register row → unit."""
    e = (entity or "").upper()
    v = (voucher_type or "").upper().strip()
    if e == "CFPL":
        if "APMC" in v: return "F-53"
        return "W-202"           # HO Sales + others default to W-202
    if e == "CDPL":
        return "A-185"
    return "OTHER"


# ──────────────────────────────────────────────────────────────────────────
# DB connect with DNS fallback (local resolver sometimes fails on this RDS hostname)
# ──────────────────────────────────────────────────────────────────────────

async def connect_db():
    try:
        return await asyncpg.create_pool(DB_URL, min_size=1, max_size=2, timeout=15)
    except (socket.gaierror, OSError):
        ip_url = DB_URL.replace(
            "wms-postgres-db.cpis084golp7.ap-south-1.rds.amazonaws.com",
            "35.154.219.249",
        )
        ctx = ssl.create_default_context()
        ctx.check_hostname = False; ctx.verify_mode = ssl.CERT_NONE
        return await asyncpg.create_pool(ip_url, min_size=1, max_size=2, timeout=15, ssl=ctx)


# ──────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────

def norm(s) -> str:
    if s is None: return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


def norm_group(g) -> str:
    if not g: return "_UNMAPPED"
    return str(g).strip().lower()


def to_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None


def fmt(v: float) -> str:
    if abs(v) < 0.5: return "0"
    return f"{v:,.0f}"


# Keyword → group inference (last-resort)
KEYWORD_GROUP_MAP: list[tuple[str, str]] = [
    ("trail mix", "trail mix"), ("nuts & seeds mix", "trail mix"),
    ("breakfast mix", "trail mix"), ("chocolate coated", "trail mix"),
    ("bars", "bars & cereals"), ("cereal", "bars & cereals"), ("muesli", "bars & cereals"),
    ("hamper", "festive hampers"), ("gift", "festive hampers"),
    ("medjoul", "dates"), ("ajwa", "dates"), ("kimia", "dates"), ("fard", "dates"),
    ("kalmi", "dates"), ("safavi", "dates"), ("khalas", "dates"), ("khanezi", "dates"),
    ("zahidi", "dates"), ("barakah", "dates"), ("date paste", "dates"),
    ("date powder", "dates"), ("date syrup", "dates"), ("dates", "dates"),
    ("almond", "almond"), ("badam", "almond"),
    ("cashew", "cashew"), ("kaju", "cashew"),
    ("pista", "pista"), ("pistachio", "pista"),
    ("raisin", "raisin"), ("kishmish", "raisin"),
    ("walnut", "walnut"), ("akhrot", "walnut"),
    ("peanut", "peanuts"), ("groundnut", "peanuts"),
    ("makhana", "makhana"),
    ("anjeer", "anjeer"), ("fig", "anjeer"),
    ("apricot", "apricot"), ("khubani", "apricot"),
    ("cranberry", "cranberry"), ("blueberry", "blueberry"),
    ("blackcurrant", "blackcurrant"), ("blackberry", "blackberry"), ("prune", "prunes"),
    ("seed", "seeds"), ("sunflower", "seeds"), ("pumpkin", "seeds"),
    ("chia", "seeds"), ("flax", "seeds"), ("watermelon", "seeds"),
    ("salt", "salt"), ("seasoning", "seasoning"), ("flavour", "seasoning"),
    ("spice", "spices"), ("tajir", "tajir"),
    ("macadamia", "premium nuts"), ("pecan", "premium nuts"),
    ("hazelnut", "premium nuts"), ("brazil nut", "premium nuts"), ("pine nut", "premium nuts"),
    ("pouch", "packaging"), ("carton", "packaging"), ("label", "packaging"),
    ("tape", "packaging"), ("box", "packaging"),
]


def infer_group_from_name(name: str) -> str | None:
    n = norm(name)
    for kw, grp in KEYWORD_GROUP_MAP:
        if kw in n: return grp
    return None


# ──────────────────────────────────────────────────────────────────────────
# Opening — load CFPL + CDPL sheets with per-warehouse breakdown
# ──────────────────────────────────────────────────────────────────────────

def load_opening_per_unit() -> dict:
    """Returns {unit: {group: kg}}."""
    bucket = {u: defaultdict(float) for u in UNITS}
    wb = openpyxl.load_workbook(OPENING_XLSX, read_only=True, data_only=True)
    for sheet_name in ("CFPL", "CDPL"):
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row or len(row) < 36:
                continue
            group = row[3]      # GROUP column
            if not group:
                continue
            g = norm_group(group)
            for col_idx, unit in CFPL_LOC_TO_UNIT.items():
                if col_idx >= len(row): continue
                val = row[col_idx]
                if val is None: continue
                try:
                    kg = float(val)
                except (TypeError, ValueError):
                    continue
                if abs(kg) < 0.001: continue
                bucket[unit][g] += kg
    wb.close()
    return {u: dict(g) for u, g in bucket.items()}


# ──────────────────────────────────────────────────────────────────────────
# Inter-unit transfers (DB) per unit
# ──────────────────────────────────────────────────────────────────────────

def site_to_unit(site: str | None) -> str:
    if not site: return "OTHER"
    s = str(site).strip().upper().replace(" ", "")
    if s == "W202": return "W-202"
    if s == "A185": return "A-185"
    if s == "F53": return "F-53"
    if s == "A68": return "A-68"
    if s == "RISHI": return "RISHI"
    if s.startswith("SAVLA"): return "SAVLA"
    if "COLD" in s: return "COLD STORAGE"
    if s == "A101": return "OTHER"
    if s == "PAWANE": return "SAVLA"
    return "OTHER"


async def load_transfers_per_unit(conn, sku_info: dict) -> tuple[dict, dict]:
    """Inter-unit transfers redistribute stock between units. Returns (out_b, in_b)."""
    out_b = {u: defaultdict(float) for u in UNITS}
    in_b = {u: defaultdict(float) for u in UNITS}
    rows = await conn.fetch(
        """
        SELECT h.from_site, h.to_site, l.item_category, l.item_desc_raw,
               COALESCE(l.net_weight, 0) AS kg
        FROM interunit_transfers_lines l
        JOIN interunit_transfers_header h ON l.header_id = h.id
        WHERE h.stock_trf_date >= $1::date
        """,
        WINDOW_START,
    )
    for r in rows:
        kg = float(r["kg"] or 0)
        if kg <= 0: continue
        from_u = site_to_unit(r["from_site"])
        to_u = site_to_unit(r["to_site"])
        if from_u == to_u: continue  # intra-unit shuffle
        # Group from item_category, fallback all_sku
        g = norm_group(r["item_category"]) if r["item_category"] else None
        if not g or g == "_UNMAPPED":
            info = sku_info.get(norm(r["item_desc_raw"]))
            g = info["group"] if info else (infer_group_from_name(r["item_desc_raw"]) or "_UNMAPPED")
        out_b[from_u][g] += kg
        in_b[to_u][g] += kg
    return ({u: dict(g) for u, g in out_b.items()},
            {u: dict(g) for u, g in in_b.items()})


# ──────────────────────────────────────────────────────────────────────────
# Inward (DB) per unit
# ──────────────────────────────────────────────────────────────────────────

async def load_inward_per_unit(conn, sku_info: dict) -> tuple[dict, dict]:
    bucket = {u: defaultdict(float) for u in UNITS}
    unmapped: dict = defaultdict(float)
    for prefix in ("cfpl", "cdpl"):
        rows = await conn.fetch(
            f"""
            SELECT b.article_description AS sku, COALESCE(b.net_weight, 0) AS kg,
                   t.warehouse
            FROM {prefix}_boxes_v2 b
            JOIN {prefix}_transactions_v2 t ON b.transaction_no = t.transaction_no
            WHERE t.entry_date >= $1::date
              AND COALESCE(t.rtv, false) = false
              AND COALESCE(t.service, false) = false
            """,
            WINDOW_START,
        )
        for r in rows:
            kg = float(r["kg"] or 0)
            if kg <= 0: continue
            name = r["sku"]
            info = sku_info.get(norm(name))
            grp = info["group"] if info else (infer_group_from_name(name) or "_UNMAPPED")
            unit = warehouse_to_unit(r["warehouse"])
            bucket[unit][grp] += kg
            if grp == "_UNMAPPED":
                unmapped[name or ""] += kg
    return {u: dict(g) for u, g in bucket.items()}, dict(unmapped)


# ──────────────────────────────────────────────────────────────────────────
# all_sku + BOM master (same as v3)
# ──────────────────────────────────────────────────────────────────────────

async def load_sku_info(conn) -> dict:
    rows = await conn.fetch("SELECT particulars, item_type, item_group, sub_group, sale_group FROM all_sku")
    return {
        norm(r["particulars"]): {
            "type": (r["item_type"] or "").lower(),
            "group": norm_group(r["item_group"]),
        }
        for r in rows if r["particulars"]
    }


async def load_bom_master(conn) -> dict:
    rows = await conn.fetch("""
        SELECT h.bom_id, h.fg_sku_name, h.customer_name,
               l.material_sku_name, l.quantity_per_unit, l.item_type, l.line_number
        FROM bom_header h JOIN bom_line l ON l.bom_id = h.bom_id
        ORDER BY h.fg_sku_name, (h.customer_name IS NOT NULL), h.bom_id, l.line_number
    """)
    bom: dict = {}; chosen: dict = {}
    for r in rows:
        fg_key = norm(r["fg_sku_name"])
        if fg_key not in chosen:
            chosen[fg_key] = r["bom_id"]; bom[fg_key] = []
        if r["bom_id"] != chosen[fg_key]: continue
        bom[fg_key].append((
            r["material_sku_name"], float(r["quantity_per_unit"] or 0),
            (r["item_type"] or "rm").lower(),
        ))
    # Merge CDPL xlsx BOM
    cdpl_path = XLSX_DIR / "BOM 18-05 CDPL.xlsx"
    if cdpl_path.exists():
        wb = openpyxl.load_workbook(cdpl_path, read_only=True, data_only=True)
        ws = wb["BOM of Stock Item"]
        for r in ws.iter_rows(min_row=3, values_only=True):
            stock, _bn, fg_qty, rm, _gd, _typ, bom_qty = r
            if not stock or not rm or fg_qty in (None, 0) or bom_qty is None: continue
            try: per_unit = float(bom_qty) / float(fg_qty)
            except (TypeError, ValueError, ZeroDivisionError): continue
            it = "pm" if str(rm).upper().startswith("PM") else "rm"
            key = norm(stock)
            if key not in bom: bom[key] = []
            bom[key].append((rm, per_unit, it))
        wb.close()
    return bom


def build_bom_aliases(bom: dict) -> dict:
    aliases: dict = {}
    for canonical in bom:
        for v in {
            canonical,
            canonical.replace(" ", ""),
            re.sub(r"(\d)\s*(gm|g|kg)\b", r"\1\2", canonical),
            re.sub(r"(\d)(gm|g|kg)\b", r"\1 \2", canonical),
            re.sub(r"\s*\([^)]*\)", "", canonical).strip(),
        }:
            if v: aliases.setdefault(v, canonical)
    return aliases


def bom_lookup(key: str, bom: dict, aliases: dict):
    if key in bom: return bom[key]
    for variant in {
        key, key.replace(" ", ""),
        re.sub(r"(\d)\s*(gm|g|kg)\b", r"\1\2", key),
        re.sub(r"(\d)(gm|g|kg)\b", r"\1 \2", key),
        re.sub(r"\s*\([^)]*\)", "", key).strip(),
    }:
        canon = aliases.get(variant)
        if canon and canon in bom: return bom[canon]
    return None


# ──────────────────────────────────────────────────────────────────────────
# Sales register
# ──────────────────────────────────────────────────────────────────────────

def _iter_sheet(path: Path, sheet: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close(); return
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=3, values_only=True):
        yield row
    wb.close()


def load_excel_sales() -> tuple[list, list]:
    sales, cns = [], []
    for p in [SALES_XLSX_APR, SALES_XLSX_MAY]:
        for r in _iter_sheet(p, "Sales Report"):
            if not r or len(r) < 18: continue
            d = to_date(r[0])
            comp = str(r[10] or "").strip().upper()
            if comp not in ("CFPL", "CDPL"): continue
            try:
                pcs = float(r[15]) if r[15] is not None else None
                kg = float(r[17]) if r[17] is not None else None
            except (TypeError, ValueError):
                continue
            if d is None or kg is None: continue
            unit = voucher_to_unit(comp, r[9])
            sales.append({"date": d, "entity": comp, "unit": unit,
                          "article": r[4], "qty_pcs": pcs, "kg": kg})
        for r in _iter_sheet(p, "Cancel Inv"):
            if not r or len(r) < 18: continue
            d = to_date(r[0])
            comp = str(r[10] or "").strip().upper()
            if comp not in ("CFPL", "CDPL"): continue
            try:
                pcs = float(r[15]) if r[15] is not None else None
                kg = float(r[17]) if r[17] is not None else None
            except (TypeError, ValueError): continue
            if d is None or kg is None: continue
            unit = voucher_to_unit(comp, r[9])
            cns.append({"date": d, "entity": comp, "unit": unit,
                        "article": r[4], "qty_pcs": pcs, "kg": kg, "source": "Excel"})
    return sales, cns


def load_pdf_cns() -> list:
    """Read CN data from PDF extractions. Two sources:
       _pdf_units_extracted.csv (latest, with unit info) takes priority over
       _cn_extracted.csv (older, no unit)."""
    out = []
    if PDF_UNITS_CSV.exists():
        with open(PDF_UNITS_CSV, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                if row.get("doc_type") != "CN":
                    continue
                try: kg = float(row.get("qty_kg") or 0)
                except (TypeError, ValueError): continue
                if kg <= 0: continue
                entity = (row.get("entity") or "").upper()
                unit = (row.get("unit") or "").upper()
                # Bug-fix: CDPL CN PDF rows often mis-tagged W-202
                if row.get("source_pdf") == "CDPL_CN":
                    unit = "A-185"
                if unit not in UNITS:
                    unit = "W-202" if entity == "CFPL" else "A-185"
                out.append({"date": None, "entity": entity, "unit": unit,
                            "article": row.get("product"), "qty_pcs": None, "kg": kg,
                            "source": f"PDF:{row.get('source_pdf','')}"})
    elif CN_PDF_CSV.exists():
        with open(CN_PDF_CSV, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                try: kg = float(row.get("qty_kg") or 0)
                except (TypeError, ValueError): continue
                if kg <= 0: continue
                entity = (row.get("entity") or "").upper()
                unit = "W-202" if entity == "CFPL" else "A-185"
                out.append({"date": None, "entity": entity, "unit": unit,
                            "article": row.get("product"), "qty_pcs": None, "kg": kg,
                            "source": f"PDF:{row.get('source_pdf','')}"})
    return out


# ──────────────────────────────────────────────────────────────────────────
# Core: convert sales/CN lines → per-unit per-group RM bucket via BOM
# ──────────────────────────────────────────────────────────────────────────

def lines_to_unit_rm(lines: list, sku_info: dict, bom: dict, aliases: dict):
    bucket = {u: defaultdict(float) for u in UNITS}
    counts = {"direct_rm": 0, "fg_bom": 0, "fg_no_bom": 0, "inferred": 0, "unknown": 0}
    unmapped: list = []

    def add(unit: str, grp: str, kg: float):
        if unit not in bucket: unit = "OTHER"
        bucket[unit][grp] += kg

    for s in lines:
        art = s.get("article")
        kg = abs(float(s.get("kg") or 0))
        if not art or kg <= 0: continue
        unit = s.get("unit") or "OTHER"
        key = norm(art)
        info = sku_info.get(key)
        b = bom_lookup(key, bom, aliases)
        pcs = s.get("qty_pcs")
        try: pcs = float(pcs) if pcs is not None else None
        except (TypeError, ValueError): pcs = None

        if info and info["type"] == "pm":
            continue

        if b and pcs:
            for rm_name, q_per, it in b:
                if it == "pm": continue
                rm_kg = q_per * pcs
                rm_info = sku_info.get(norm(rm_name))
                rm_grp = rm_info["group"] if rm_info else (
                    infer_group_from_name(rm_name) or "_UNMAPPED"
                )
                add(unit, rm_grp, rm_kg)
            counts["fg_bom"] += 1
            continue

        if info and info["type"] == "rm":
            add(unit, info["group"], kg); counts["direct_rm"] += 1; continue
        if info and info["type"] == "fg":
            add(unit, info["group"], kg); counts["fg_no_bom"] += 1; continue
        inferred = infer_group_from_name(art)
        if inferred:
            add(unit, inferred, kg); counts["inferred"] += 1
        else:
            add(unit, "_UNMAPPED", kg); counts["unknown"] += 1
            unmapped.append((art, kg))
    return {u: dict(g) for u, g in bucket.items()}, counts, unmapped


# ──────────────────────────────────────────────────────────────────────────
# HTML report with multi-option toggle
# ──────────────────────────────────────────────────────────────────────────

def write_html_report(opening_u, inward_u, txfr_in_u, txfr_out_u, cn_u, outward_u, closing_u, *,
                       sales_count, cn_excel_count, cn_pdf_count,
                       out_classification, inward_unmapped, top_unmapped_out):
    all_groups = set()
    for buckets in (opening_u, inward_u, txfr_in_u, txfr_out_u, cn_u, outward_u, closing_u):
        for u in buckets:
            for g in buckets[u]:
                all_groups.add(g)
    groups = sorted(all_groups)

    def sum_units(bucket: dict, units: list[str], group: str) -> float:
        return sum(bucket.get(u, {}).get(group, 0) for u in units)

    def render_view(units: list[str]):
        rows = []; tot = [0.0] * 7   # op, iw, tin, tout, cn, ow, cl
        for g in groups:
            op = sum_units(opening_u, units, g)
            iw = sum_units(inward_u, units, g)
            tin = sum_units(txfr_in_u, units, g)
            tout = sum_units(txfr_out_u, units, g)
            cn = sum_units(cn_u, units, g)
            ow = sum_units(outward_u, units, g)
            cl = sum_units(closing_u, units, g)
            if all(abs(x) < 0.5 for x in [op, iw, tin, tout, cn, ow, cl]): continue
            cl_cls = "cl neg" if cl < -0.5 else "cl"
            rows.append(
                f"<tr><td class='grp'>{g}</td>"
                f"<td>{fmt(op)}</td><td>{fmt(iw)}</td>"
                f"<td>{fmt(tin)}</td><td>{fmt(tout)}</td>"
                f"<td>{fmt(cn)}</td><td>{fmt(ow)}</td>"
                f"<td class='{cl_cls}'>{fmt(cl)}</td></tr>"
            )
            for i, v in enumerate([op, iw, tin, tout, cn, ow, cl]):
                tot[i] += v
        rows.append(
            f"<tr class='tot-row'><td class='grp tot'>TOTAL</td>"
            + "".join(f"<td class='tot'>{fmt(v)}</td>" for v in tot[:6])
            + f"<td class='cl tot'>{fmt(tot[6])}</td></tr>"
        )
        return "".join(rows), tot

    # Views to render
    view_defs = [
        ("total", "Consolidated (All Units)", UNITS),
        ("cfpl", "CFPL (W-202 + F-53 + A-68 + Rishi)", ["W-202", "F-53", "A-68", "RISHI"]),
        ("cdpl", "CDPL (A-185 + Cold Storage + Savla)", ["A-185", "COLD STORAGE", "SAVLA"]),
        ("w202", "W-202", ["W-202"]),
        ("a185", "A-185", ["A-185"]),
        ("f53", "F-53 (APMC)", ["F-53"]),
        ("cold", "Cold Storage", ["COLD STORAGE"]),
        ("a68", "A-68", ["A-68"]),
        ("rishi", "Rishi", ["RISHI"]),
        ("savla", "Savla", ["SAVLA"]),
        ("other", "Other", ["OTHER"]),
    ]

    view_tables = {}
    view_totals = {}
    for vid, _label, units in view_defs:
        table, tot = render_view(units)
        view_tables[vid] = table
        view_totals[vid] = tot

    tot = view_totals["total"]

    # Render the toggle buttons + view divs
    buttons = "".join(
        f"<button class=\"{'active' if vid == 'total' else ''}\" "
        f"data-view='{vid}' onclick=\"switchView('{vid}')\">{label}</button>"
        for vid, label, _ in view_defs
    )
    views_html = "".join(
        f"<div class=\"view {'active' if vid == 'total' else ''}\" id='view-{vid}'>"
        f"<table><thead><tr>"
        f"<th>Group</th><th>Opening</th><th>+ Inward</th>"
        f"<th>+ Xfr-In</th><th>− Xfr-Out</th>"
        f"<th>+ CN (as RM)</th><th>− Outward (as RM)</th><th>= Closing</th>"
        f"</tr></thead><tbody>{view_tables[vid]}</tbody></table></div>"
        for vid, _label, _ in view_defs
    )
    js_summaries = ", ".join(
        f"{vid}: {{closing:'{fmt(view_totals[vid][6])}', opening:'{fmt(view_totals[vid][0])}',"
        f"inward:'{fmt(view_totals[vid][1])}', outward:'{fmt(view_totals[vid][5])}'}}"
        for vid, _, _ in view_defs
    )
    js_hints = ", ".join(
        f"{vid}: '{label}'"
        for vid, label, _ in view_defs
    )

    unmapped_html = ""
    if top_unmapped_out:
        items = sorted(top_unmapped_out, key=lambda x: -x[1])[:12]
        rows = "".join(f"<tr><td>{a}</td><td class='num'>{fmt(v)}</td></tr>" for a, v in items)
        unmapped_html = (
            f"<section class='card'><h3>Top unmapped sale SKUs</h3>"
            f"<table class='small'><thead><tr><th>Article</th><th>Sale KG</th></tr></thead>"
            f"<tbody>{rows}</tbody></table></section>"
        )

    html = f"""<!doctype html>
<html lang='en'><head><meta charset='utf-8'>
<title>Candor Foods — Closing Stock v4 — Per-Unit (Apr 1 → May 12, 2026)</title>
<style>
  :root {{ --bg:#f7f7f5; --card:#fff; --border:#d9d9d6; --accent:#16513f;
            --neg:#b6260d; --tot-bg:#eef3f0; --muted:#666; }}
  body {{ font-family:-apple-system,"Segoe UI",sans-serif; background:var(--bg);
          margin:0; padding:24px; color:#222; }}
  h1 {{ margin:0 0 4px; font-size:22px; color:var(--accent); }}
  .sub {{ color:var(--muted); margin-bottom:24px; font-size:13px; }}
  .card {{ background:var(--card); border:1px solid var(--border); border-radius:8px;
            padding:18px 22px; margin-bottom:18px; }}
  .summary {{ display:grid; grid-template-columns:repeat(4,1fr); gap:14px; }}
  .summary .lbl {{ color:var(--muted); font-size:11px; text-transform:uppercase;
                    letter-spacing:0.05em; }}
  .summary .val {{ font-size:24px; font-weight:600; color:var(--accent); }}
  .summary .val.neg {{ color:var(--neg); }}
  h2 {{ margin:0 0 12px; font-size:16px; color:var(--accent); }}
  h3 {{ margin:0 0 10px; font-size:14px; color:var(--accent); }}
  table {{ border-collapse:collapse; width:100%; font-size:13px; font-variant-numeric:tabular-nums; }}
  th,td {{ padding:7px 10px; text-align:right; border-bottom:1px solid #f0f0ee; }}
  th {{ background:#f4f5f1; font-weight:600; color:#444; font-size:11px;
        text-transform:uppercase; letter-spacing:0.04em; }}
  td.grp, th:first-child {{ text-align:left; }}
  td.grp {{ font-weight:500; }}
  td.cl {{ font-weight:600; color:var(--accent); }}
  td.cl.neg, td.neg {{ color:var(--neg); }}
  td.num {{ text-align:right; }}
  tr.tot-row td {{ background:var(--tot-bg); border-top:2px solid #ccc; font-weight:600; }}
  td.tot {{ font-weight:600; }}
  table.small {{ font-size:12px; }}
  table.small td:first-child {{ text-align:left; }}
  /* Multi-option toggle (pill chips) */
  .toggle {{ display:flex; flex-wrap:wrap; gap:6px; }}
  .toggle button {{ background:#eef0eb; border:1px solid transparent;
                     padding:6px 12px; font-size:12px; color:#555;
                     border-radius:14px; cursor:pointer; font-family:inherit;
                     font-weight:500; transition:all 0.15s; }}
  .toggle button:hover {{ color:var(--accent); border-color:var(--accent); }}
  .toggle button.active {{ background:var(--accent); color:#fff; border-color:var(--accent); }}
  .toggle-row {{ display:flex; align-items:start; gap:18px;
                  margin-bottom:14px; flex-wrap:wrap; }}
  .toggle-row h2 {{ margin:0; flex-shrink:0; }}
  .view-hint {{ font-size:12px; color:var(--muted); margin-bottom:10px;
                  background:#f8f8f5; padding:6px 10px; border-radius:4px;
                  border-left:3px solid var(--accent); }}
  .view {{ display:none; }}
  .view.active {{ display:block; }}
  .formula {{ font-family:ui-monospace,Consolas,monospace; background:#f0f1ed;
              padding:12px 14px; border-radius:6px; font-size:12.5px; line-height:1.6; }}
  .formula .lbl {{ color:var(--accent); font-weight:600; }}
  .pill {{ display:inline-block; background:#e2eae5; color:var(--accent);
            padding:2px 8px; border-radius:10px; font-size:11px; margin-right:4px; }}
  footer {{ color:var(--muted); font-size:11px; margin-top:24px; text-align:center; }}
</style></head><body>

<h1>Candor Foods — Closing Stock Report (v4 — Per-Unit)</h1>
<div class='sub'>
  Window <b>2026-04-01 → 2026-05-12</b> &nbsp;·&nbsp; All quantities in KG &nbsp;·&nbsp;
  ALL FG sales converted to RM groups via BOM &nbsp;·&nbsp;
  Toggle below to view per unit
</div>

<div class='card summary' id='summary'>
  <div><div class='lbl'>Closing Stock</div><div class='val' id='sum-closing'>{fmt(tot[6])} kg</div></div>
  <div><div class='lbl'>Opening (31-Mar)</div><div class='val' id='sum-opening'>{fmt(tot[0])} kg</div></div>
  <div><div class='lbl'>Inward (Apr 1+)</div><div class='val' id='sum-inward'>{fmt(tot[1])} kg</div></div>
  <div><div class='lbl'>Outward (Apr 1+)</div><div class='val' id='sum-outward'>{fmt(tot[5])} kg</div></div>
</div>

<div class='card'>
  <h2>Formula &amp; Unit allocation rules</h2>
  <div class='formula'>
    <span class='lbl'>Closing[group, unit]</span> &nbsp;=&nbsp; Opening + Inward + Xfr-In − Xfr-Out + CN-as-RM − Outward-as-RM
    <br><br>
    <span style='color:#666'>
      <b>Opening</b>: from CFPL/CDPL sheets — per-warehouse columns mapped to units<br>
      <b>Inward</b>: DB warehouse column → unit (W202/A185/F53/A68/Cold/Rishi/Savla)<br>
      <b>Xfr-In/Out</b>: <code>interunit_transfers</code> table — from_site/to_site mapped to units<br>
      <b>Outward</b>: voucher-type rule (CFPL+HO Sales → W-202; CFPL+APMC Sale → F-53; CDPL → A-185)<br>
      <b>CN</b>: PDF-extracted unit per credit-note row (CFPL_CN→W-202, CDPL_CN→A-185)
    </span>
  </div>
</div>

<div class='card'>
  <div class='toggle-row'>
    <h2>Closing Stock per RM Group</h2>
    <div class='toggle'>{buttons}</div>
  </div>
  <div class='view-hint' id='view-hint'>Showing Consolidated (All Units).</div>
  {views_html}
</div>

<div class='card'>
  <h2>Sale-line classification</h2>
  <p>
    <span class='pill'>Direct RM sales: {out_classification.get('direct_rm', 0)}</span>
    <span class='pill'>FG with BOM: {out_classification.get('fg_bom', 0)}</span>
    <span class='pill'>FG without BOM: {out_classification.get('fg_no_bom', 0)}</span>
    <span class='pill'>Keyword-inferred: {out_classification.get('inferred', 0)}</span>
    <span class='pill'>Unknown: {out_classification.get('unknown', 0)}</span>
  </p>
  <p style='font-size:12px;color:var(--muted)'>
    Sales lines: {sales_count:,} &nbsp;·&nbsp;
    CN from Excel: {cn_excel_count:,} &nbsp;·&nbsp;
    CN from PDFs: {cn_pdf_count:,}
  </p>
</div>

{unmapped_html}

<script>
  const summaries = {{ {js_summaries} }};
  const hints     = {{ {js_hints} }};
  function switchView(view) {{
    document.querySelectorAll('.toggle button').forEach(b => {{
      b.classList.toggle('active', b.dataset.view === view);
    }});
    document.querySelectorAll('.view').forEach(v => {{
      v.classList.toggle('active', v.id === 'view-' + view);
    }});
    const s = summaries[view];
    document.getElementById('sum-closing').textContent = s.closing + ' kg';
    document.getElementById('sum-opening').textContent = s.opening + ' kg';
    document.getElementById('sum-inward').textContent  = s.inward  + ' kg';
    document.getElementById('sum-outward').textContent = s.outward + ' kg';
    document.getElementById('view-hint').textContent = 'Showing ' + hints[view] + '.';
  }}
</script>

<footer>Generated by <code>_closing_stock_calc_v4_probe.py</code></footer>
</body></html>
"""
    OUT_HTML.write_text(html, encoding="utf-8")


# ──────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────

async def main():
    print("Loading opening per unit (CFPL+CDPL sheets)...")
    opening_u = load_opening_per_unit()

    print("Loading Excel sales + CNs...")
    sales, excel_cns = load_excel_sales()

    print("Loading PDF-extracted CNs (unit-aware)...")
    pdf_cns = load_pdf_cns()
    print(f"  Sales lines: {len(sales)} | Excel CNs: {len(excel_cns)} | PDF CNs: {len(pdf_cns)}")

    pool = await connect_db()
    try:
        async with pool.acquire() as c:
            print("Loading all_sku + BOM master...")
            sku_info = await load_sku_info(c)
            bom = await load_bom_master(c)
            print(f"  all_sku rows: {len(sku_info)} | BOM FGs: {len(bom)}")

            print("Loading inward per unit...")
            inward_u, inward_unmapped = await load_inward_per_unit(c, sku_info)

            print("Loading inter-unit transfers...")
            txfr_out_u, txfr_in_u = await load_transfers_per_unit(c, sku_info)
    finally:
        await pool.close()

    aliases = build_bom_aliases(bom)
    print(f"  BOM aliases built: {len(aliases)} variant keys")

    print("Converting outward sales → per-unit RM groups via BOM ...")
    outward_u, out_class, out_unmap = lines_to_unit_rm(sales, sku_info, bom, aliases)

    print("Converting CN inward → per-unit RM groups via BOM ...")
    all_cns = excel_cns + pdf_cns
    cn_u, _, _ = lines_to_unit_rm(all_cns, sku_info, bom, aliases)

    # ─── Compute closing per (unit, group) ───
    # closing = opening + inward + transfer_in - transfer_out + cn - outward
    closing_u = {u: {} for u in UNITS}
    all_g = set()
    for bucket in (opening_u, inward_u, txfr_in_u, txfr_out_u, cn_u, outward_u):
        for u in bucket:
            for g in bucket[u]:
                all_g.add(g)
    for u in UNITS:
        for g in all_g:
            closing_u[u][g] = (
                opening_u.get(u, {}).get(g, 0)
                + inward_u.get(u, {}).get(g, 0)
                + txfr_in_u.get(u, {}).get(g, 0)
                - txfr_out_u.get(u, {}).get(g, 0)
                + cn_u.get(u, {}).get(g, 0)
                - outward_u.get(u, {}).get(g, 0)
            )

    # ─── Text output ───
    print()
    print(f"{'UNIT':<14} {'Opening':>10} {'+Inward':>10} {'+XfrIn':>9} {'-XfrOut':>9} "
          f"{'+CN':>8} {'-Outward':>10} {'= Closing':>12}")
    print("-" * 92)
    for u in UNITS:
        op = sum(opening_u.get(u, {}).values())
        iw = sum(inward_u.get(u, {}).values())
        tin = sum(txfr_in_u.get(u, {}).values())
        tout = sum(txfr_out_u.get(u, {}).values())
        cn = sum(cn_u.get(u, {}).values())
        ow = sum(outward_u.get(u, {}).values())
        cl = sum(closing_u.get(u, {}).values())
        if all(abs(x) < 0.5 for x in [op, iw, tin, tout, cn, ow, cl]):
            continue
        print(f"{u:<14} {fmt(op):>10} {fmt(iw):>10} {fmt(tin):>9} {fmt(tout):>9} "
              f"{fmt(cn):>8} {fmt(ow):>10} {fmt(cl):>12}")
    print()
    print(f"Sale-line classification: {out_class}")

    # ─── HTML output ───
    write_html_report(
        opening_u, inward_u, txfr_in_u, txfr_out_u, cn_u, outward_u, closing_u,
        sales_count=len(sales), cn_excel_count=len(excel_cns), cn_pdf_count=len(pdf_cns),
        out_classification=out_class, inward_unmapped=inward_unmapped,
        top_unmapped_out=out_unmap,
    )
    print(f"\nHTML report: {OUT_HTML}")


if __name__ == "__main__":
    asyncio.run(main())
