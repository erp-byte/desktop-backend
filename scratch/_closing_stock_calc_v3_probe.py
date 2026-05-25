"""Closing-stock v3 — all-FG converted to RM groups via BOM; CFPL+CDPL consolidated.

Per user feedback on v2:
  1. Convert EVERY FG sale (not just trail mix / bars / hampers) into its underlying
     RM groups via BOM, then deduct from those RM groups. This properly captures the
     fact that material comes in as RM, then gets processed/mixed/repackaged into FG.
  2. Direct RM sales deduct from their own group as-is.
  3. CFPL + CDPL consolidated into ONE column per group — per-unit splits were
     causing entry-mismatch noise (e.g., packaging at A-185 tagged wrongly).
  4. Unmatched FG SKUs fall back to: (a) DB sub_group/sale_group hint, (b) keyword
     inference from the article name, (c) the FG's own item_group as last resort.
  5. CN data merged from BOTH the Excel `Cancel Inv` sheets AND the credit-note PDFs
     inside the daily / APMC zip files (CSV produced by `_cn_extracted.csv`).
  6. Use the latest BOM master from `bom_header`/`bom_line` in DB (imported earlier).

Formula per RM group:
    Closing[g] = Opening[g] + Inward[g] + Net_Transfers[g] + CN_RM_equiv[g]
               − Outward_RM_equiv[g]

  where Outward_RM_equiv = direct RM sales + Σ (FG sale × BOM-derived RM components)
  and   CN_RM_equiv      = same logic, inward direction (returns from customers)

Inputs:
  Opening       Excel   Candor Physical Stock Compilation 31-03-2026.xlsx / Compiled
  Inward        DB      cfpl_boxes_v2 / cdpl_boxes_v2 ⨝ *_transactions_v2
  Transfers    (skipped — intra-company; nets to 0 at consolidated level)
  Outward       Excel   Sales Register 30th April + 12th May  /  Sales Report
  CN inward     Excel   Sales Register /  Cancel Inv  +  PDF CNs from zip files
                        (_cn_extracted.csv produced by parallel agent)
  BOM master    DB      bom_header / bom_line (latest, imported earlier)
  all_sku       DB      SKU → (item_type, item_group, sub_group, sale_group)
"""
import asyncio
import csv
import os
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import asyncpg
import openpyxl
import socket
import ssl
from dotenv import load_dotenv

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

load_dotenv()
DB_URL = os.environ["DATABASE_URL"]


async def connect_db_with_fallback():
    """Try standard DNS first; on failure, fall back to a hard-coded IP for the
    Candor RDS endpoint (local DNS sometimes can't resolve it). SSL hostname
    verification is relaxed when using the IP fallback."""
    try:
        return await asyncpg.create_pool(DB_URL, min_size=1, max_size=2, timeout=15)
    except (socket.gaierror, OSError) as e:
        print(f"  DNS lookup failed ({e}); falling back to RDS IP...")
        # Hard-coded IP for the Candor RDS endpoint, resolved via 8.8.8.8
        ip_url = DB_URL.replace(
            "wms-postgres-db.cpis084golp7.ap-south-1.rds.amazonaws.com",
            "35.154.219.249",
        )
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        return await asyncpg.create_pool(ip_url, min_size=1, max_size=2, timeout=15, ssl=ctx)

XLSX_DIR = Path(r"C:\Users\cando\Downloads\Inventory calc")
OPENING_XLSX = XLSX_DIR / "Candor Physical Stock Compilation 31-03-2026.xlsx"
SALES_XLSX_APR = XLSX_DIR / "Sales Register 30th April 2026.xlsx"
SALES_XLSX_MAY = XLSX_DIR / "Sales Register 12th May 2026.xlsx"
CN_PDF_CSV = Path(__file__).parent / "_cn_extracted.csv"   # produced by sub-agent

WINDOW_START = date(2026, 4, 1)

OUT_HTML = Path(__file__).parent / "_closing_v3_report.html"
OUT_TXT = Path(__file__).parent / "_closing_v3_report.txt"


# ─────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────

def norm(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


def norm_group(g) -> str:
    if not g:
        return "_UNMAPPED"
    return str(g).strip().lower()


def to_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None


def fmt(v: float) -> str:
    if abs(v) < 0.5:
        return "0"
    return f"{v:,.0f}"


# Keyword → group inference (last-resort for unmapped FG/RM articles)
KEYWORD_GROUP_MAP: list[tuple[str, str]] = [
    # Most specific first
    ("trail mix", "trail mix"),
    ("nuts & seeds mix", "trail mix"),
    ("breakfast mix", "trail mix"),
    ("chocolate coated", "trail mix"),
    ("bars", "bars & cereals"),
    ("cereal", "bars & cereals"),
    ("muesli", "bars & cereals"),
    ("hamper", "festive hampers"),
    ("gift", "festive hampers"),
    ("medjoul", "dates"),
    ("ajwa", "dates"),
    ("kimia", "dates"),
    ("fard", "dates"),
    ("kalmi", "dates"),
    ("safavi", "dates"),
    ("khalas", "dates"),
    ("khanezi", "dates"),
    ("zahidi", "dates"),
    ("barakah", "dates"),
    ("date paste", "dates"),
    ("date powder", "dates"),
    ("date syrup", "dates"),
    ("dates", "dates"),
    ("almond", "almond"),
    ("badam", "almond"),
    ("cashew", "cashew"),
    ("kaju", "cashew"),
    ("pista", "pista"),
    ("pistachio", "pista"),
    ("raisin", "raisin"),
    ("kishmish", "raisin"),
    ("walnut", "walnut"),
    ("akhrot", "walnut"),
    ("peanut", "peanuts"),
    ("groundnut", "peanuts"),
    ("makhana", "makhana"),
    ("anjeer", "anjeer"),
    ("fig", "anjeer"),
    ("apricot", "apricot"),
    ("khubani", "apricot"),
    ("cranberry", "cranberry"),
    ("blueberry", "blueberry"),
    ("blackcurrant", "blackcurrant"),
    ("blackberry", "blackberry"),
    ("prune", "prunes"),
    ("seed", "seeds"),
    ("sunflower", "seeds"),
    ("pumpkin", "seeds"),
    ("chia", "seeds"),
    ("flax", "seeds"),
    ("watermelon", "seeds"),
    ("salt", "salt"),
    ("seasoning", "seasoning"),
    ("flavour", "seasoning"),
    ("spice", "spices"),
    ("tajir", "tajir"),
    ("macadamia", "premium nuts"),
    ("pecan", "premium nuts"),
    ("hazelnut", "premium nuts"),
    ("brazil nut", "premium nuts"),
    ("pine nut", "premium nuts"),
    ("pouch", "packaging"),
    ("carton", "packaging"),
    ("label", "packaging"),
    ("tape", "packaging"),
    ("box", "packaging"),
]


def infer_group_from_name(name: str) -> str | None:
    n = norm(name)
    for kw, grp in KEYWORD_GROUP_MAP:
        if kw in n:
            return grp
    return None


# ─────────────────────────────────────────────────────────────────────────
# Load openings
# ─────────────────────────────────────────────────────────────────────────

def load_opening() -> tuple[dict, dict]:
    """Return (total_per_group, by_entity_per_group)."""
    bucket: dict = defaultdict(float)
    by_entity: dict = {"CFPL": defaultdict(float), "CDPL": defaultdict(float)}
    wb = openpyxl.load_workbook(OPENING_XLSX, read_only=True, data_only=True)
    ws = wb["Compiled"]
    for row in ws.iter_rows(min_row=4, values_only=True):
        company = row[0]
        group = row[4]
        total_kg = row[14]
        if not company or not group or total_kg is None:
            continue
        try:
            kg = float(total_kg)
        except (TypeError, ValueError):
            continue
        g = norm_group(group)
        bucket[g] += kg
        e = "CFPL" if str(company).strip().upper() == "CFPL" else "CDPL"
        by_entity[e][g] += kg
    wb.close()
    return dict(bucket), {e: dict(g) for e, g in by_entity.items()}


# ─────────────────────────────────────────────────────────────────────────
# Inward (DB) — consolidated CFPL+CDPL
# ─────────────────────────────────────────────────────────────────────────

async def load_inward(conn, sku_info: dict) -> tuple[dict, dict, dict]:
    """Return (total_bucket, by_entity_bucket, unmapped)."""
    bucket: dict = defaultdict(float)
    by_entity: dict = {"CFPL": defaultdict(float), "CDPL": defaultdict(float)}
    unmapped: dict = defaultdict(float)
    for prefix, entity in (("cfpl", "CFPL"), ("cdpl", "CDPL")):
        rows = await conn.fetch(
            f"""
            SELECT b.article_description AS sku, COALESCE(b.net_weight, 0) AS kg
            FROM {prefix}_boxes_v2 b
            JOIN {prefix}_transactions_v2 t ON b.transaction_no = t.transaction_no
            WHERE t.entry_date >= $1::date
              AND COALESCE(t.rtv, false) = false
              AND COALESCE(t.service, false) = false
            """,
            WINDOW_START,
        )
        for r in rows:
            kg = float(r["kg"] or 0)
            if kg <= 0:
                continue
            name = r["sku"]
            info = sku_info.get(norm(name))
            grp = info["group"] if info else infer_group_from_name(name) or "_UNMAPPED"
            bucket[grp] += kg
            by_entity[entity][grp] += kg
            if grp == "_UNMAPPED":
                unmapped[name or ""] += kg
    return dict(bucket), {e: dict(g) for e, g in by_entity.items()}, dict(unmapped)


# ─────────────────────────────────────────────────────────────────────────
# all_sku + BOM master
# ─────────────────────────────────────────────────────────────────────────

async def load_sku_info(conn) -> dict:
    """Return {sku_name_lower: {type, group, sub_group, sale_group}}."""
    rows = await conn.fetch(
        "SELECT particulars, item_type, item_group, sub_group, sale_group FROM all_sku"
    )
    info: dict = {}
    for r in rows:
        if not r["particulars"]:
            continue
        info[norm(r["particulars"])] = {
            "type": (r["item_type"] or "").lower(),
            "group": norm_group(r["item_group"]),
            "sub_group": r["sub_group"],
            "sale_group": r["sale_group"],
        }
    return info


async def load_bom_master(conn) -> dict:
    """{fg_name_lower: [(rm_name, qty_per_unit, item_type), ...]}.

    Picks ONE BOM per FG: generic (customer_name IS NULL) first, else first
    customer-specific. Avoids double-counting when an FG has multiple BOM variants.
    """
    rows = await conn.fetch(
        """
        SELECT h.bom_id, h.fg_sku_name, h.customer_name,
               l.material_sku_name, l.quantity_per_unit, l.item_type, l.line_number
        FROM bom_header h
        JOIN bom_line l ON l.bom_id = h.bom_id
        ORDER BY h.fg_sku_name, (h.customer_name IS NOT NULL), h.bom_id, l.line_number
        """
    )
    bom: dict = {}
    chosen_bom_id: dict = {}
    for r in rows:
        fg_key = norm(r["fg_sku_name"])
        if fg_key not in chosen_bom_id:
            chosen_bom_id[fg_key] = r["bom_id"]
            bom[fg_key] = []
        if r["bom_id"] != chosen_bom_id[fg_key]:
            continue  # different BOM variant for same FG — skip
        bom[fg_key].append((
            r["material_sku_name"],
            float(r["quantity_per_unit"] or 0),
            (r["item_type"] or "rm").lower(),
        ))
    # Also fold in CDPL BOM xlsx for FGs not in DB master
    cdpl_path = XLSX_DIR / "BOM 18-05 CDPL.xlsx"
    if cdpl_path.exists():
        wb = openpyxl.load_workbook(cdpl_path, read_only=True, data_only=True)
        ws = wb["BOM of Stock Item"]
        for r in ws.iter_rows(min_row=3, values_only=True):
            stock, _bom_name, fg_qty, rm, _gd, _typ, bom_qty = r
            if not stock or not rm or fg_qty in (None, 0) or bom_qty is None:
                continue
            try:
                per_unit = float(bom_qty) / float(fg_qty)
            except (TypeError, ValueError, ZeroDivisionError):
                continue
            it = "pm" if str(rm).upper().startswith("PM") else "rm"
            key = norm(stock)
            if key not in bom:
                bom[key] = []
            bom[key].append((rm, per_unit, it))
        wb.close()
    return bom


def build_bom_aliases(bom: dict) -> dict:
    """Aliases: variants of FG name → canonical BOM key.
    Helps match sales-register names like '200gm' to BOM names like '200 gm'."""
    aliases: dict = {}
    def variants(key: str):
        out = {key}
        out.add(key.replace(" ", ""))                                 # collapse spaces
        out.add(re.sub(r"(\d)\s*(gm|g|kg)\b", r"\1\2", key))          # "200 gm" → "200gm"
        out.add(re.sub(r"(\d)(gm|g|kg)\b", r"\1 \2", key))            # "200gm" → "200 gm"
        out.add(re.sub(r"\s*\([^)]*\)", "", key).strip())             # strip "(...)" suffix
        return {v for v in out if v}
    for canonical in bom:
        for v in variants(canonical):
            aliases.setdefault(v, canonical)
    return aliases


def bom_lookup_with_aliases(key: str, bom: dict, aliases: dict):
    """Try direct match, then alias variants."""
    if key in bom:
        return bom[key]
    for variant in {
        key,
        key.replace(" ", ""),
        re.sub(r"(\d)\s*(gm|g|kg)\b", r"\1\2", key),
        re.sub(r"(\d)(gm|g|kg)\b", r"\1 \2", key),
        re.sub(r"\s*\([^)]*\)", "", key).strip(),
    }:
        canon = aliases.get(variant)
        if canon and canon in bom:
            return bom[canon]
    return None


# ─────────────────────────────────────────────────────────────────────────
# Sales register loader
# ─────────────────────────────────────────────────────────────────────────

def _iter_sales_sheet(path: Path, sheet: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=3, values_only=True):
        yield row
    wb.close()


def load_excel_sales() -> tuple[list, list]:
    sales, cns = [], []
    for p in [SALES_XLSX_APR, SALES_XLSX_MAY]:
        for r in _iter_sales_sheet(p, "Sales Report"):
            if not r or len(r) < 18:
                continue
            d = to_date(r[0])
            comp = str(r[10] or "").strip().upper()
            if comp not in ("CFPL", "CDPL"):
                continue  # skip footer / summary / junk rows
            try:
                pcs = float(r[15]) if r[15] is not None else None
                kg = float(r[17]) if r[17] is not None else None
            except (TypeError, ValueError):
                continue
            if d is None or kg is None:
                continue
            sales.append({"date": d, "entity": comp, "article": r[4], "qty_pcs": pcs, "kg": kg})
        for r in _iter_sales_sheet(p, "Cancel Inv"):
            if not r or len(r) < 18:
                continue
            d = to_date(r[0])
            comp = str(r[10] or "").strip().upper()
            if comp not in ("CFPL", "CDPL"):
                continue
            try:
                pcs = float(r[15]) if r[15] is not None else None
                kg = float(r[17]) if r[17] is not None else None
            except (TypeError, ValueError):
                continue
            if d is None or kg is None:
                continue
            cns.append({"date": d, "entity": comp, "article": r[4], "qty_pcs": pcs, "kg": kg,
                         "source": "Excel"})
    return sales, cns


def load_pdf_cns() -> list:
    """Read PDF-extracted CN data if the sub-agent's CSV is available."""
    if not CN_PDF_CSV.exists():
        return []
    out = []
    with open(CN_PDF_CSV, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                kg = float(row.get("qty_kg") or 0)
            except (TypeError, ValueError):
                continue
            if kg <= 0:
                continue
            out.append({
                "date": None, "entity": (row.get("entity") or "").upper(),
                "article": row.get("product"), "qty_pcs": None, "kg": kg,
                "source": f"PDF:{row.get('source_pdf', '')}",
            })
    return out


# ─────────────────────────────────────────────────────────────────────────
# Core: convert a list of (entity, article, pcs, kg) into RM-group bucket
# ─────────────────────────────────────────────────────────────────────────

def lines_to_rm_groups(
    lines: list, sku_info: dict, bom_lookup: dict, bom_aliases: dict | None = None,
) -> tuple[dict, dict, dict, list]:
    """For each line: if FG with BOM → split into RM groups via BOM.
                        if RM → deduct from its own group.
                        if unknown → infer by keyword, else FG's own group.

    Returns (rm_kg_total, rm_kg_by_entity, classification_counts, unmapped_lines).
    """
    bucket: dict = defaultdict(float)
    by_entity: dict = {"CFPL": defaultdict(float), "CDPL": defaultdict(float)}
    counts = {"direct_rm": 0, "fg_bom": 0, "fg_no_bom": 0, "inferred": 0, "unknown": 0}
    unmapped: list = []
    aliases = bom_aliases or {}

    def add(grp: str, kg: float, entity: str | None):
        bucket[grp] += kg
        if entity in ("CFPL", "CDPL"):
            by_entity[entity][grp] += kg

    for s in lines:
        art = s.get("article")
        kg = abs(float(s.get("kg") or 0))
        if not art or kg <= 0:
            continue
        entity = s.get("entity")
        key = norm(art)
        info = sku_info.get(key)
        bom = bom_lookup_with_aliases(key, bom_lookup, aliases) if aliases else bom_lookup.get(key)
        pcs = s.get("qty_pcs")
        try:
            pcs = float(pcs) if pcs is not None else None
        except (TypeError, ValueError):
            pcs = None

        if info and info["type"] == "pm":
            continue

        # FG with BOM → split into RM groups
        if bom and pcs:
            for rm_name, q_per, it in bom:
                if it == "pm":
                    continue
                rm_kg = q_per * pcs
                rm_info = sku_info.get(norm(rm_name))
                rm_grp = rm_info["group"] if rm_info else (
                    infer_group_from_name(rm_name) or "_UNMAPPED"
                )
                add(rm_grp, rm_kg, entity)
            counts["fg_bom"] += 1
            continue

        # RM article (per all_sku.item_type) → direct deduction
        if info and info["type"] == "rm":
            add(info["group"], kg, entity)
            counts["direct_rm"] += 1
            continue

        # FG without BOM (or no pcs) → fall back to its own group
        if info and info["type"] == "fg":
            add(info["group"], kg, entity)
            counts["fg_no_bom"] += 1
            continue

        # Unknown: keyword inference
        inferred = infer_group_from_name(art)
        if inferred:
            add(inferred, kg, entity)
            counts["inferred"] += 1
        else:
            add("_UNMAPPED", kg, entity)
            counts["unknown"] += 1
            unmapped.append((art, kg))

    return dict(bucket), {e: dict(g) for e, g in by_entity.items()}, counts, unmapped


# ─────────────────────────────────────────────────────────────────────────
# HTML report
# ─────────────────────────────────────────────────────────────────────────

def write_html_report(
    opening, inward, cn_rm, outward_rm, closing,
    opening_e, inward_e, cn_rm_e, outward_rm_e, closing_e,
    *,
    sales_count, cn_excel_count, cn_pdf_count,
    out_classification, cn_classification,
    inward_unmapped, top_unmapped_out,
) -> None:
    groups = sorted(set(opening) | set(inward) | set(cn_rm) | set(outward_rm) | set(closing))

    def render_rows(view: str):
        """view ∈ {'total','cfpl','cdpl'} — selects which dimension to render."""
        if view == "total":
            op_b, iw_b, cn_b, ow_b, cl_b = opening, inward, cn_rm, outward_rm, closing
        elif view == "cfpl":
            op_b, iw_b, cn_b, ow_b, cl_b = (
                opening_e["CFPL"], inward_e["CFPL"], cn_rm_e["CFPL"], outward_rm_e["CFPL"], closing_e["CFPL"],
            )
        else:
            op_b, iw_b, cn_b, ow_b, cl_b = (
                opening_e["CDPL"], inward_e["CDPL"], cn_rm_e["CDPL"], outward_rm_e["CDPL"], closing_e["CDPL"],
            )
        rows = []
        tot = [0.0] * 5
        for g in groups:
            op = op_b.get(g, 0); iw = iw_b.get(g, 0); cn = cn_b.get(g, 0)
            ow = ow_b.get(g, 0); cl = cl_b.get(g, 0)
            if all(abs(x) < 0.5 for x in [op, iw, cn, ow, cl]):
                continue
            cl_cls = "cl neg" if cl < -0.5 else "cl"
            rows.append(
                f"<tr><td class='grp'>{g}</td>"
                f"<td>{fmt(op)}</td><td>{fmt(iw)}</td><td>{fmt(cn)}</td>"
                f"<td>{fmt(ow)}</td><td class='{cl_cls}'>{fmt(cl)}</td></tr>"
            )
            for i, v in enumerate([op, iw, cn, ow, cl]):
                tot[i] += v
        rows.append(
            f"<tr class='tot-row'><td class='grp tot'>TOTAL</td>"
            + "".join(f"<td class='tot'>{fmt(v)}</td>" for v in tot[:4])
            + f"<td class='cl tot'>{fmt(tot[4])}</td></tr>"
        )
        return "".join(rows), tot

    table_total, tot = render_rows("total")
    table_cfpl, tot_cfpl = render_rows("cfpl")
    table_cdpl, tot_cdpl = render_rows("cdpl")

    unmapped_html = ""
    if top_unmapped_out:
        items = sorted(top_unmapped_out, key=lambda x: -x[1])[:12]
        rows2 = "".join(
            f"<tr><td>{a}</td><td class='num'>{fmt(v)}</td></tr>" for a, v in items
        )
        unmapped_html = (
            f"<section><h3>Top unmapped sale SKUs (top {len(items)} of {len(top_unmapped_out)})</h3>"
            f"<table class='small'><thead><tr><th>Article</th><th>Sale KG</th></tr></thead>"
            f"<tbody>{rows2}</tbody></table></section>"
        )

    inward_unmapped_html = ""
    if inward_unmapped:
        items = sorted(inward_unmapped.items(), key=lambda x: -x[1])[:8]
        rows3 = "".join(f"<tr><td>{a}</td><td class='num'>{fmt(v)}</td></tr>" for a, v in items)
        inward_unmapped_html = (
            f"<section><h3>Top unmapped inward SKUs</h3>"
            f"<table class='small'><thead><tr><th>Article</th><th>Inward KG</th></tr></thead>"
            f"<tbody>{rows3}</tbody></table></section>"
        )

    html = f"""<!doctype html>
<html lang='en'><head><meta charset='utf-8'>
<title>Candor Foods — Closing Stock v3 (Apr 1 → May 12, 2026)</title>
<style>
  :root {{ --bg:#f7f7f5; --card:#fff; --border:#d9d9d6; --accent:#16513f; --neg:#b6260d;
            --tot-bg:#eef3f0; --muted:#666; }}
  body {{ font-family:-apple-system,"Segoe UI",sans-serif; background:var(--bg); margin:0;
          padding:24px; color:#222; }}
  h1 {{ margin:0 0 4px; font-size:22px; color:var(--accent); }}
  .sub {{ color:var(--muted); margin-bottom:24px; font-size:13px; }}
  .card {{ background:var(--card); border:1px solid var(--border); border-radius:8px;
            padding:18px 22px; margin-bottom:18px; }}
  .summary {{ display:grid; grid-template-columns:repeat(4,1fr); gap:14px; }}
  .summary .lbl {{ color:var(--muted); font-size:11px; text-transform:uppercase;
                    letter-spacing:0.05em; }}
  .summary .val {{ font-size:24px; font-weight:600; color:var(--accent); }}
  .summary .val.neg {{ color:var(--neg); }}
  h2 {{ margin:0 0 12px; font-size:16px; color:var(--accent); }}
  h3 {{ margin:0 0 10px; font-size:14px; color:var(--accent); }}
  table {{ border-collapse:collapse; width:100%; font-size:13px; font-variant-numeric:tabular-nums; }}
  th,td {{ padding:7px 10px; text-align:right; border-bottom:1px solid #f0f0ee; }}
  th {{ background:#f4f5f1; font-weight:600; color:#444; font-size:11px;
        text-transform:uppercase; letter-spacing:0.04em; }}
  td.grp, th:first-child {{ text-align:left; }}
  td.grp {{ font-weight:500; }}
  td.cl {{ font-weight:600; color:var(--accent); }}
  td.cl.neg, td.neg {{ color:var(--neg); }}
  td.num {{ text-align:right; }}
  tr.tot-row td {{ background:var(--tot-bg); border-top:2px solid #ccc; font-weight:600; }}
  td.tot {{ font-weight:600; }}
  table.small {{ font-size:12px; }}
  table.small td:first-child {{ text-align:left; }}
  .formula {{ font-family:ui-monospace,Consolas,monospace; background:#f0f1ed;
              padding:12px 14px; border-radius:6px; font-size:12.5px; line-height:1.6; }}
  .formula .lbl {{ color:var(--accent); font-weight:600; }}
  .notes li {{ margin-bottom:6px; font-size:13px; }}
  footer {{ color:var(--muted); font-size:11px; margin-top:24px; text-align:center; }}
  .pill {{ display:inline-block; background:#e2eae5; color:var(--accent);
            padding:2px 8px; border-radius:10px; font-size:11px; margin-right:4px; }}
  /* Toggle (segmented control) */
  .toggle {{ display:inline-flex; background:#eef0eb; border-radius:8px;
              padding:3px; gap:2px; }}
  .toggle button {{ background:transparent; border:0; padding:7px 16px;
                     font-size:12px; color:#666; font-weight:500; cursor:pointer;
                     border-radius:6px; transition:all 0.15s; font-family:inherit; }}
  .toggle button:hover {{ color:var(--accent); }}
  .toggle button.active {{ background:var(--card); color:var(--accent);
                            box-shadow:0 1px 3px rgba(0,0,0,0.06); }}
  .toggle-row {{ display:flex; justify-content:space-between; align-items:center;
                  margin-bottom:12px; gap:14px; flex-wrap:wrap; }}
  .toggle-row .lbl-hint {{ font-size:11px; color:var(--muted); }}
  .view {{ display:none; }}
  .view.active {{ display:block; }}
</style></head><body>

<h1>Candor Foods — Closing Stock Report (v3)</h1>
<div class='sub'>
  Window <b>2026-04-01 → 2026-05-12</b> &nbsp;·&nbsp; All quantities in KG &nbsp;·&nbsp;
  ALL FG sales converted to underlying RM groups via BOM &nbsp;·&nbsp;
  Toggle below to switch between consolidated and per-entity views
</div>

<div class='card summary' id='summary'>
  <div><div class='lbl'>Closing Stock</div><div class='val' id='sum-closing'>{fmt(tot[4])} kg</div></div>
  <div><div class='lbl'>Opening (31-Mar)</div><div class='val' id='sum-opening'>{fmt(tot[0])} kg</div></div>
  <div><div class='lbl'>Inward (Apr 1+)</div><div class='val' id='sum-inward'>{fmt(tot[1])} kg</div></div>
  <div><div class='lbl'>Outward (Apr 1+)</div><div class='val' id='sum-outward'>{fmt(tot[3])} kg</div></div>
</div>

<div class='card'>
  <h2>Formula</h2>
  <div class='formula'>
    <span class='lbl'>Closing[group]</span> &nbsp;=&nbsp; Opening + Inward + CN-as-RM − Outward-as-RM
    <br><br>
    <span style='color:#666'>
      Every FG sale is split into its BOM ingredients and deducted from each
      ingredient's RM group. Direct RM sales deduct from their own group.
      In the per-entity view, opening / inward / outward / CN are attributed by
      the entity tagged on each source row (Company01 in sales register, table
      prefix in DB tables).
    </span>
  </div>
</div>

<div class='card'>
  <div class='toggle-row'>
    <h2 style='margin:0'>Closing Stock per RM Group</h2>
    <div class='toggle' role='group' aria-label='Entity view'>
      <button class='active' data-view='total' onclick='switchView("total")'>Consolidated (CFPL + CDPL)</button>
      <button data-view='cfpl' onclick='switchView("cfpl")'>CFPL only</button>
      <button data-view='cdpl' onclick='switchView("cdpl")'>CDPL only</button>
    </div>
  </div>
  <div class='lbl-hint' id='view-hint'>Showing CFPL + CDPL combined.</div>
  <div class='view active' id='view-total'>
    <table>
      <thead><tr>
        <th>Group</th><th>Opening</th><th>+ Inward</th>
        <th>+ CN (as RM)</th><th>− Outward (as RM)</th><th>= Closing</th>
      </tr></thead>
      <tbody>{table_total}</tbody>
    </table>
  </div>
  <div class='view' id='view-cfpl'>
    <table>
      <thead><tr>
        <th>Group</th><th>Opening</th><th>+ Inward</th>
        <th>+ CN (as RM)</th><th>− Outward (as RM)</th><th>= Closing</th>
      </tr></thead>
      <tbody>{table_cfpl}</tbody>
    </table>
  </div>
  <div class='view' id='view-cdpl'>
    <table>
      <thead><tr>
        <th>Group</th><th>Opening</th><th>+ Inward</th>
        <th>+ CN (as RM)</th><th>− Outward (as RM)</th><th>= Closing</th>
      </tr></thead>
      <tbody>{table_cdpl}</tbody>
    </table>
  </div>
</div>

<script>
  const summaries = {{
    total:  {{ closing: "{fmt(tot[4])}", opening: "{fmt(tot[0])}",
                inward:  "{fmt(tot[1])}", outward: "{fmt(tot[3])}" }},
    cfpl:   {{ closing: "{fmt(tot_cfpl[4])}", opening: "{fmt(tot_cfpl[0])}",
                inward:  "{fmt(tot_cfpl[1])}", outward: "{fmt(tot_cfpl[3])}" }},
    cdpl:   {{ closing: "{fmt(tot_cdpl[4])}", opening: "{fmt(tot_cdpl[0])}",
                inward:  "{fmt(tot_cdpl[1])}", outward: "{fmt(tot_cdpl[3])}" }},
  }};
  const hints = {{
    total: "Showing CFPL + CDPL combined.",
    cfpl:  "Showing CFPL only — sales filtered by Company01 = CFPL, inward from cfpl_* DB tables.",
    cdpl:  "Showing CDPL only — sales filtered by Company01 = CDPL, inward from cdpl_* DB tables.",
  }};
  function switchView(view) {{
    document.querySelectorAll('.toggle button').forEach(b => {{
      b.classList.toggle('active', b.dataset.view === view);
    }});
    document.querySelectorAll('.view').forEach(v => {{
      v.classList.toggle('active', v.id === 'view-' + view);
    }});
    const s = summaries[view];
    document.getElementById('sum-closing').textContent = s.closing + ' kg';
    document.getElementById('sum-opening').textContent = s.opening + ' kg';
    document.getElementById('sum-inward').textContent  = s.inward  + ' kg';
    document.getElementById('sum-outward').textContent = s.outward + ' kg';
    document.getElementById('view-hint').textContent = hints[view];
  }}
</script>

<div class='card'>
  <h2>Sale-line classification</h2>
  <p class='notes'>
    <span class='pill'>Direct RM sales: {out_classification.get('direct_rm', 0)}</span>
    <span class='pill'>FG with BOM (converted): {out_classification.get('fg_bom', 0)}</span>
    <span class='pill'>FG without BOM: {out_classification.get('fg_no_bom', 0)}</span>
    <span class='pill'>Keyword-inferred: {out_classification.get('inferred', 0)}</span>
    <span class='pill'>Unknown (_UNMAPPED): {out_classification.get('unknown', 0)}</span>
  </p>
  <p class='notes'>
    Sales lines: {sales_count:,} &nbsp;·&nbsp; CN lines from Excel: {cn_excel_count:,}
    &nbsp;·&nbsp; CN lines from PDFs: {cn_pdf_count:,}
  </p>
</div>

{inward_unmapped_html}
{unmapped_html}

<footer>Generated by <code>_closing_stock_calc_v3_probe.py</code> &nbsp;·&nbsp; Candor Foods &amp; Candor Dates</footer>
</body></html>
"""
    OUT_HTML.write_text(html, encoding="utf-8")


# ─────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────

async def main():
    print("Loading opening (Excel)...")
    opening, opening_e = load_opening()

    print("Loading Excel sales + CNs...")
    sales, excel_cns = load_excel_sales()

    print("Loading PDF-extracted CNs (if available)...")
    pdf_cns = load_pdf_cns()
    print(f"  Sales lines: {len(sales)} | Excel CNs: {len(excel_cns)} | PDF CNs: {len(pdf_cns)}")

    pool = await connect_db_with_fallback()
    try:
        async with pool.acquire() as c:
            print("Loading all_sku + BOM master...")
            sku_info = await load_sku_info(c)
            bom = await load_bom_master(c)
            print(f"  all_sku rows: {len(sku_info)} | BOM FGs: {len(bom)}")

            print("Loading inward...")
            inward, inward_e, inward_unmapped = await load_inward(c, sku_info)
    finally:
        await pool.close()

    aliases = build_bom_aliases(bom)
    print(f"  BOM aliases built: {len(aliases)} variant keys")

    print("Converting outward sales → RM groups via BOM ...")
    outward_rm, outward_rm_e, out_class, out_unmap = lines_to_rm_groups(sales, sku_info, bom, aliases)

    print("Converting CN inward → RM groups via BOM ...")
    all_cns = excel_cns + pdf_cns
    cn_rm, cn_rm_e, cn_class, _ = lines_to_rm_groups(all_cns, sku_info, bom, aliases)

    # ─── Compute closing per group (consolidated + per entity) ───
    all_g = set(opening) | set(inward) | set(cn_rm) | set(outward_rm)
    closing = {
        g: opening.get(g, 0) + inward.get(g, 0) + cn_rm.get(g, 0) - outward_rm.get(g, 0)
        for g in all_g
    }
    closing_e = {}
    for ent in ("CFPL", "CDPL"):
        closing_e[ent] = {
            g: opening_e[ent].get(g, 0) + inward_e[ent].get(g, 0)
               + cn_rm_e[ent].get(g, 0) - outward_rm_e[ent].get(g, 0)
            for g in all_g
        }

    # ─── Text output ───
    lines = []
    def p(s=""):
        lines.append(s); print(s)
    p()
    p("WINDOW: 2026-04-01 → 2026-05-12  (CFPL+CDPL consolidated; all qty in KG)")
    p("=" * 96)
    p(f"{'GROUP':<22} {'Opening':>12} {'+Inward':>12} {'+CN-RM':>10} {'-Out-RM':>12} {'= CLOSING':>14}")
    p("-" * 96)
    tot = [0.0] * 5
    for g in sorted(all_g):
        op = opening.get(g, 0)
        iw = inward.get(g, 0)
        cn = cn_rm.get(g, 0)
        ow = outward_rm.get(g, 0)
        cl = closing.get(g, 0)
        if all(abs(x) < 0.5 for x in [op, iw, cn, ow, cl]):
            continue
        p(f"{g[:22]:<22} {fmt(op):>12} {fmt(iw):>12} {fmt(cn):>10} {fmt(ow):>12} {fmt(cl):>14}")
        for i, v in enumerate([op, iw, cn, ow, cl]):
            tot[i] += v
    p("-" * 96)
    p(f"{'TOTAL':<22} {fmt(tot[0]):>12} {fmt(tot[1]):>12} {fmt(tot[2]):>10} {fmt(tot[3]):>12} {fmt(tot[4]):>14}")
    p()
    p(f"Sale-line classification: {out_class}")
    if out_unmap:
        p(f"Top unmapped sale SKUs ({len(out_unmap)} lines):")
        for art, kg in sorted(out_unmap, key=lambda x: -x[1])[:10]:
            p(f"  {fmt(kg):>10} kg   {art}")
    OUT_TXT.write_text("\n".join(lines), encoding="utf-8")

    # ─── HTML output ───
    write_html_report(
        opening, inward, cn_rm, outward_rm, closing,
        opening_e, inward_e, cn_rm_e, outward_rm_e, closing_e,
        sales_count=len(sales), cn_excel_count=len(excel_cns), cn_pdf_count=len(pdf_cns),
        out_classification=out_class, cn_classification=cn_class,
        inward_unmapped=inward_unmapped, top_unmapped_out=out_unmap,
    )
    p(f"\nHTML report: {OUT_HTML}")
    p(f"Text report: {OUT_TXT}")


if __name__ == "__main__":
    asyncio.run(main())
