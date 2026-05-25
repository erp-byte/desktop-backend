"""Deeper probe: focus on conversion sheets - Sheet2 (CFPL Conv), Bars & Trail later rows,
and look for SF (semi-finished) handling clues, the [N] external link tokens, and pivot caches."""
from openpyxl import load_workbook
import os

PATHS = {
    "Hamper": r"C:\Users\cando\Downloads\Inventory calc\BOM 5 JUN CFPL working for Conv - Hamper.xlsx",
    "Conv":   r"C:\Users\cando\Downloads\Inventory calc\BOM 5 JUN CFPL working for Conv.xlsx",
    "April":  r"C:\Users\cando\Downloads\Inventory calc\BOM Item Wise CFPL April 13.xlsx",
}

out_path = r"d:\Consumption\New\Backend\_bom_working_formulas.txt"


def dump_sheet_rows(wb_v, wb_f, sname, row_ranges, label, out):
    sh_v = wb_v[sname]
    sh_f = wb_f[sname]
    out.write(f"\n## {label} - sheet {sname!r} rows={sh_f.max_row}, cols={sh_f.max_column}\n")
    # header
    header = [sh_v.cell(1, c).value for c in range(1, sh_f.max_column + 1)]
    out.write("HEADER: " + " | ".join(str(h) for h in header) + "\n")
    for rng in row_ranges:
        for r in rng:
            if r > sh_f.max_row:
                continue
            vals = [sh_v.cell(r, c).value for c in range(1, sh_f.max_column + 1)]
            fmls = [sh_f.cell(r, c).value if (isinstance(sh_f.cell(r, c).value, str) and sh_f.cell(r, c).value.startswith("=")) else "" for c in range(1, sh_f.max_column + 1)]
            out.write(f"  R{r} VAL: " + " | ".join(str(v)[:50] if v is not None else "" for v in vals) + "\n")
            if any(fmls):
                out.write(f"  R{r} FML: " + " | ".join(str(f)[:60] for f in fmls) + "\n")


def find_unique_formulas(sh_f, max_per_pattern=3):
    """Group formulas by their column letter + structural signature."""
    from collections import defaultdict
    patterns = defaultdict(list)
    for r in range(1, sh_f.max_row + 1):
        for c in range(1, sh_f.max_column + 1):
            v = sh_f.cell(r, c).value
            if isinstance(v, str) and v.startswith("="):
                # Normalise pattern: replace digits in row refs with #
                import re
                norm = re.sub(r"(?<=[A-Z])\d+", "#", v)
                key = (c, norm)
                if len(patterns[key]) < max_per_pattern:
                    patterns[key].append((sh_f.cell(r, c).coordinate, v))
    return patterns


def main():
    with open(out_path, "a", encoding="utf-8") as out:
        out.write("\n\n" + "#" * 100 + "\n")
        out.write("DEEPER PROBE: structural formula uniques per sheet + targeted row dumps\n")
        out.write("#" * 100 + "\n")

        # 1) Hamper file - Sheet2 (Conv sheet) full structural patterns and look at rows 20-30 (last data rows)
        for label, path in PATHS.items():
            out.write(f"\n========== {label} - {os.path.basename(path)} ==========\n")
            wb_f = load_workbook(path, data_only=False)
            wb_v = load_workbook(path, data_only=True)
            for sname in wb_f.sheetnames:
                sh_f = wb_f[sname]
                if sh_f.max_row < 3:
                    continue
                patterns = find_unique_formulas(sh_f)
                if not patterns:
                    continue
                out.write(f"\n  Sheet {sname!r} unique formula patterns ({len(patterns)}):\n")
                # de-dup pattern strings
                seen = set()
                for (c, norm), examples in sorted(patterns.items()):
                    if norm in seen:
                        continue
                    seen.add(norm)
                    hdr = sh_f.cell(1, c).value
                    out.write(f"    col {c} ({hdr!r}): {norm}\n")
                    for coord, fml in examples[:1]:
                        out.write(f"        e.g. {coord} = {fml}\n")

        # 2) Hamper Sheet2 - dump rows up to end (only 31 rows) and Sheet1 (small lookup)
        out.write("\n\n=== Hamper file: Sheet2 full dump (this is the CONV / final aggregation sheet) ===\n")
        wb_f = load_workbook(PATHS["Hamper"], data_only=False)
        wb_v = load_workbook(PATHS["Hamper"], data_only=True)
        dump_sheet_rows(wb_v, wb_f, "Sheet2", [range(1, 32)], "Hamper Sheet2", out)
        out.write("\n=== Hamper file: Sheet1 full dump (lookup K col + pivot M:N) ===\n")
        dump_sheet_rows(wb_v, wb_f, "Sheet1", [range(1, 48)], "Hamper Sheet1", out)
        out.write("\n=== Hamper file: Sheet4 full dump (Bulks / SF) ===\n")
        dump_sheet_rows(wb_v, wb_f, "Sheet4", [range(1, 14)], "Hamper Sheet4", out)

        # 3) Conv file: Bars & Trail middle and tail
        out.write("\n\n=== Conv file: Bars & Trail rows 1255-1265 (where formulas start) ===\n")
        wb_f = load_workbook(PATHS["Conv"], data_only=False)
        wb_v = load_workbook(PATHS["Conv"], data_only=True)
        dump_sheet_rows(wb_v, wb_f, "Bars & Trail", [range(1255, 1270)], "Conv Bars & Trail", out)
        out.write("\n=== Conv file: Sheet1 sample (cols include 'Raw Grp' + Lookup=336+477) ===\n")
        dump_sheet_rows(wb_v, wb_f, "Sheet1", [range(1, 10), range(370, 382)], "Conv Sheet1", out)
        out.write("\n=== Conv file: Sheet3 (CURATE DATE - looks like a single FG roll up) ===\n")
        dump_sheet_rows(wb_v, wb_f, "Sheet3", [range(1, 21)], "Conv Sheet3", out)
        out.write("\n=== Conv file: Sheet2 (pivot-like RM totals) ===\n")
        dump_sheet_rows(wb_v, wb_f, "Sheet2", [range(1, 76)], "Conv Sheet2", out)

        # 4) April file: BOM of Stock Item (2) sample + Sheet3 (calculation) + Stock (master)
        out.write("\n\n=== April file: BOM of Stock Item (2) rows showing semi-finished/multi-level ===\n")
        wb_f = load_workbook(PATHS["April"], data_only=False)
        wb_v = load_workbook(PATHS["April"], data_only=True)
        # Look for rows where FG Qty != 1 (semi-finished often has FG Qty=10 etc)
        sh_v = wb_v["BOM of Stock Item (2)"]
        sh_f = wb_f["BOM of Stock Item (2)"]
        out.write(f"HEADER: {[sh_v.cell(1,c).value for c in range(1, sh_f.max_column+1)]}\n")
        found = 0
        for r in range(2, sh_f.max_row + 1):
            fg_qty = sh_v.cell(r, 6).value  # FG Qty col F
            if fg_qty and fg_qty != 1 and found < 8:
                vals = [sh_v.cell(r, c).value for c in range(1, sh_f.max_column + 1)]
                out.write(f"  R{r}: " + " | ".join(str(v)[:40] if v is not None else "" for v in vals) + "\n")
                found += 1
        # And rows where Raw Material looks like a finished product
        out.write("\nRows where RM appears to be a FG (heuristic: contains 'Carnival' or 'Bulk'):\n")
        found = 0
        for r in range(2, sh_f.max_row + 1):
            raw = sh_v.cell(r, 7).value
            if raw and ("Bulk" in str(raw) or "Carnival" in str(raw)) and found < 8:
                vals = [sh_v.cell(r, c).value for c in range(1, sh_f.max_column + 1)]
                out.write(f"  R{r}: " + " | ".join(str(v)[:40] if v is not None else "" for v in vals) + "\n")
                found += 1

        out.write("\n=== April file: Sheet3 (calculation sheet) - structural columns ===\n")
        sh_v = wb_v["Sheet3"]
        sh_f = wb_f["Sheet3"]
        out.write(f"HEADER: {[sh_v.cell(1,c).value for c in range(1, sh_f.max_column+1)]}\n")
        for r in range(2, min(15, sh_f.max_row + 1)):
            vals = [sh_v.cell(r, c).value for c in range(1, sh_f.max_column + 1)]
            out.write(f"  R{r}: " + " | ".join(str(v)[:40] if v is not None else "" for v in vals) + "\n")

        out.write("\n=== April file: Stock sheet (FG inventory with FG-to-Bulk lookup) ===\n")
        sh_v = wb_v["Stock"]
        sh_f = wb_f["Stock"]
        out.write(f"HEADER row 2: {[sh_v.cell(2,c).value for c in range(1, sh_f.max_column+1)]}\n")
        for r in range(3, min(15, sh_f.max_row + 1)):
            vals = [sh_v.cell(r, c).value for c in range(1, sh_f.max_column + 1)]
            out.write(f"  R{r}: " + " | ".join(str(v)[:40] if v is not None else "" for v in vals) + "\n")

    print("Done.")


if __name__ == "__main__":
    main()
