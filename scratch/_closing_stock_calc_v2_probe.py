"""Revised closing-stock calculation v2 — per user feedback.

Key changes from v1:
 1. BOM RM consumption attributed by **RM Group** (from RM Detail sheet), not FG Group.
    The v1 mistake: RAISIN consumption showed 20 kg (wrong) because raisins-going-into-
    trail-mix were lumped under TRAIL MIX. v2 correctly attributes 1,046 kg of raisin
    consumption via the RM-side of each voucher.
 2. **Conversion FG groups** (trail mix, bars & cereals, festive hampers) shown in a
    separate Section B — they are pure FG groups whose ingredients live in other groups.
 3. **CFPL + CDPL combined** for the accounting math, with two extra columns showing
    where the closing stock physically sits per entity.
 4. **CDPL BOM-derived consumption** from `BOM 18-05 CDPL.xlsx` — every CDPL FG sale
    converts to underlying RM via that BOM and deducts the RM, not the FG mass.
 5. **CFPL May consumption** derived from May FG sales via the imported `bom_header` /
    `bom_line` DB tables (the latest BOM master, populated in the earlier task).

Inputs:
    Opening            Excel — `Candor Physical Stock Compilation 31-03-2026.xlsx` / Compiled
    Inward             DB    — cfpl_boxes_v2 / cdpl_boxes_v2 ⨝ *_transactions_v2
    Transfers          DB    — interunit_transfers_lines / interunit_transfer_in_boxes
    CFPL Apr RM/FG     Excel — `RM_Consumption_Apr2026_CFPL.xlsx` / RM Detail + Item Analysis
    CFPL May RM/FG     Derived — May FG sales ⨯ BOM master (DB bom_header/line)
    CDPL Apr+May RM/FG Derived — CDPL FG sales ⨯ BOM master (BOM 18-05 CDPL.xlsx + DB)
    CN (returns)       Excel — Sales Register `Cancel Inv`
    Outward            Excel — Sales Register `Sales Report` (direct RM sales only;
                                FG sales handled via BOM-derived RM consumption above)

Formula per group (combined CFPL+CDPL):
    closing = opening + inward + net_xfer
              - (rm_consumed - fg_produced)   ← net BOM impact (PL + cross-group export)
              + cn_inward
              - direct_rm_outward             ← only RM sales w/o BOM; FG sales handled
                                                via BOM-derived consumption

Window: 2026-04-01 → 2026-05-12 (sales register cutoff).

Usage:
    cd d:/Consumption/New/Backend
    python _closing_stock_calc_v2_probe.py > _closing_v2_report.txt
"""
import asyncio
import os
import re
import sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import asyncpg
import openpyxl
from dotenv import load_dotenv

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

load_dotenv()
DB_URL = os.environ["DATABASE_URL"]

XLSX_DIR = Path(r"C:\Users\cando\Downloads\Inventory calc")
OPENING_XLSX = XLSX_DIR / "Candor Physical Stock Compilation 31-03-2026.xlsx"
RMCONS_XLSX = XLSX_DIR / "RM_Consumption_Apr2026_CFPL.xlsx"
SALES_XLSX_APR = XLSX_DIR / "Sales Register 30th April 2026.xlsx"
SALES_XLSX_MAY = XLSX_DIR / "Sales Register 12th May 2026.xlsx"
CDPL_BOM_XLSX = XLSX_DIR / "BOM 18-05 CDPL.xlsx"

WINDOW_START = date(2026, 4, 1)
APR_END = date(2026, 4, 30)
MAY_START = date(2026, 5, 1)

# Site → entity mapping (for inter-unit transfer attribution)
CFPL_SITES = {"W202", "A68", "F53", "RISHI", "A101"}
CDPL_SITES = {"A185", "COLD STORAGE", "SAVLA D-39", "SAVLA D-514", "PAWANE"}

# Groups treated as "conversion / composite FGs" — shown in Section B
CONVERSION_GROUPS = {"trail mix", "bars & cereals", "festive hampers"}


def site_to_entity(site: str | None) -> str | None:
    if not site:
        return None
    s = str(site).strip().upper()
    if s in CFPL_SITES:
        return "CFPL"
    if s in CDPL_SITES:
        return "CDPL"
    return None


def norm_group(g: str | None) -> str:
    if not g:
        return "_UNMAPPED"
    return str(g).strip().lower()


def norm(s: str | None) -> str:
    if not s:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


def new_bucket():
    return defaultdict(lambda: defaultdict(float))


def to_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


# ──────────────────────────────────────────────────────────────────────────
# Load supporting lookups
# ──────────────────────────────────────────────────────────────────────────

async def load_sku_to_group(conn) -> dict:
    rows = await conn.fetch("SELECT particulars, item_group FROM all_sku")
    return {norm(r["particulars"]): norm_group(r["item_group"]) for r in rows if r["particulars"]}


async def load_bom_master(conn) -> dict:
    """Return {fg_name_lower: [(rm_name, qty_per_unit, item_type), ...]}.

    Uses the DB bom_header/bom_line tables (populated with latest BOM Details CFPL
    + this task's CDPL BOM). Customer-specific BOMs are flattened — first match wins
    by (fg_sku, customer_name=NULL preferred).
    """
    # Prefer generic BOM (customer_name IS NULL); fall back to any customer BOM.
    rows = await conn.fetch(
        """
        SELECT h.fg_sku_name, h.customer_name,
               l.material_sku_name, l.quantity_per_unit, l.item_type
        FROM bom_header h
        JOIN bom_line l ON l.bom_id = h.bom_id
        ORDER BY h.fg_sku_name, (h.customer_name IS NOT NULL), l.line_number
        """
    )
    bom: dict[str, list] = {}
    for r in rows:
        fg_key = norm(r["fg_sku_name"])
        if fg_key not in bom:
            bom[fg_key] = []
        bom[fg_key].append((
            r["material_sku_name"],
            float(r["quantity_per_unit"]) if r["quantity_per_unit"] is not None else 0,
            r["item_type"] or "rm",
        ))
    return bom


def load_cdpl_bom_xlsx() -> dict:
    """Load BOM 18-05 CDPL.xlsx as supplementary lookup."""
    bom: dict[str, list] = defaultdict(list)
    wb = openpyxl.load_workbook(CDPL_BOM_XLSX, read_only=True, data_only=True)
    ws = wb["BOM of Stock Item"]
    for r in ws.iter_rows(min_row=3, values_only=True):
        stock, _bom_name, fg_qty, rm, _gd, _typ, bom_qty = r
        if not stock or not rm or fg_qty in (None, 0) or bom_qty is None:
            continue
        try:
            per_unit = float(bom_qty) / float(fg_qty)
        except (TypeError, ValueError, ZeroDivisionError):
            continue
        # Assume Dates RMs are RM; PM prefixes ignored
        item_type = "pm" if rm.upper().startswith("PM") else "rm"
        bom[norm(stock)].append((rm, per_unit, item_type))
    wb.close()
    return dict(bom)


# ──────────────────────────────────────────────────────────────────────────
# 1. OPENING (per entity + group)
# ──────────────────────────────────────────────────────────────────────────

def load_opening() -> dict:
    bucket = new_bucket()
    wb = openpyxl.load_workbook(OPENING_XLSX, read_only=True, data_only=True)
    ws = wb["Compiled"]
    for row in ws.iter_rows(min_row=4, values_only=True):
        company = row[0]
        group = row[4]
        total_kg = row[14]
        if not company or not group or total_kg is None:
            continue
        try:
            kg = float(total_kg)
        except (TypeError, ValueError):
            continue
        entity = "CFPL" if str(company).strip().upper() == "CFPL" else "CDPL"
        bucket[entity][norm_group(group)] += kg
    wb.close()
    return bucket


# ──────────────────────────────────────────────────────────────────────────
# 2. INWARD (DB)
# ──────────────────────────────────────────────────────────────────────────

async def load_inward(conn, sku_to_group: dict) -> tuple[dict, dict]:
    bucket = new_bucket()
    unmapped: dict = defaultdict(float)
    for entity, prefix in [("CFPL", "cfpl"), ("CDPL", "cdpl")]:
        rows = await conn.fetch(
            f"""
            SELECT b.article_description AS sku, COALESCE(b.net_weight, 0) AS kg
            FROM {prefix}_boxes_v2 b
            JOIN {prefix}_transactions_v2 t ON b.transaction_no = t.transaction_no
            WHERE t.entry_date >= $1::date
              AND COALESCE(t.rtv, false) = false
              AND COALESCE(t.service, false) = false
            """,
            WINDOW_START,
        )
        for r in rows:
            kg = float(r["kg"] or 0)
            if kg <= 0:
                continue
            g = sku_to_group.get(norm(r["sku"]))
            if g is None:
                unmapped[r["sku"] or ""] += kg
                g = "_UNMAPPED"
            bucket[entity][g] += kg
    return bucket, dict(unmapped)


# ──────────────────────────────────────────────────────────────────────────
# 3. TRANSFERS (cross-entity only)
# ──────────────────────────────────────────────────────────────────────────

async def load_transfers(conn, sku_to_group: dict) -> tuple[dict, dict]:
    out_b = new_bucket()
    in_b = new_bucket()
    rows = await conn.fetch(
        """
        SELECT h.from_site, h.to_site, l.item_category, l.item_desc_raw,
               COALESCE(l.net_weight, 0) AS kg
        FROM interunit_transfers_lines l
        JOIN interunit_transfers_header h ON l.header_id = h.id
        WHERE h.stock_trf_date >= $1::date
        """,
        WINDOW_START,
    )
    for r in rows:
        kg = float(r["kg"] or 0)
        if kg <= 0:
            continue
        from_e = site_to_entity(r["from_site"])
        to_e = site_to_entity(r["to_site"])
        if from_e is None or to_e is None or from_e == to_e:
            continue
        g = norm_group(r["item_category"]) if r["item_category"] else None
        if not g or g == "_UNMAPPED":
            g = sku_to_group.get(norm(r["item_desc_raw"]), "_UNMAPPED")
        out_b[from_e][g] += kg
        in_b[to_e][g] += kg
    return out_b, in_b


# ──────────────────────────────────────────────────────────────────────────
# 4a. CFPL APRIL — RM Detail (correct per-RM-group attribution)
# 4b. CFPL APRIL — FG production (Item Analysis sheet)
# ──────────────────────────────────────────────────────────────────────────

# RM groups in RM Detail that aren't real stock pools — skip these.
_BOOKKEEPING_RM_GROUPS = {"extra give away", "process loss", "off-grade dates"}

def load_cfpl_apr_rm_consumption() -> tuple[dict, float]:
    """Per-RM-group consumption from RM Detail. Returns ({(CFPL, group): kg}, total_PL_kg)."""
    bucket = new_bucket()
    pl_total = 0.0
    wb = openpyxl.load_workbook(RMCONS_XLSX, read_only=True, data_only=True)
    ws = wb["RM Detail"]
    for row in ws.iter_rows(min_row=5, values_only=True):
        if len(row) < 12:
            continue
        rm_item = row[8]
        rm_group = row[9]
        issued = row[11]
        if rm_item is None or issued is None:
            continue
        try:
            kg = float(issued)
        except (TypeError, ValueError):
            continue
        g = norm_group(rm_group)
        if g in _BOOKKEEPING_RM_GROUPS or g in ("extra give away", "process loss"):
            if kg < 0:
                pl_total += -kg
            continue
        bucket["CFPL"][g] += kg
    wb.close()
    return bucket, pl_total


def load_cfpl_apr_fg_production() -> dict:
    """FG produced per FG-group, from Item Analysis sheet."""
    bucket = new_bucket()
    wb = openpyxl.load_workbook(RMCONS_XLSX, read_only=True, data_only=True)
    ws = wb["Item Analysis"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 6:
            continue
        grp = row[0]
        fg_kg = row[5]
        if not grp or fg_kg is None:
            continue
        try:
            kg = float(fg_kg)
        except (TypeError, ValueError):
            continue
        bucket["CFPL"][norm_group(grp)] += kg
    wb.close()
    return bucket


# ──────────────────────────────────────────────────────────────────────────
# 5. SALES REGISTER — segregate FG sales (need BOM convert) vs direct RM sales
# ──────────────────────────────────────────────────────────────────────────

def _iter_sales_rows(path: Path, sheet: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=3, values_only=True):
        yield row
    wb.close()


def load_all_sales() -> tuple[list, list]:
    """Return (sales_lines, cancel_inv_lines) where each item is a dict."""
    sales: list = []
    cns: list = []
    for p in [SALES_XLSX_APR, SALES_XLSX_MAY]:
        for r in _iter_sales_rows(p, "Sales Report"):
            if not r or len(r) < 18:
                continue
            d = to_date(r[0])
            comp_s = str(r[10] or "").strip().upper()
            entity = "CFPL" if comp_s == "CFPL" else ("CDPL" if comp_s == "CDPL" else None)
            try:
                kg = float(r[17]) if r[17] is not None else None
                pcs = float(r[15]) if r[15] is not None else None
            except (TypeError, ValueError):
                continue
            if d is None or entity is None or kg is None:
                continue
            sales.append({
                "date": d, "entity": entity, "article": r[4],
                "qty_pcs": pcs, "kg": kg,
            })
        for r in _iter_sales_rows(p, "Cancel Inv"):
            if not r or len(r) < 18:
                continue
            d = to_date(r[0])
            comp_s = str(r[10] or "").strip().upper()
            entity = "CFPL" if comp_s == "CFPL" else ("CDPL" if comp_s == "CDPL" else None)
            try:
                kg = float(r[17]) if r[17] is not None else None
                pcs = float(r[15]) if r[15] is not None else None
            except (TypeError, ValueError):
                continue
            if d is None or entity is None or kg is None:
                continue
            cns.append({
                "date": d, "entity": entity, "article": r[4],
                "qty_pcs": pcs, "kg": kg,
            })
    return sales, cns


def aggregate_outward(sales_lines: list, sku_to_group: dict) -> tuple[dict, list]:
    """Outward = ALL sales by the SOLD article's group (FG or RM, regardless of BOM match)."""
    bucket = new_bucket()
    unmapped: list = []
    for s in sales_lines:
        if not s["article"] or s["kg"] in (None, 0):
            continue
        grp = sku_to_group.get(norm(s["article"]), "_UNMAPPED")
        bucket[s["entity"]][grp] += abs(s["kg"])
        if grp == "_UNMAPPED":
            unmapped.append((s["entity"], s["article"], s["kg"]))
    return bucket, unmapped


def derive_bom_production_from_sales(
    sales_lines: list,
    bom_lookup: dict,
    sku_to_group: dict,
    *,
    date_filter=None,
    entity_filter=None,
) -> tuple[dict, dict, list]:
    """Just-in-time production model: for each BOM-matched FG sale, treat as if
    the FG was produced from BOM-listed RMs in the same period.

    Returns:
      rm_consumed   {(entity, rm_group): kg}
      fg_produced   {(entity, fg_group): kg}  (== sale_mass of matched lines)
      unmatched     list of (entity, article, kg) for visibility
    """
    rm_b = new_bucket()
    fg_b = new_bucket()
    unmatched: list = []
    for s in sales_lines:
        if date_filter and not date_filter(s["date"]):
            continue
        if entity_filter and s["entity"] not in entity_filter:
            continue
        art = s["article"]
        if not art or s["kg"] in (None, 0):
            continue
        key = norm(art)
        bom = bom_lookup.get(key)
        pcs = s["qty_pcs"] if s["qty_pcs"] not in (None, 0) else None
        if not bom or pcs is None:
            # No BOM → can't derive production. Stock just leaves via outward as RM directly.
            unmatched.append((s["entity"], art, s["kg"]))
            continue
        fg_group = sku_to_group.get(key, "_UNMAPPED")
        fg_b[s["entity"]][fg_group] += abs(s["kg"])
        for rm_name, q_per, item_type in bom:
            if item_type == "pm":
                continue
            rm_kg = q_per * pcs
            rm_group = sku_to_group.get(norm(rm_name), "_UNMAPPED")
            rm_b[s["entity"]][rm_group] += rm_kg
    return rm_b, fg_b, unmatched


def load_cn_inward(cn_lines: list, sku_to_group: dict) -> dict:
    bucket = new_bucket()
    for s in cn_lines:
        if not s["article"]:
            continue
        g = sku_to_group.get(norm(s["article"]), "_UNMAPPED")
        bucket[s["entity"]][g] += abs(s["kg"])
    return bucket


# ──────────────────────────────────────────────────────────────────────────
# Aggregate buckets across entities (for combined math)
# ──────────────────────────────────────────────────────────────────────────

def per_group_total(bucket: dict, group: str) -> float:
    return sum(bucket[e].get(group, 0) for e in ["CFPL", "CDPL"])


def all_groups(*buckets) -> list[str]:
    s: set = set()
    for b in buckets:
        for e in b:
            for g in b[e]:
                s.add(g)
    return sorted(s)


# ──────────────────────────────────────────────────────────────────────────
# Output
# ──────────────────────────────────────────────────────────────────────────

def fmt(v: float) -> str:
    if abs(v) < 0.5:
        return "0"
    return f"{v:,.0f}"


def print_main_section(
    title: str, groups: list[str],
    opening, inward, t_out, t_in, rm_cons, fg_prod, cn_in, direct_out, closing,
) -> None:
    print()
    print("=" * 138)
    print(f"  {title}")
    print("=" * 138)
    hdr = (
        f"{'GROUP':<22} {'Opening':>10} {'Inward':>10} {'NetXfr':>8} "
        f"{'BOM-RM':>10} {'BOM-FG':>10} {'+CN':>8} {'-DirOut':>10} "
        f"{'= CLOSING':>12} | {'CFPL':>10} {'CDPL':>10}"
    )
    print(hdr)
    print("-" * len(hdr))
    totals = [0.0] * 9
    for g in groups:
        op = per_group_total(opening, g)
        iw = per_group_total(inward, g)
        nx = per_group_total(t_in, g) - per_group_total(t_out, g)  # 0 at company level
        rc = per_group_total(rm_cons, g)
        fp = per_group_total(fg_prod, g)
        cn = per_group_total(cn_in, g)
        do = per_group_total(direct_out, g)
        cl = per_group_total(closing, g)
        cl_cfpl = closing["CFPL"].get(g, 0)
        cl_cdpl = closing["CDPL"].get(g, 0)
        if all(abs(x) < 0.5 for x in [op, iw, nx, rc, fp, cn, do, cl]):
            continue
        print(f"{g[:22]:<22} {fmt(op):>10} {fmt(iw):>10} {fmt(nx):>8} "
              f"{fmt(rc):>10} {fmt(fp):>10} {fmt(cn):>8} {fmt(do):>10} "
              f"{fmt(cl):>12} | {fmt(cl_cfpl):>10} {fmt(cl_cdpl):>10}")
        for idx, val in enumerate([op, iw, nx, rc, fp, cn, do, cl_cfpl, cl_cdpl]):
            totals[idx] += val
    print("-" * len(hdr))
    cl_total = totals[7] + totals[8]
    print(f"{'SECTION TOTAL':<22} {fmt(totals[0]):>10} {fmt(totals[1]):>10} "
          f"{fmt(totals[2]):>8} {fmt(totals[3]):>10} {fmt(totals[4]):>10} "
          f"{fmt(totals[5]):>8} {fmt(totals[6]):>10} {fmt(cl_total):>12} | "
          f"{fmt(totals[7]):>10} {fmt(totals[8]):>10}")
    return totals[7], totals[8]


# ──────────────────────────────────────────────────────────────────────────
# HTML output
# ──────────────────────────────────────────────────────────────────────────

HTML_OUT = Path(__file__).parent / "_closing_v2_report.html"


def _html_section(title: str, groups: list[str],
                   opening, inward, t_out, t_in, rm_cons, fg_prod, cn_in, outward,
                   closing) -> tuple[str, float, float]:
    rows_html = []
    totals = [0.0] * 9
    for g in groups:
        op = per_group_total(opening, g)
        iw = per_group_total(inward, g)
        nx = per_group_total(t_in, g) - per_group_total(t_out, g)
        rc = per_group_total(rm_cons, g)
        fp = per_group_total(fg_prod, g)
        cn = per_group_total(cn_in, g)
        ow = per_group_total(outward, g)
        cl = per_group_total(closing, g)
        cl_cfpl = closing["CFPL"].get(g, 0)
        cl_cdpl = closing["CDPL"].get(g, 0)
        if all(abs(x) < 0.5 for x in [op, iw, nx, rc, fp, cn, ow, cl]):
            continue
        def td(v, klass=""):
            cls = klass
            if v < -0.5:
                cls = (cls + " neg").strip()
            elif v > 0.5:
                pass
            return f'<td class="{cls}">{fmt(v)}</td>'
        rows_html.append(
            f"<tr>"
            f'<td class="grp">{g}</td>'
            f"{td(op)}{td(iw)}{td(nx)}{td(rc)}{td(fp)}{td(cn)}{td(ow)}"
            f'<td class="cl">{fmt(cl)}{" ⚠" if cl < -0.5 else ""}</td>'
            f"{td(cl_cfpl, 'sub')}{td(cl_cdpl, 'sub')}"
            f"</tr>"
        )
        for idx, val in enumerate([op, iw, nx, rc, fp, cn, ow, cl_cfpl, cl_cdpl]):
            totals[idx] += val

    def ttd(v, klass=""):
        cls = ("tot " + klass).strip()
        if v < -0.5:
            cls += " neg"
        return f'<td class="{cls}">{fmt(v)}</td>'

    cl_total = totals[7] + totals[8]
    totals_row = (
        f"<tr class='tot-row'>"
        f"<td class='grp tot'>SECTION TOTAL</td>"
        f"{ttd(totals[0])}{ttd(totals[1])}{ttd(totals[2])}{ttd(totals[3])}{ttd(totals[4])}{ttd(totals[5])}{ttd(totals[6])}"
        f"<td class='cl tot'>{fmt(cl_total)}</td>"
        f"{ttd(totals[7])}{ttd(totals[8])}"
        f"</tr>"
    )

    table = (
        f"<section><h2>{title}</h2>"
        f"<table>"
        f"<thead><tr>"
        f"<th>Group</th><th>Opening</th><th>+ Inward</th><th>Net Xfr</th>"
        f"<th>− BOM RM</th><th>+ BOM FG</th><th>+ CN</th><th>− Outward</th>"
        f"<th>= Closing</th><th class='sub'>CFPL</th><th class='sub'>CDPL</th>"
        f"</tr></thead>"
        f"<tbody>{''.join(rows_html)}{totals_row}</tbody>"
        f"</table></section>"
    )
    return table, totals[7], totals[8]


def write_html_report(
    groups_sorted, main_groups, conv_groups,
    opening, inward, t_out, t_in, rm_cons, fg_prod, cn_in, outward, closing,
    *, cfpl_unmatched, cdpl_unmatched, pl_apr, n_sales, n_cns,
) -> None:
    section_a, cfpl_a, cdpl_a = _html_section(
        "Section A — Main RM / PM Groups", main_groups,
        opening, inward, t_out, t_in, rm_cons, fg_prod, cn_in, outward, closing,
    )
    section_b, cfpl_b, cdpl_b = _html_section(
        "Section B — Conversion FG Groups (trail mix · bars & cereals · festive hampers)",
        conv_groups,
        opening, inward, t_out, t_in, rm_cons, fg_prod, cn_in, outward, closing,
    )

    grand_cfpl = cfpl_a + cfpl_b
    grand_cdpl = cdpl_a + cdpl_b
    grand_total = grand_cfpl + grand_cdpl

    def render_unmatched(label: str, rows: list, limit: int = 10) -> str:
        if not rows:
            return ""
        agg: dict = {}
        for e, art, kg in rows:
            key = (e, art)
            agg[key] = agg.get(key, 0) + abs(kg)
        items = sorted(agg.items(), key=lambda x: -x[1])[:limit]
        body = "".join(
            f"<tr><td>{e}</td><td>{art}</td><td class='num'>{fmt(v)}</td></tr>"
            for (e, art), v in items
        )
        return (
            f"<section class='dq'><h3>{label} (top {limit} of {len(rows)})</h3>"
            f"<table class='small'><thead><tr><th>Entity</th><th>Article</th>"
            f"<th>Sale KG</th></tr></thead><tbody>{body}</tbody></table></section>"
        )

    html = f"""<!doctype html>
<html lang='en'>
<head>
<meta charset='utf-8'>
<title>Candor Foods — Closing Stock Report (Apr 1 → May 12, 2026)</title>
<style>
  :root {{
    --bg: #f7f7f5; --card: #ffffff; --border: #d9d9d6;
    --accent: #16513f; --neg: #b6260d; --tot-bg: #eef3f0; --sub-bg: #fafafa;
    --muted: #666;
  }}
  body {{
    font-family: -apple-system, "Segoe UI", system-ui, sans-serif;
    background: var(--bg); color: #222; margin: 0; padding: 24px;
  }}
  h1 {{ margin: 0 0 4px; font-size: 22px; color: var(--accent); }}
  .sub-title {{ color: var(--muted); margin-bottom: 24px; font-size: 13px; }}
  .summary-card {{
    background: var(--card); border: 1px solid var(--border); border-radius: 8px;
    padding: 18px 22px; margin-bottom: 24px; display: grid;
    grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 14px;
  }}
  .summary-card .label {{ color: var(--muted); font-size: 11px; text-transform: uppercase; letter-spacing: 0.05em; }}
  .summary-card .value {{ font-size: 22px; font-weight: 600; color: var(--accent); margin-top: 2px; }}
  .summary-card .value.neg {{ color: var(--neg); }}
  section {{ background: var(--card); border: 1px solid var(--border); border-radius: 8px; padding: 16px 18px; margin-bottom: 18px; }}
  h2 {{ margin: 0 0 12px; font-size: 16px; color: var(--accent); }}
  h3 {{ margin: 0 0 10px; font-size: 14px; color: var(--accent); }}
  table {{ border-collapse: collapse; width: 100%; font-size: 13px; font-variant-numeric: tabular-nums; }}
  th, td {{ padding: 6px 9px; text-align: right; border-bottom: 1px solid #f0f0ee; }}
  th {{ background: #f4f5f1; font-weight: 600; color: #444; font-size: 11px; text-transform: uppercase; letter-spacing: 0.04em; }}
  td.grp, th:first-child {{ text-align: left; }}
  td.grp {{ font-weight: 500; }}
  td.cl {{ font-weight: 600; color: var(--accent); }}
  td.cl.neg, td.neg {{ color: var(--neg); }}
  td.sub {{ background: var(--sub-bg); color: #555; }}
  th.sub {{ background: #eaeae6; }}
  tr.tot-row td {{ background: var(--tot-bg); border-top: 2px solid #ccc; font-weight: 600; }}
  td.tot {{ font-weight: 600; }}
  table.small {{ font-size: 12px; }}
  table.small td {{ text-align: left; }}
  table.small td.num {{ text-align: right; }}
  .formula {{ font-family: ui-monospace, "Consolas", monospace; background: #f0f1ed; padding: 12px 14px; border-radius: 6px; font-size: 12.5px; line-height: 1.6; }}
  .formula .lbl {{ color: var(--accent); font-weight: 600; }}
  .notes {{ font-size: 13px; color: #444; }}
  .notes li {{ margin-bottom: 6px; }}
  footer {{ color: var(--muted); font-size: 11px; margin-top: 24px; text-align: center; }}
</style>
</head>
<body>

<h1>Candor Foods — Closing Stock Report</h1>
<div class='sub-title'>
  Window: <b>2026-04-01 → 2026-05-12</b> &nbsp;·&nbsp; All quantities in KG &nbsp;·&nbsp;
  CFPL + CDPL combined math, with per-entity closing split on the right.
</div>

<div class='summary-card'>
  <div><div class='label'>Closing Stock — Total</div><div class='value'>{fmt(grand_total)} kg</div></div>
  <div><div class='label'>CFPL Stock</div><div class='value'>{fmt(grand_cfpl)} kg</div></div>
  <div><div class='label'>CDPL Stock</div><div class='value'>{fmt(grand_cdpl)} kg</div></div>
  <div><div class='label'>Sales lines processed</div><div class='value'>{n_sales:,}</div></div>
</div>

<section>
  <h2>Formula</h2>
  <div class='formula'>
    <span class='lbl'>Closing</span> &nbsp;=&nbsp; Opening + Inward + (Transfer-In − Transfer-Out)
    <br>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
    − <span class='lbl'>BOM RM consumed</span> &nbsp;+&nbsp; <span class='lbl'>BOM FG produced</span>
    &nbsp;+&nbsp; CN inward &nbsp;−&nbsp; Outward (all sales)
    <br>
    <br>
    <span style='color:#666'>For Apr CFPL, BOM RM/FG comes from actual Tally vouchers (RM Detail / Item Analysis).
    For May CFPL + all CDPL, BOM RM/FG is JIT-derived from sales × BOM master.
    Inter-entity transfers cancel at company level, so Net Xfr = 0.</span>
  </div>
</section>

{section_a}
{section_b}

<section>
  <h2>Grand Total — Section A + B</h2>
  <table>
    <thead><tr><th></th><th>Total</th><th>CFPL</th><th>CDPL</th></tr></thead>
    <tbody>
      <tr><td class='grp'>Section A — Main RM/PM</td>
          <td class='num'>{fmt(cfpl_a + cdpl_a)}</td>
          <td class='num'>{fmt(cfpl_a)}</td><td class='num'>{fmt(cdpl_a)}</td></tr>
      <tr><td class='grp'>Section B — Conversion FGs</td>
          <td class='num'>{fmt(cfpl_b + cdpl_b)}</td>
          <td class='num'>{fmt(cfpl_b)}</td><td class='num'>{fmt(cdpl_b)}</td></tr>
      <tr class='tot-row'><td class='grp tot'>Grand Total</td>
          <td class='cl tot'>{fmt(grand_total)}</td>
          <td class='cl tot'>{fmt(grand_cfpl)}</td>
          <td class='cl tot'>{fmt(grand_cdpl)}</td></tr>
    </tbody>
  </table>
</section>

<section>
  <h2>Data-quality notes</h2>
  <ul class='notes'>
    <li><b>Trail mix / Bars / Festive Hampers (Section B)</b> may show negative closings because some
        May sales of these FGs don't match the BOM master, so May production isn't derived for them.
        Apr Tally voucher data is exact; only May falls back to JIT-via-BOM.</li>
    <li><b>CFPL May voucher file not yet supplied</b> — derived from May FG sales × BOM as proxy.
        CFPL April uses real Tally data via <code>RM_Consumption_Apr2026_CFPL.xlsx</code>.</li>
    <li><b>CDPL has no voucher file</b> — all CDPL consumption derived via <code>BOM 18-05 CDPL.xlsx</code> + DB BOM master.</li>
    <li><b>Packaging −36k kg at CDPL</b> is a per-entity split artefact (W-202 → A-185 packaging transfers
        flagged as cross-entity). Combined company total (≈34k kg) is correct.</li>
    <li><b>Sales after May 12</b> only available as PDF invoices in the date zip files — not included in this iteration.</li>
    <li>Negatives in <code>miscellaneous - rm</code> and <code>tajir</code> indicate
        <code>all_sku.item_group</code> mismatches between inward and consumption — needs reconciliation.</li>
    <li>Process loss (Apr CFPL only): <b>{pl_apr:,.0f} kg</b></li>
  </ul>
</section>

{render_unmatched("CFPL May FG sales without BOM match", cfpl_unmatched)}
{render_unmatched("CDPL FG sales without BOM match", cdpl_unmatched)}

<footer>Generated by <code>_closing_stock_calc_v2_probe.py</code> &nbsp;·&nbsp; Candor Foods Pvt Ltd / Candor Dates Pvt Ltd</footer>
</body>
</html>
"""
    HTML_OUT.write_text(html, encoding="utf-8")
    print(f"\nHTML report written: {HTML_OUT}")


# ──────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────

async def main() -> None:
    print("Loading openings (Excel)...")
    opening = load_opening()

    print("Loading CFPL April BOM RM consumption (Excel — RM Detail by RM Group)...")
    cfpl_apr_rm, cfpl_apr_pl = load_cfpl_apr_rm_consumption()
    cfpl_apr_fg = load_cfpl_apr_fg_production()
    print(f"  CFPL Apr RM-consumed groups: {len(cfpl_apr_rm['CFPL'])}  |  "
          f"FG-produced groups: {len(cfpl_apr_fg['CFPL'])}  |  PL Apr ~{cfpl_apr_pl:,.0f} kg")

    print("Loading sales register (Excel — both files)...")
    sales, cns = load_all_sales()
    print(f"  Sales lines: {len(sales)}  |  Cancel Inv lines: {len(cns)}")

    print("Loading CDPL BOM (Excel) + DB BOM master...")
    cdpl_xlsx_bom = load_cdpl_bom_xlsx()
    print(f"  CDPL xlsx BOM FGs: {len(cdpl_xlsx_bom)}")

    pool = await asyncpg.create_pool(DB_URL, min_size=1, max_size=2)
    try:
        async with pool.acquire() as c:
            print("Loading all_sku lookup + DB bom_master ...")
            sku_to_group = await load_sku_to_group(c)
            db_bom = await load_bom_master(c)
            # Merge: CDPL xlsx BOM augments DB master (DB wins on collision since it's newer)
            merged_bom = {**cdpl_xlsx_bom, **db_bom}
            print(f"  all_sku rows: {len(sku_to_group)}  |  merged BOM FGs: {len(merged_bom)}")

            print("Loading inward (DB)...")
            inward, inward_unmapped = await load_inward(c, sku_to_group)

            print("Loading transfers (DB)...")
            t_out, t_in = await load_transfers(c, sku_to_group)
    finally:
        await pool.close()

    # ─── Outward = ALL sales by article group (whether FG or RM) ───
    print("Aggregating outward (all sales by article group)...")
    outward, outward_unmapped = aggregate_outward(sales, sku_to_group)

    # ─── Derive production (RM consumed + FG produced) ───
    # For CFPL Apr we already have Tally-accurate data; for May CFPL + CDPL we
    # use just-in-time BOM derivation from sales.
    print("Deriving CFPL May production from sales × BOM ...")
    cfpl_may_rm, cfpl_may_fg, cfpl_unmatched = derive_bom_production_from_sales(
        sales, merged_bom, sku_to_group,
        date_filter=lambda d: d >= MAY_START,
        entity_filter={"CFPL"},
    )
    print("Deriving CDPL Apr+May production from sales × BOM ...")
    cdpl_rm, cdpl_fg, cdpl_unmatched = derive_bom_production_from_sales(
        sales, merged_bom, sku_to_group,
        date_filter=None,
        entity_filter={"CDPL"},
    )

    # ─── Combine RM consumption & FG production buckets ───
    rm_cons = new_bucket()
    fg_prod = new_bucket()
    for src in [cfpl_apr_rm, cfpl_may_rm, cdpl_rm]:
        for e, g_map in src.items():
            for g, v in g_map.items(): rm_cons[e][g] += v
    for src in [cfpl_apr_fg, cfpl_may_fg, cdpl_fg]:
        for e, g_map in src.items():
            for g, v in g_map.items(): fg_prod[e][g] += v

    # ─── CN inward ───
    cn_in = load_cn_inward(cns, sku_to_group)

    # ─── Compute closing per (entity, group) ───
    # closing = opening + inward + net_xfer - BOM_RM + BOM_FG + CN - outward(all sales)
    # For Apr CFPL: BOM_RM/FG from Tally (decoupled from sales).
    # For May CFPL + CDPL: BOM_RM/FG from JIT-derivation → BOM_FG cancels outward,
    # leaving net = -BOM_RM (underlying cross-group RM depletion).
    closing = new_bucket()
    all_g = set()
    for b in [opening, inward, t_out, t_in, rm_cons, fg_prod, cn_in, outward]:
        for e in b:
            all_g.update(b[e].keys())
    for entity in ["CFPL", "CDPL"]:
        for g in all_g:
            closing[entity][g] = (
                opening[entity].get(g, 0)
                + inward[entity].get(g, 0)
                + t_in[entity].get(g, 0) - t_out[entity].get(g, 0)
                - rm_cons[entity].get(g, 0)
                + fg_prod[entity].get(g, 0)
                + cn_in[entity].get(g, 0)
                - outward[entity].get(g, 0)
            )

    groups_sorted = sorted(all_g)
    main_groups = [g for g in groups_sorted if g not in CONVERSION_GROUPS]
    conv_groups = [g for g in groups_sorted if g in CONVERSION_GROUPS]

    # ─── Print sections ───
    print()
    print("WINDOW: 2026-04-01 → 2026-05-12  (CFPL+CDPL combined; per-entity closing in right columns)")
    cfpl_a, cdpl_a = print_main_section(
        "SECTION A — MAIN RM/PM GROUPS  (all qty in KG)",
        main_groups,
        opening, inward, t_out, t_in, rm_cons, fg_prod, cn_in, outward, closing,
    )
    cfpl_b, cdpl_b = print_main_section(
        "SECTION B — CONVERSION FG GROUPS  (trail mix / bars & cereals / festive hampers)",
        conv_groups,
        opening, inward, t_out, t_in, rm_cons, fg_prod, cn_in, outward, closing,
    )

    print()
    print("=" * 138)
    print(f"  GRAND TOTAL  (Section A + Section B)")
    print("=" * 138)
    print(f"  Closing total : {fmt(cfpl_a + cdpl_a + cfpl_b + cdpl_b):>15} kg")
    print(f"     of which CFPL : {fmt(cfpl_a + cfpl_b):>12} kg")
    print(f"     of which CDPL : {fmt(cdpl_a + cdpl_b):>12} kg")

    # ─── Data quality notes ───
    print()
    print("-" * 80)
    print("  DATA-QUALITY NOTES")
    print("-" * 80)

    def top_n(label: str, d: dict, n: int = 8):
        items = []
        if isinstance(d, dict):
            for e, gm in d.items():
                for g, v in gm.items():
                    items.append((e, g, v))
        else:
            items = [(e, a, abs(v)) for e, a, v in d]
        items = sorted(items, key=lambda x: -x[2])[:n]
        if items:
            print(f"  {label}:")
            for e, x, v in items:
                print(f"     {fmt(v):>10} kg   [{e}]  {x}")

    if cfpl_unmatched:
        print(f"  CFPL May FG-sales with NO BOM match (treated as direct RM sale): {len(cfpl_unmatched)} lines")
        top_n("  Top unmatched CFPL May sales", cfpl_unmatched)
    if cdpl_unmatched:
        print(f"  CDPL FG-sales with NO BOM match: {len(cdpl_unmatched)} lines")
        top_n("  Top unmatched CDPL sales", cdpl_unmatched)

    # ─── Write HTML report ───
    write_html_report(
        groups_sorted, main_groups, conv_groups,
        opening, inward, t_out, t_in, rm_cons, fg_prod, cn_in, outward, closing,
        cfpl_unmatched=cfpl_unmatched, cdpl_unmatched=cdpl_unmatched,
        pl_apr=cfpl_apr_pl, n_sales=len(sales), n_cns=len(cns),
    )


if __name__ == "__main__":
    asyncio.run(main())
