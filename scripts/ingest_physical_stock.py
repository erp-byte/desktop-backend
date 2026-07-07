"""
Ingest physical stock take Excel files into floor_inventory table.

Files:
  1. Candor Physical Stock Compilation 31-03-2026.xlsx  (RM + FG by location, CFPL & CDPL sheets)
  2. PM Stock as on 31st March.xlsx                      (PM closing balance, CFPL only)

This script REPLACES all existing floor_inventory data.
"""

import asyncio
import sys
import os
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
import asyncpg

DATABASE_URL = "postgresql://wmsadmin:Candorfoods@wms-postgres-db.cpis084golp7.ap-south-1.rds.amazonaws.com:5432/warehouse_db"

# ---------------------------------------------------------------------------
# Location mapping: Excel location name → floor_inventory.floor_location
# ---------------------------------------------------------------------------
LOCATION_MAP = {
    "w-202 store":      "rm_store",
    "w-202 lower":      "production_floor",
    "w-202 upper":      "production_floor",
    "w-202 1st floor":  "production_floor",
    "w-202 2nd floor":  "production_floor",
    "w-202 barline":    "production_floor",
    "w-202 terrace":    "production_floor",
    "w-202-off grade":  "offgrade_store",
    "w-202 off-grade":  "offgrade_store",
    "a-185":            "rm_store",
    "a-185 cold":       "cold_store",
    "a-185 off-grade":  "offgrade_store",
    "apmc-f53":         "rm_store",
    "rishi":            "rm_store",
    "savla":            "rm_store",
    "supreme cold":     "cold_store",
}


def _safe_float(val):
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _parse_physical_stock(filepath: str) -> list[dict]:
    """Parse the RM/FG physical stock compilation Excel."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    records = []

    for sheet_name in ("CFPL", "CDPL"):
        if sheet_name not in wb.sheetnames:
            continue

        entity = sheet_name.lower()
        ws = wb[sheet_name]

        # Row 2 has location names at even columns starting from 6
        row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        locations = {}  # col_index (of qty_kg) → location_name
        for col_idx, val in enumerate(row2):
            if val and str(val).strip() and col_idx >= 6:
                loc_name = str(val).strip().lower()
                if loc_name in ("total stock",):
                    continue
                # KG column is col_idx + 1
                locations[col_idx + 1] = loc_name

        # Data starts at row 4
        for row in ws.iter_rows(min_row=4, values_only=True):
            vals = list(row)
            item_name = str(vals[1]).strip() if vals[1] else ""
            if not item_name:
                continue
            # Skip total/summary rows
            if item_name.lower().startswith(("total", "grand total")):
                continue

            item_type_raw = str(vals[2]).strip().upper() if vals[2] else ""
            item_type = "rm" if item_type_raw == "RM" else "fg" if item_type_raw == "FG" else item_type_raw.lower()
            group = str(vals[3]).strip() if vals[3] else ""
            sub_group = str(vals[4]).strip() if vals[4] else ""
            pack_size = _safe_float(vals[5]) if vals[5] else 1.0

            for kg_col, loc_name in locations.items():
                qty_kg = _safe_float(vals[kg_col]) if kg_col < len(vals) else 0.0
                if qty_kg <= 0:
                    continue

                floor_loc = LOCATION_MAP.get(loc_name, "rm_store")

                records.append({
                    "sku_name": item_name,
                    "item_type": item_type,
                    "floor_location": floor_loc,
                    "quantity_kg": qty_kg,
                    "lot_number": f"STOCKTAKE-2026-03-31",
                    "entity": entity,
                })

    wb.close()
    return records


def _parse_pm_stock(filepath: str) -> list[dict]:
    """Parse the PM closing balance Excel."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb["Packing Material"]
    records = []

    # Data starts at row 18: col 0=Particulars, col 1=Quantity, col 2=Rate, col 3=Value
    for row in ws.iter_rows(min_row=18, values_only=True):
        vals = list(row)
        item_name = str(vals[0]).strip() if vals[0] else ""
        if not item_name:
            continue
        if item_name.lower().startswith(("total", "grand total")):
            continue

        qty = _safe_float(vals[1])
        if qty <= 0:
            continue

        records.append({
            "sku_name": item_name,
            "item_type": "pm",
            "floor_location": "pm_store",
            "quantity_kg": qty,  # Actually units for PM, but stored in same column
            "lot_number": f"STOCKTAKE-2026-03-31",
            "entity": "cfpl",
        })

    wb.close()
    return records


async def main():
    downloads = Path(os.path.expanduser("~")) / "Downloads"

    physical_file = downloads / "Candor Physical Stock Compilation 31-03-2026.xlsx"
    pm_file = downloads / "PM Stock as on 31st March.xlsx"

    print("Parsing physical stock...")
    rm_fg_records = _parse_physical_stock(str(physical_file))
    print(f"  RM/FG records: {len(rm_fg_records)}")

    print("Parsing PM stock...")
    pm_records = _parse_pm_stock(str(pm_file))
    print(f"  PM records: {len(pm_records)}")

    all_records = rm_fg_records + pm_records

    # Aggregate: same (sku_name, floor_location, entity, lot_number) → sum qty
    agg = {}
    for r in all_records:
        key = (r["sku_name"], r["item_type"], r["floor_location"], r["lot_number"], r["entity"])
        if key in agg:
            agg[key]["quantity_kg"] += r["quantity_kg"]
        else:
            agg[key] = r.copy()

    records = list(agg.values())
    print(f"  Total unique records after aggregation: {len(records)}")

    # Summary by type
    by_type = {}
    for r in records:
        t = r["item_type"]
        by_type[t] = by_type.get(t, 0) + 1
    for t, c in sorted(by_type.items()):
        print(f"    {t}: {c}")

    conn = await asyncpg.connect(DATABASE_URL)

    # Clear existing inventory
    await conn.execute("TRUNCATE TABLE floor_inventory CASCADE;")
    print("\nCleared floor_inventory table")

    # Insert new records
    inserted = 0
    for r in records:
        await conn.execute(
            """
            INSERT INTO floor_inventory (sku_name, item_type, floor_location, quantity_kg, lot_number, entity)
            VALUES ($1, $2, $3, $4, $5, $6)
            ON CONFLICT (sku_name, floor_location, lot_number, entity)
            DO UPDATE SET quantity_kg = EXCLUDED.quantity_kg, item_type = EXCLUDED.item_type, last_updated = NOW()
            """,
            r["sku_name"], r["item_type"], r["floor_location"],
            r["quantity_kg"], r["lot_number"], r["entity"],
        )
        inserted += 1

    total = await conn.fetchval("SELECT COUNT(*) FROM floor_inventory")
    total_kg = await conn.fetchval("SELECT COALESCE(SUM(quantity_kg), 0) FROM floor_inventory")
    print(f"\nInserted {inserted} records")
    print(f"floor_inventory now has {total} records, total qty: {total_kg:.2f}")

    # Breakdown
    rows = await conn.fetch(
        "SELECT entity, item_type, floor_location, COUNT(*), SUM(quantity_kg) "
        "FROM floor_inventory GROUP BY entity, item_type, floor_location ORDER BY 1, 2, 3"
    )
    print("\nBreakdown:")
    for r in rows:
        print(f"  {r[0]} | {r[1]:4s} | {r[2]:20s} | {r[3]:4d} items | {float(r[4]):>12.2f} kg")

    await conn.close()
    print("\nDone!")


if __name__ == "__main__":
    asyncio.run(main())
